from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import os
import json
import time
import requests  # 新增：用于微信登录
import pymysql   # 确保有这行！
import pymysql.cursors  # 新增：用于DictCursor
from datetime import datetime, timedelta
from config import Config
from utils import (
    get_db_connection, allowed_file, save_uploaded_file,
    calculate_age_from_id_card, get_gender_from_id_card,
    success_response, error_response
)
from auth import login_required, admin_required, admin_or_teacher_required, get_current_user, wechat_login, create_or_update_wechat_user
from datetime import datetime, timedelta
import json
# 在 from auth import ... 这行之后添加

from functools import wraps
from collections import defaultdict
import threading

from logger import Logger
from notifier import Notifier


# 简单的内存限流器
rate_limiter = defaultdict(list)
rate_limit_lock = threading.Lock()


def rate_limit(max_requests=100, time_window=60):
    """限流装饰器"""

    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            client_ip = request.remote_addr
            current_time = time.time()

            with rate_limit_lock:
                rate_limiter[client_ip] = [t for t in rate_limiter[client_ip]
                                           if current_time - t < time_window]

                if len(rate_limiter[client_ip]) >= max_requests:
                    return jsonify(error_response("请求过于频繁，请稍后再试", 429)), 429

                rate_limiter[client_ip].append(current_time)

            return f(*args, **kwargs)

        return wrapped

    return decorator

app = Flask(__name__)
CORS(app)

# 配置
app.config.from_object(Config)

# 确保上传目录存在
os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)


# 静态文件服务
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(Config.UPLOAD_FOLDER, filename)


# ==================== 认证相关 ====================
@app.route('/api/auth/check', methods=['GET'])
def check_auth():
    """检查登录状态"""
    user = get_current_user()
    if user:
        return jsonify(success_response(user, "已登录"))
    return jsonify(error_response("未登录", 401)), 401


@app.route('/api/auth/admin-check', methods=['GET'])
@admin_required
def check_admin():
    """检查管理员权限"""
    return jsonify(success_response(request.current_user, "管理员权限验证通过"))


# ==================== 活动相关 ====================
@app.route('/api/activities', methods=['GET'])
def get_activities():
    """获取活动列表"""
    try:
        category = request.args.get('category', '')
        search = request.args.get('search', '')
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        offset = (page - 1) * limit

        conn = get_db_connection()
        cursor = conn.cursor()

        # 构建查询条件
        where_clause = "WHERE status = 'published'"
        params = []

        if category and category != '全部':
            where_clause += " AND category = %s"
            params.append(category)

        # 新增：搜索条件
        if search:
            where_clause += " AND title LIKE %s"
            params.append(f'%{search}%')

        # 查询活动列表
        sql = f"""
        SELECT * FROM activities 
        {where_clause}
        ORDER BY is_top DESC, sort_order ASC, created_at DESC 
        LIMIT %s OFFSET %s
        """
        params.extend([limit, offset])

        cursor.execute(sql, params)
        activities = cursor.fetchall()

        # 处理每个活动
        activities_list = []
        for activity in activities:
            # 转换日期
            date_fields = ['registration_start', 'registration_end', 'activity_start', 'activity_end', 'created_at',
                           'updated_at']
            for field in date_fields:
                if field in activity and activity[field] is not None:
                    if hasattr(activity[field], 'strftime'):
                        activity[field] = activity[field].strftime('%Y-%m-%d %H:%M:%S')

            # 转换 Decimal
            from decimal import Decimal
            decimal_fields = ['base_fee', 'insurance_fee', 'transport_fee', 'meal_fee']
            for field in decimal_fields:
                if field in activity and isinstance(activity.get(field), Decimal):
                    activity[field] = float(activity[field])

            # 处理封面图片
            if activity.get('cover_images'):
                try:
                    if isinstance(activity['cover_images'], str):
                        activity['cover_images'] = json.loads(activity['cover_images'])
                except:
                    activity['cover_images'] = []
            else:
                activity['cover_images'] = []

            # 如果没有封面图片，设置默认图片
            if not activity['cover_images']:
                activity['cover_images'] = ['https://picsum.photos/750/400?random=1']

            # 为了兼容前端，添加 cover_image_url 字段（取第一张图）
            activity['cover_image_url'] = activity['cover_images'][0] if activity[
                'cover_images'] else 'https://picsum.photos/750/400?random=1'

            # 处理JSON字段
            if activity.get('additional_fees'):
                try:
                    if isinstance(activity['additional_fees'], str):
                        activity['additional_fees'] = json.loads(activity['additional_fees'])
                except:
                    activity['additional_fees'] = []
            else:
                activity['additional_fees'] = []

            activities_list.append(activity)

        cursor.close()
        conn.close()

        return jsonify(success_response(activities_list))
    except Exception as e:
        return jsonify(error_response(f"获取活动列表失败: {str(e)}")), 500


@app.route('/api/activities/carousel', methods=['GET'])
def get_carousel_activities():
    """获取轮播活动"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 查询轮播活动（只取前10个）
        cursor.execute("""
            SELECT * FROM activities 
            WHERE status = 'published' AND is_carousel = 1
            ORDER BY sort_order ASC, created_at DESC 
            LIMIT 10
        """)

        activities = cursor.fetchall()

        # 处理每个活动
        activities_list = []
        for activity in activities:
            # 转换日期
            date_fields = ['registration_start', 'registration_end', 'activity_start', 'activity_end', 'created_at',
                           'updated_at']
            for field in date_fields:
                if field in activity and activity[field] is not None:
                    if hasattr(activity[field], 'strftime'):
                        activity[field] = activity[field].strftime('%Y-%m-%d %H:%M:%S')

            # 转换 Decimal
            from decimal import Decimal
            decimal_fields = ['base_fee', 'insurance_fee', 'transport_fee', 'meal_fee']
            for field in decimal_fields:
                if field in activity and isinstance(activity.get(field), Decimal):
                    activity[field] = float(activity[field])

            # 处理封面图片
            if activity.get('cover_images'):
                try:
                    if isinstance(activity['cover_images'], str):
                        activity['cover_images'] = json.loads(activity['cover_images'])
                except:
                    activity['cover_images'] = []
            else:
                activity['cover_images'] = []

            # 如果没有封面图片，设置默认图片
            if not activity['cover_images']:
                activity['cover_images'] = ['https://picsum.photos/750/400?random=1']

            # 为了兼容前端，添加 cover_image_url 字段（取第一张图）
            activity['cover_image_url'] = activity['cover_images'][0] if activity[
                'cover_images'] else 'https://picsum.photos/750/400?random=1'

            activities_list.append(activity)

        cursor.close()
        conn.close()

        return jsonify(success_response(activities_list))
    except Exception as e:
        return jsonify(error_response(f"获取轮播活动失败: {str(e)}")), 500

@app.route('/api/activities/<int:activity_id>', methods=['GET'])
def get_activity_detail(activity_id):
    """获取活动详情"""
    print(f"\n===== 开始处理活动详情请求 =====")
    print(f"活动ID: {activity_id}")


    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        print("数据库连接成功")


        cursor.execute("SELECT * FROM activities WHERE id = %s", (activity_id,))
        activity = cursor.fetchone()
        print(f"数据库查询完成，结果类型: {type(activity)}")


        if not activity:
            print("活动不存在")
            cursor.close()
            conn.close()
            return jsonify(error_response("活动不存在", 404)), 404


        activity_dict = activity


        print(f"活动字典键: {activity_dict.keys()}")


        # ★ 修复：添加 cancel_deadline 到日期字段列表
        date_fields = ['registration_start', 'registration_end', 'activity_start', 'activity_end', 'cancel_deadline', 'created_at', 'updated_at']
        for field in date_fields:
            if field in activity_dict and activity_dict[field] is not None:
                value = activity_dict[field]
                if hasattr(value, 'strftime'):
                    activity_dict[field] = value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    activity_dict[field] = str(value)


        # 转换 Decimal 类型为 float
        from decimal import Decimal
        decimal_fields = ['base_fee', 'insurance_fee', 'transport_fee', 'meal_fee']
        for field in decimal_fields:
            if field in activity_dict and isinstance(activity_dict.get(field), Decimal):
                activity_dict[field] = float(activity_dict[field])


        # 转换整数布尔值为真正的布尔值
        if 'is_top' in activity_dict:
            activity_dict['is_top'] = bool(activity_dict['is_top'])
        if 'is_carousel' in activity_dict:
            activity_dict['is_carousel'] = bool(activity_dict['is_carousel'])
        if 'no_review_needed' in activity_dict:
            activity_dict['no_review_needed'] = bool(activity_dict['no_review_needed'])


        # 处理封面图片
        if activity_dict.get('cover_images'):
            try:
                if isinstance(activity_dict['cover_images'], str):
                    activity_dict['cover_images'] = json.loads(activity_dict['cover_images'])
            except:
                activity_dict['cover_images'] = []
        else:
            activity_dict['cover_images'] = []


        if not activity_dict['cover_images']:
            activity_dict['cover_images'] = ['https://picsum.photos/750/400?random=1']


        # 处理JSON字段
        if activity_dict.get('additional_fees'):
            try:
                if isinstance(activity_dict['additional_fees'], str):
                    activity_dict['additional_fees'] = json.loads(activity_dict['additional_fees'])
            except:
                activity_dict['additional_fees'] = []
        else:
            activity_dict['additional_fees'] = []


        cursor.close()
        conn.close()


        print("活动详情处理完成，准备返回")
        return jsonify(success_response(activity_dict))


    except Exception as e:
        print(f"\n!!!!! 发生异常 !!!!!")
        print(f"异常类型: {type(e)}")
        print(f"异常信息: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"获取活动详情失败: {str(e)}")), 500



@app.route('/api/activities/<int:activity_id>/register', methods=['POST'])
@login_required
@rate_limit(max_requests=10, time_window=60)
def register_activity(activity_id):
    """报名活动"""
    conn = None
    cursor = None
    try:
        data = request.get_json()
        user_id = request.current_user['id']

        print(f"\n===== 开始处理报名 =====")
        print(f"用户ID: {user_id}, 活动ID: {activity_id}")
        print(f"接收数据: {data}")

        conn = get_db_connection()
        cursor = conn.cursor()

        # 开始事务
        conn.begin()

        # 检查是否已经报名（加锁）
        cursor.execute(
            "SELECT id, status FROM user_activities WHERE user_id = %s AND activity_id = %s FOR UPDATE",
            (user_id, activity_id)
        )
        existing = cursor.fetchone()

        if existing:
            if existing.get('status') in ('rejected', 'cancelled'):
                # ★ 修复：审核未通过或已取消的记录，删除后允许重新报名
                cursor.execute("DELETE FROM user_activities WHERE id = %s", (existing['id'],))
                print(f"删除状态为{existing.get('status')}的报名记录 id={existing['id']}，允许重新报名")
            else:
                conn.rollback()
                print(f"用户已报名，状态: {existing.get('status')}")
                return jsonify(error_response("您已经报名了此活动")), 400


        # 检查活动（加锁防止超卖）
        cursor.execute("SELECT * FROM activities WHERE id = %s FOR UPDATE", (activity_id,))
        activity = cursor.fetchone()

        if not activity:
            conn.rollback()
            print("活动不存在")
            return jsonify(error_response("活动不存在")), 404

        # 检查报名时间
        now = datetime.now()
        if now < activity['registration_start']:
            conn.rollback()
            print("报名尚未开始")
            return jsonify(error_response("报名尚未开始")), 400
        if now > activity['registration_end']:
            conn.rollback()
            print("报名已截止")
            return jsonify(error_response("报名已截止")), 400

        # 检查人数限制
        if activity['current_participants'] >= activity['max_participants']:
            conn.rollback()
            print("报名人数已满")
            return jsonify(error_response("报名人数已满")), 400

        # 插入报名记录
        skip_insurance = data.get('skip_insurance', False)
        skip_transport = data.get('skip_transport', False)
        skip_meal = data.get('skip_meal', False)
        total_amount = data.get('total_amount', 0)

        print(f"插入报名记录 - 金额: {total_amount}, 跳过保险: {skip_insurance}")

        # 判断是否免审核
        no_review = activity.get('no_review_needed', 0)
        if no_review:
            reg_status = 'approved'
        else:
            reg_status = 'pending'

        cursor.execute("""
            INSERT INTO user_activities
            (user_id, activity_id, total_amount, skip_insurance, skip_transport, skip_meal, status, payment_status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'unpaid')
        """, (user_id, activity_id, total_amount, skip_insurance, skip_transport, skip_meal, reg_status))

        registration_id = cursor.lastrowid
        print(f"报名记录创建成功，ID: {registration_id}")

        # 更新活动报名人数
        cursor.execute(
            "UPDATE activities SET current_participants = current_participants + 1 WHERE id = %s",
            (activity_id,)
        )

        # 提交事务
        conn.commit()
        print("事务提交成功")

        # ===== 新增：记录日志和发送通知 =====
        try:
            # 记录操作日志
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='报名活动',
                operation_module='活动报名',
                operation_desc=f"报名活动ID:{activity_id}, 金额:{total_amount}",
                request_params=data
            )

            # 免审核不通知管理员
            if not no_review:
                Notifier.notify_admin_new_registration(
                    activity_title=activity['title'],
                    user_name=request.current_user.get('real_name', '用户'),
                    item_type='活动'
                )
        except Exception as e:
            print(f"日志/通知记录失败（不影响主流程）: {str(e)}")

        if no_review:
            return jsonify(success_response({
                'registration_id': registration_id,
                'need_payment': True
            }, "报名成功，请完成支付"))
        else:
            return jsonify(success_response({
                'registration_id': registration_id,
                'need_payment': False
            }, "报名成功，等待审核"))


    except Exception as e:
        # 发生错误时回滚
        if conn:
            conn.rollback()

        print(f"\n!!!!! 报名失败 !!!!!")
        print(f"异常类型: {type(e)}")
        print(f"异常信息: {str(e)}")

        # 记录错误日志
        try:
            Logger.log_error(
                user_id=user_id,
                error_type='报名失败',
                error_message=str(e),
                error_stack=traceback.format_exc(),
                request_params=data
            )
        except:
            pass

        import traceback
        traceback.print_exc()

        return jsonify(error_response(f"报名失败: {str(e)}")), 500

    finally:
        # 确保关闭连接
        if cursor:
            cursor.close()
        if conn:
            conn.close()
        print("===== 报名处理结束 =====\n")


@app.route('/api/activities/<int:activity_id>/photos', methods=['GET'])
def get_activity_photos(activity_id):
    """获取活动照片"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT ap.*, u.real_name as uploader_name 
            FROM activity_photos ap 
            JOIN users u ON ap.user_id = u.id 
            WHERE ap.activity_id = %s 
            ORDER BY ap.upload_time DESC
        """, (activity_id,))

        photos = cursor.fetchall()
        cursor.close()
        conn.close()

        return jsonify(success_response(photos))
    except Exception as e:
        return jsonify(error_response(f"获取活动照片失败: {str(e)}")), 500


@app.route('/api/activities/<int:activity_id>/photos', methods=['POST'])
@login_required
def upload_activity_photo(activity_id):
    """上传活动照片"""
    try:
        timestamp = request.args.get('t', 'no_timestamp')
        random_id = request.args.get('r', 'no_random')

        print(f"\n===== 活动照片上传请求 =====")
        print(f"活动ID: {activity_id}")
        print(f"时间戳: {timestamp}, 随机数: {random_id}")
        print(f"请求文件: {request.files}")

        if 'photo' not in request.files:
            return jsonify(error_response("没有上传文件")), 400

        file = request.files['photo']

        print(f"文件名: {file.filename}")

        if file.filename == '':
            return jsonify(error_response("没有选择文件")), 400

        photo_url = save_uploaded_file(file, 'activity_photos')

        print(f"保存后的URL: {photo_url}")

        if not photo_url:
            return jsonify(error_response("文件上传失败")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO activity_photos (activity_id, user_id, photo_url) 
            VALUES (%s, %s, %s)
        """, (activity_id, request.current_user['id'], photo_url))

        photo_id = cursor.lastrowid

        conn.commit()
        cursor.close()
        conn.close()

        print(f"数据库插入成功，photo_id: {photo_id}")
        print(f"===== 上传完成 =====\n")

        return jsonify(success_response({
            'photo_id': photo_id,
            'photo_url': photo_url
        }, "照片上传成功"))
    except Exception as e:
        print(f"上传照片失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"照片上传失败: {str(e)}")), 500


# ==================== 用户相关 ====================
@app.route('/api/users/profile', methods=['GET'])
@login_required
def get_user_profile():
    """获取用户信息"""
    try:
        user = request.current_user.copy()

        # 检查年险状态
        if user['has_annual_insurance'] and user['annual_insurance_end']:
            end_date = user['annual_insurance_end']
            if hasattr(end_date, 'date'):
                end_date = end_date.date()
            elif isinstance(end_date, str):
                try:
                    end_date = datetime.strptime(end_date.split(' ')[0], '%Y-%m-%d').date()
                except:
                    end_date = None

            if end_date:
                user['insurance_valid'] = end_date >= datetime.now().date()
            else:
                user['insurance_valid'] = False
        else:
            user['insurance_valid'] = False

        # ===== 新增：检查是否有进行中的活动（临时保险） =====
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT a.activity_start, a.activity_end, a.title
            FROM user_activities ua
            JOIN activities a ON ua.activity_id = a.id
            WHERE ua.user_id = %s 
            AND ua.status IN ('approved', 'completed')
            AND a.activity_end >= NOW()
            ORDER BY a.activity_start ASC
            LIMIT 1
        """, (user['id'],))

        ongoing_activity = cursor.fetchone()

        if ongoing_activity:
            user['has_ongoing_activity'] = True
            user['ongoing_activity_title'] = ongoing_activity['title']
            user['ongoing_activity_start'] = format_datetime_to_chinese(ongoing_activity['activity_start'])
            user['ongoing_activity_end'] = format_datetime_to_chinese(ongoing_activity['activity_end'])

            # 如果没有年险或年险过期，但有进行中的活动，视为有临时保险
            if not user['insurance_valid']:
                user['has_temporary_insurance'] = True
        else:
            user['has_ongoing_activity'] = False
            user['has_temporary_insurance'] = False

        cursor.close()
        conn.close()

        # 格式化日期为中文
        if user.get('annual_insurance_start'):
            user['annual_insurance_start_cn'] = format_datetime_to_chinese(user['annual_insurance_start'])
        if user.get('annual_insurance_end'):
            user['annual_insurance_end_cn'] = format_datetime_to_chinese(user['annual_insurance_end'])

        return jsonify(success_response(user))
    except Exception as e:
        print(f"\n!!!!! 获取用户信息失败 !!!!!")
        print(f"异常信息: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"获取用户信息失败: {str(e)}")), 500


@app.route('/api/users/training-stats', methods=['GET'])
@login_required
def get_user_training_stats():
    """获取用户训练统计数据"""
    try:
        user = request.current_user
        user_id = user['id']


        conn = get_db_connection()
        cursor = conn.cursor()


        # 累计天数
        created_at = user.get('created_at')
        if created_at:
            if hasattr(created_at, 'date'):
                reg_date = created_at.date()
            elif isinstance(created_at, str):
                try:
                    reg_date = datetime.strptime(created_at.split(' ')[0], '%Y-%m-%d').date()
                except:
                    reg_date = datetime.now().date()
            else:
                reg_date = datetime.now().date()
            cumulative_days = (datetime.now().date() - reg_date).days
        else:
            cumulative_days = 0


        # 获取所有已完成/已批准的活动
        cursor.execute("""
            SELECT a.activity_start, a.activity_end
            FROM user_activities ua
            JOIN activities a ON ua.activity_id = a.id
            WHERE ua.user_id = %s AND ua.status IN ('approved', 'completed')
            AND a.activity_start IS NOT NULL AND a.activity_end IS NOT NULL
        """, (user_id,))
        activities = cursor.fetchall()


        # 获取所有已完成/已批准的课程预约
        cursor.execute("""
            SELECT tcb.booking_date, tcs.schedule_date, tc.duration
            FROM teacher_course_bookings tcb
            LEFT JOIN teacher_courses tc ON tcb.course_id = tc.id
            LEFT JOIN teacher_course_schedules tcs ON tcb.schedule_id = tcs.id
            WHERE tcb.user_id = %s AND tcb.status IN ('approved', 'completed')
        """, (user_id,))
        course_bookings = cursor.fetchall()


        # ===== 新增：获取会员卡核销记录 =====
        cursor.execute("""
            SELECT created_at FROM membership_consume_logs
            WHERE user_id = %s
        """, (user_id,))
        consume_records = cursor.fetchall()


        cursor.close()
        conn.close()


        # 计算累计训练次数（活动按天累计 + 课程每次1）
        cumulative_training_count = 0
        all_training_dates = []


        for act in activities:
            start = act['activity_start']
            end = act['activity_end']
            if hasattr(start, 'date'):
                start_date = start.date()
            else:
                try:
                    start_date = datetime.strptime(str(start).split(' ')[0], '%Y-%m-%d').date()
                except:
                    continue
            if hasattr(end, 'date'):
                end_date = end.date()
            else:
                try:
                    end_date = datetime.strptime(str(end).split(' ')[0], '%Y-%m-%d').date()
                except:
                    continue
            days = (end_date - start_date).days + 1
            if days < 1:
                days = 1
            cumulative_training_count += days
            # 收集每个训练日期
            for i in range(days):
                d = start_date + timedelta(days=i)
                all_training_dates.append(d)


        for booking in course_bookings:
            cumulative_training_count += 1
            # 课程预约日期
            booking_date = booking.get('schedule_date') or booking.get('booking_date')
            if booking_date:
                if hasattr(booking_date, 'date'):
                    all_training_dates.append(booking_date.date())
                elif isinstance(booking_date, str):
                    try:
                        all_training_dates.append(datetime.strptime(str(booking_date).split(' ')[0], '%Y-%m-%d').date())
                    except:
                        pass
                else:
                    try:
                        all_training_dates.append(booking_date)
                    except:
                        pass


        # ===== 新增：会员卡核销也计为训练 =====
        for rec in consume_records:
            cumulative_training_count += 1
            c_date = rec['created_at']
            if hasattr(c_date, 'date'):
                c_date = c_date.date()
            elif isinstance(c_date, str):
                try:
                    c_date = datetime.strptime(str(c_date).split(' ')[0], '%Y-%m-%d').date()
                except:
                    continue
            else:
                try:
                    c_date = c_date.date() if hasattr(c_date, 'date') else c_date
                except:
                    continue
            all_training_dates.append(c_date)


        # 本月统计
        today = datetime.now().date()
        month_start = today.replace(day=1)
        # 下个月第一天
        if today.month == 12:
            month_end = today.replace(year=today.year + 1, month=1, day=1)
        else:
            month_end = today.replace(month=today.month + 1, day=1)


        monthly_training_count = 0
        monthly_training_dates = set()
        monthly_training_minutes = 0


        for act in activities:
            start = act['activity_start']
            end = act['activity_end']
            if hasattr(start, 'date'):
                start_date = start.date()
            else:
                try:
                    start_date = datetime.strptime(str(start).split(' ')[0], '%Y-%m-%d').date()
                except:
                    continue
            if hasattr(end, 'date'):
                end_date = end.date()
            else:
                try:
                    end_date = datetime.strptime(str(end).split(' ')[0], '%Y-%m-%d').date()
                except:
                    continue


            # 与本月的交集
            overlap_start = max(start_date, month_start)
            overlap_end = min(end_date, month_end - timedelta(days=1))
            if overlap_start <= overlap_end:
                overlap_days = (overlap_end - overlap_start).days + 1
                monthly_training_count += overlap_days
                for i in range(overlap_days):
                    monthly_training_dates.add(overlap_start + timedelta(days=i))


                # 计算时长（分钟）
                if hasattr(start, 'hour') and hasattr(end, 'hour'):
                    total_minutes = int((end - start).total_seconds() / 60)
                    if total_minutes > 0:
                        total_days = (end_date - start_date).days + 1
                        if total_days > 0:
                            monthly_training_minutes += int(total_minutes * overlap_days / total_days)


        for booking in course_bookings:
            booking_date = booking.get('schedule_date') or booking.get('booking_date')
            if booking_date:
                if hasattr(booking_date, 'date'):
                    b_date = booking_date.date()
                elif isinstance(booking_date, str):
                    try:
                        b_date = datetime.strptime(str(booking_date).split(' ')[0], '%Y-%m-%d').date()
                    except:
                        continue
                else:
                    try:
                        b_date = booking_date
                    except:
                        continue


                if month_start <= b_date < month_end:
                    monthly_training_count += 1
                    monthly_training_dates.add(b_date)
                    duration = booking.get('duration') or 0
                    try:
                        monthly_training_minutes += int(duration)
                    except:
                        pass


        # ===== 新增：核销记录的本月统计 =====
        for rec in consume_records:
            c_date = rec['created_at']
            if hasattr(c_date, 'date'):
                c_date = c_date.date()
            elif isinstance(c_date, str):
                try:
                    c_date = datetime.strptime(str(c_date).split(' ')[0], '%Y-%m-%d').date()
                except:
                    continue
            else:
                try:
                    c_date = c_date.date() if hasattr(c_date, 'date') else c_date
                except:
                    continue


            if month_start <= c_date < month_end:
                monthly_training_count += 1
                monthly_training_dates.add(c_date)


        result = {
            'cumulative_days': cumulative_days,
            'cumulative_training_count': cumulative_training_count,
            'monthly_training_count': monthly_training_count,
            'monthly_training_days': len(monthly_training_dates),
            'monthly_training_minutes': monthly_training_minutes
        }


        return jsonify(success_response(result, "获取成功"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"获取训练统计失败: {str(e)}")), 500



@app.route('/api/users/profile', methods=['PUT'])
@login_required
def update_user_profile():
    """更新用户信息"""
    try:
        data = request.get_json()
        user_id = request.current_user['id']


        # ★ 修复：添加 gender 和 age 到可更新字段
        updatable_fields = [
            'real_name', 'id_card', 'phone', 'gender', 'age', 'nation',
            'emergency_contact_name', 'emergency_contact_phone',
            'blood_type', 'allergy_history', 'contraindications',
            'avatar_url', 'teacher_cover_image', 'teacher_intro', 'teacher_detail'
        ]


        # 使用字典收集更新，避免重复字段
        updates = {}
        for field in updatable_fields:
            if field in data:
                updates[field] = data[field]


        # 如果更新了身份证号，自动计算年龄和性别（覆盖手动输入）
        if 'id_card' in data and data['id_card']:
            age = calculate_age_from_id_card(data['id_card'])
            gender = get_gender_from_id_card(data['id_card'])
            if age:
                updates['age'] = age
            if gender:
                updates['gender'] = gender


        if not updates:
            return jsonify(error_response("没有可更新的字段")), 400


        update_fields = [f"{k} = %s" for k in updates.keys()]
        params = list(updates.values())
        params.append(user_id)


        conn = get_db_connection()
        cursor = conn.cursor()


        sql = f"UPDATE users SET {', '.join(update_fields)} WHERE id = %s"
        cursor.execute(sql, params)


        conn.commit()


        try:
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='修改个人信息',
                operation_module='用户管理',
                operation_desc=f"更新个人信息",
                request_params=data
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")


        cursor.close()
        conn.close()


        return jsonify(success_response(None, "用户信息更新成功"))
    except Exception as e:
        return jsonify(error_response(f"更新用户信息失败: {str(e)}")), 500






@app.route('/api/users/activities/<int:registration_id>/cancel', methods=['POST'])
@login_required
def cancel_activity_registration(registration_id):
    """取消活动报名（仅限未支付）"""
    try:
        data = request.get_json()
        cancel_reason = data.get('cancel_reason', '用户主动取消')
        user_id = request.current_user['id']


        conn = get_db_connection()
        cursor = conn.cursor()


        cursor.execute("""
            SELECT ua.*, a.title 
            FROM user_activities ua JOIN activities a ON ua.activity_id = a.id
            WHERE ua.id = %s AND ua.user_id = %s
        """, (registration_id, user_id))


        registration = cursor.fetchone()
        if not registration:
            cursor.close()
            conn.close()
            return jsonify(error_response("报名记录不存在")), 404


        if registration['status'] in ['cancelled', 'completed']:
            cursor.close()
            conn.close()
            return jsonify(error_response("该报名已取消或已完成，无法再次取消")), 400


        # ★ 新增：已付款的报名不能直接取消，需走退款流程
        if registration['payment_status'] == 'paid':
            cursor.close()
            conn.close()
            return jsonify(error_response("已支付的报名请通过退款流程取消")), 400


        # 直接取消，不需要审核
        cursor.execute("""
            UPDATE user_activities 
            SET status = 'cancelled', cancel_reason = %s 
            WHERE id = %s
        """, (cancel_reason, registration_id))


        # 减少活动报名人数
        cursor.execute("""
            UPDATE activities 
            SET current_participants = GREATEST(0, current_participants - 1)
            WHERE id = %s
        """, (registration['activity_id'],))


        conn.commit()


        try:
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='取消报名',
                operation_module='活动报名',
                operation_desc=f"取消报名活动: {registration['title']}, 原因: {cancel_reason}",
                request_params=data
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")


        cursor.close()
        conn.close()


        # ★ 修复：直接取消，无需审核
        return jsonify(success_response(None, "已取消报名"))
    except Exception as e:
        return jsonify(error_response(f"取消报名失败: {str(e)}")), 500


@app.route('/api/users/insurance-status', methods=['GET'])
@login_required
def get_user_insurance_status():
    """获取用户保险状态"""
    try:
        user = request.current_user

        result = {
            'has_annual_insurance': user['has_annual_insurance'],
            'start_date': user['annual_insurance_start'],
            'end_date': user['annual_insurance_end'],
            'is_valid': False
        }

        if user['has_annual_insurance'] and user['annual_insurance_end']:
            result['is_valid'] = user['annual_insurance_end'] >= datetime.now().date()

        return jsonify(success_response(result))
    except Exception as e:
        return jsonify(error_response(f"获取保险状态失败: {str(e)}")), 500

# ==================== 商品和订单相关 ====================
@app.route('/api/products', methods=['GET'])
def get_products():
    """获取商品列表"""
    try:
        search = request.args.get('search', '')
        category = request.args.get('category', '')
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        offset = (page - 1) * limit

        conn = get_db_connection()
        cursor = conn.cursor()

        where_clause = "WHERE status = 'active'"
        params = []

        if search:
            where_clause += " AND name LIKE %s"
            params.append(f'%{search}%')

        if category:
            where_clause += " AND category = %s"
            params.append(category)

        sql = f"""
        SELECT * FROM products
        {where_clause}
        ORDER BY created_at DESC
        LIMIT %s OFFSET %s
        """
        params.extend([limit, offset])

        cursor.execute(sql, params)
        products = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify(success_response(products))
    except Exception as e:
        return jsonify(error_response(f"获取商品列表失败: {str(e)}")), 500

@app.route('/api/products/<int:product_id>', methods=['GET'])
def get_product_detail(product_id):
    """获取商品详情"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM products WHERE id = %s", (product_id,))
        product = cursor.fetchone()

        if not product:
            return jsonify(error_response("商品不存在", 404)), 404

        cursor.close()
        conn.close()

        return jsonify(success_response(product))
    except Exception as e:
        return jsonify(error_response(f"获取商品详情失败: {str(e)}")), 500


# ==================== 商品分类管理 ====================
@app.route('/api/product-categories', methods=['GET'])
def get_product_categories():
    """获取商品分类列表"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM product_categories
            WHERE status = 'active'
            ORDER BY sort_order ASC, id ASC
        """)
        categories = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify(success_response(categories, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取商品分类失败: {str(e)}")), 500


@app.route('/api/product-categories', methods=['POST'])
@admin_required
def create_product_category():
    """创建商品分类（管理员）"""
    try:
        data = request.json
        category_name = data.get('category_name', '')
        sort_order = data.get('sort_order', 0)

        if not category_name:
            return jsonify(error_response("分类名称不能为空")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO product_categories (category_name, sort_order, status)
            VALUES (%s, %s, 'active')
        """, (category_name, sort_order))

        category_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response({'id': category_id}, "创建成功"))
    except Exception as e:
        return jsonify(error_response(f"创建分类失败: {str(e)}")), 500


@app.route('/api/product-categories/<int:category_id>', methods=['PUT'])
@admin_required
def update_product_category(category_id):
    """更新商品分类（管理员）"""
    try:
        data = request.json

        conn = get_db_connection()
        cursor = conn.cursor()

        updates = []
        params = []

        for field in ['category_name', 'sort_order', 'status']:
            if field in data:
                updates.append(f"{field} = %s")
                params.append(data[field])

        if updates:
            params.append(category_id)
            cursor.execute(f"""
                UPDATE product_categories SET {', '.join(updates)} WHERE id = %s
            """, params)
            conn.commit()

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "更新成功"))
    except Exception as e:
        return jsonify(error_response(f"更新分类失败: {str(e)}")), 500


@app.route('/api/product-categories/<int:category_id>', methods=['DELETE'])
@admin_required
def delete_product_category(category_id):
    """删除商品分类（管理员）"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("UPDATE product_categories SET status = 'inactive' WHERE id = %s", (category_id,))
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(None, "删除成功"))
    except Exception as e:
        return jsonify(error_response(f"删除分类失败: {str(e)}")), 500


@app.route('/api/cart', methods=['GET'])
@login_required
def get_cart():
    """获取购物车"""
    try:
        user_id = request.current_user['id']


        conn = get_db_connection()
        cursor = conn.cursor()


        cursor.execute("""
            SELECT c.*, p.name, p.price, p.image_url, p.stock, p.category,
                   u_teacher.real_name as delivery_teacher_name
            FROM cart c
            JOIN products p ON c.product_id = p.id
            LEFT JOIN users u_teacher ON c.delivery_teacher_id = u_teacher.id
            WHERE c.user_id = %s AND p.status = 'active'
            ORDER BY c.created_at DESC
        """, (user_id,))


        cart_items = cursor.fetchall()
        cursor.close()
        conn.close()


        return jsonify(success_response(cart_items))
    except Exception as e:
        return jsonify(error_response(f"获取购物车失败: {str(e)}")), 500


@app.route('/api/cart', methods=['POST'])
@login_required
def add_to_cart():
    """添加到购物车"""
    try:
        data = request.get_json()
        user_id = request.current_user['id']
        product_id = data.get('product_id')
        quantity = data.get('quantity', 1)
        delivery_teacher_id = data.get('delivery_teacher_id')  # 新增


        conn = get_db_connection()
        cursor = conn.cursor()


        # 检查商品是否存在
        cursor.execute("SELECT * FROM products WHERE id = %s AND status = 'active'", (product_id,))
        product = cursor.fetchone()
        if not product:
            cursor.close()
            conn.close()
            return jsonify(error_response("商品不存在或已下架")), 404


        # 检查库存
        if product['stock'] < quantity:
            cursor.close()
            conn.close()
            return jsonify(error_response("库存不足")), 400


        # 检查购物车中是否已有此商品
        cursor.execute("SELECT * FROM cart WHERE user_id = %s AND product_id = %s", (user_id, product_id))
        existing = cursor.fetchone()


        if existing:
            new_quantity = existing['quantity'] + quantity
            if new_quantity > product['stock']:
                cursor.close()
                conn.close()
                return jsonify(error_response("加入数量超过库存限制")), 400


            cursor.execute("""
                UPDATE cart SET quantity = %s, delivery_teacher_id = %s
                WHERE user_id = %s AND product_id = %s
            """, (new_quantity, delivery_teacher_id, user_id, product_id))
        else:
            cursor.execute("""
                INSERT INTO cart (user_id, product_id, quantity, delivery_teacher_id)
                VALUES (%s, %s, %s, %s)
            """, (user_id, product_id, quantity, delivery_teacher_id))


        conn.commit()


        try:
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='添加购物车',
                operation_module='购物车',
                operation_desc=f"添加商品ID {product_id} 到购物车，数量: {quantity}",
                request_params=data
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")


        cursor.close()
        conn.close()


        return jsonify(success_response(None, "已添加到购物车"))
    except Exception as e:
        return jsonify(error_response(f"添加到购物车失败: {str(e)}")), 500


@app.route('/api/cart/<int:cart_id>', methods=['PUT'])
@login_required
def update_cart_item(cart_id):
    """更新购物车商品数量"""
    try:
        data = request.get_json()
        quantity = data.get('quantity', 1)
        user_id = request.current_user['id']

        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查购物车项是否存在
        cursor.execute("""
            SELECT c.*, p.stock 
            FROM cart c 
            JOIN products p ON c.product_id = p.id 
            WHERE c.id = %s AND c.user_id = %s
        """, (cart_id, user_id))

        cart_item = cursor.fetchone()
        if not cart_item:
            cursor.close()
            conn.close()
            return jsonify(error_response("购物车项不存在")), 404

        if quantity > cart_item['stock']:
            cursor.close()
            conn.close()
            return jsonify(error_response("数量超过库存限制")), 400

        cursor.execute("UPDATE cart SET quantity = %s WHERE id = %s", (quantity, cart_id))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(None, "购物车已更新"))
    except Exception as e:
        return jsonify(error_response(f"更新购物车失败: {str(e)}")), 500

@app.route('/api/cart/<int:cart_id>', methods=['DELETE'])
@login_required
def delete_cart_item(cart_id):
    """删除购物车商品"""
    try:
        user_id = request.current_user['id']

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM cart WHERE id = %s AND user_id = %s", (cart_id, user_id))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(None, "商品已从购物车移除"))
    except Exception as e:
        return jsonify(error_response(f"删除购物车商品失败: {str(e)}")), 500


@app.route('/api/orders', methods=['POST'])
@login_required
def create_order():
    """创建订单"""
    try:
        data = request.get_json()
        user_id = request.current_user['id']
        items = data.get('items', [])
        address_id = data.get('address_id')


        if not items:
            return jsonify(error_response("订单商品不能为空")), 400


        if not address_id:
            return jsonify(error_response("请选择收货地址")), 400


        conn = get_db_connection()
        cursor = conn.cursor()


        # 获取地址信息
        cursor.execute("SELECT * FROM user_addresses WHERE id = %s AND user_id = %s", (address_id, user_id))
        address = cursor.fetchone()


        if not address:
            cursor.close()
            conn.close()
            return jsonify(error_response("收货地址不存在")), 404


        # 生成订单号
        order_no = f"OD{int(time.time())}{user_id}"
        total_amount = 0


        # 验证商品并计算总价
        for item in items:
            cursor.execute("SELECT * FROM products WHERE id = %s AND status = 'active'", (item['product_id'],))
            product = cursor.fetchone()
            if not product:
                cursor.close()
                conn.close()
                return jsonify(error_response(f"商品ID {item['product_id']} 不存在或已下架")), 400
            if product['stock'] < item['quantity']:
                cursor.close()
                conn.close()
                return jsonify(error_response(f"商品 {product['name']} 库存不足")), 400
            total_amount += product['price'] * item['quantity']


        # 从购物车中读取 delivery_teacher_id（在删除前）
        delivery_teacher_id = None
        product_ids = [item['product_id'] for item in items]
        if product_ids:
            placeholders = ','.join(['%s'] * len(product_ids))
            cursor.execute(f"""
                SELECT delivery_teacher_id FROM cart
                WHERE user_id = %s AND product_id IN ({placeholders}) AND delivery_teacher_id IS NOT NULL
                LIMIT 1
            """, [user_id] + product_ids)
            dt_row = cursor.fetchone()
            if dt_row:
                delivery_teacher_id = dt_row['delivery_teacher_id']


        shipping_address = f"{address['province']} {address['city']} {address['district']} {address['detail_address']}"


        cursor.execute("""
            INSERT INTO orders
            (order_no, user_id, total_amount, address_id, receiver_name, receiver_phone,
             shipping_address, status, payment_status, delivery_teacher_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'pending', 'unpaid', %s)
        """, (order_no, user_id, total_amount, address_id, address['receiver_name'],
              address['receiver_phone'], shipping_address, delivery_teacher_id))


        order_id = cursor.lastrowid


        for item in items:
            cursor.execute("SELECT * FROM products WHERE id = %s", (item['product_id'],))
            product = cursor.fetchone()


            cursor.execute("""
                INSERT INTO order_items (order_id, product_id, quantity, unit_price)
                VALUES (%s, %s, %s, %s)
            """, (order_id, item['product_id'], item['quantity'], product['price']))


            cursor.execute("UPDATE products SET stock = stock - %s WHERE id = %s",
                           (item['quantity'], item['product_id']))


            cursor.execute("DELETE FROM cart WHERE user_id = %s AND product_id = %s",
                           (user_id, item['product_id']))


        conn.commit()


        try:
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='创建订单',
                operation_module='订单管理',
                operation_desc=f"创建订单 {order_no}，金额: ￥{total_amount}",
                request_params={'items': items, 'address_id': address_id}
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")


        cursor.close()
        conn.close()


        return jsonify(success_response({
            'order_id': order_id,
            'order_no': order_no,
            'total_amount': total_amount
        }, "订单创建成功"))
    except Exception as e:
        return jsonify(error_response(f"创建订单失败: {str(e)}")), 500


@app.route('/api/orders', methods=['GET'])
@login_required
def get_user_orders():
    """获取用户订单列表"""
    try:
        user_id = request.current_user['id']
        status = request.args.get('status', 'all')

        conn = get_db_connection()
        cursor = conn.cursor()

        where_clause = "WHERE user_id = %s"
        params = [user_id]

        if status != 'all':
            where_clause += " AND status = %s"
            params.append(status)

        cursor.execute(f"""
            SELECT * FROM orders 
            {where_clause}
            ORDER BY created_at DESC
        """, params)

        orders = cursor.fetchall()

        # 获取每个订单的商品详情
        for order in orders:
            cursor.execute("""
                SELECT oi.*, p.name, p.image_url 
                FROM order_items oi 
                JOIN products p ON oi.product_id = p.id 
                WHERE oi.order_id = %s
            """, (order['id'],))
            order['items'] = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify(success_response(orders))
    except Exception as e:
        return jsonify(error_response(f"获取订单列表失败: {str(e)}")), 500

# ==================== 管理员相关 ====================
@app.route('/api/admin/activities', methods=['GET'])
@admin_or_teacher_required
def admin_get_activities():
    """管理员获取活动列表"""
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        offset = (page - 1) * limit

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT a.*, u.real_name as creator_name
            FROM activities a 
            LEFT JOIN users u ON a.creator_id = u.id
            ORDER BY a.is_top DESC, a.sort_order ASC, a.created_at DESC 
            LIMIT %s OFFSET %s
        """, (limit, offset))

        activities = cursor.fetchall()

        for activity in activities:
            # 处理封面图片
            if activity.get('cover_images'):
                try:
                    if isinstance(activity['cover_images'], str):
                        activity['cover_images'] = json.loads(activity['cover_images'])
                except:
                    activity['cover_images'] = []
            else:
                activity['cover_images'] = []

            # 如果没有封面图片，设置默认图片
            if not activity['cover_images']:
                activity['cover_images'] = ['https://picsum.photos/160/160?random=' + str(activity['id'])]

            # 为了兼容前端，添加 cover_image_url 字段（取第一张图）
            activity['cover_image_url'] = activity['cover_images'][0] if activity[
                'cover_images'] else 'https://picsum.photos/160/160?random=' + str(activity['id'])

            # 格式化日期为中文
            if activity.get('activity_start'):
                activity['activity_start_cn'] = format_datetime_to_chinese(activity['activity_start'])
            if activity.get('activity_end'):
                activity['activity_end_cn'] = format_datetime_to_chinese(activity['activity_end'])
            if activity.get('registration_start'):
                activity['registration_start_cn'] = format_datetime_to_chinese(activity['registration_start'])
            if activity.get('registration_end'):
                activity['registration_end_cn'] = format_datetime_to_chinese(activity['registration_end'])

            if activity.get('additional_fees'):
                try:
                    if isinstance(activity['additional_fees'], str):
                        activity['additional_fees'] = json.loads(activity['additional_fees'])
                except:
                    activity['additional_fees'] = []

        cursor.close()
        conn.close()

        return jsonify(success_response(activities))
    except Exception as e:
        return jsonify(error_response(f"获取活动列表失败: {str(e)}")), 500


@app.route('/api/admin/activities', methods=['POST'])
@admin_or_teacher_required
def admin_create_activity():
    """管理员创建活动"""
    print("\n===== 开始创建活动 =====")


    try:
        data = request.get_json()
        print(f"接收到的数据: {data}")


        creator_id = request.current_user['id']


        additional_fees = data.get('additional_fees', [])
        if additional_fees:
            additional_fees = json.dumps(additional_fees)
        else:
            additional_fees = None


        conn = get_db_connection()
        cursor = conn.cursor()


        # ★ 修复：在INSERT中添加 no_review_needed 字段
        cursor.execute("""
            INSERT INTO activities 
            (title, description, category, cover_images, registration_start, registration_end,
             activity_start, activity_end, cancel_deadline, location, latitude, longitude, registration_requirements,
             fee_details, base_fee, insurance_fee, transport_fee, meal_fee, additional_fees, max_participants, notices,
             is_top, is_carousel, sort_order, creator_id, status, no_review_needed)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data.get('title', ''),
            data.get('description', ''),
            data.get('category', '徒步'),
            json.dumps(data.get('cover_images', [])) if data.get('cover_images') else None,
            data.get('registration_start', None),
            data.get('registration_end', None),
            data.get('activity_start', None),
            data.get('activity_end', None),
            data.get('cancel_deadline', None),
            data.get('location', ''),
            data.get('latitude', None),
            data.get('longitude', None),
            data.get('registration_requirements', ''),
            data.get('fee_details', ''),
            data.get('base_fee', 0),
            data.get('insurance_fee', 0),
            data.get('transport_fee', 0),
            data.get('meal_fee', 0),
            additional_fees,
            data.get('max_participants', 50),
            data.get('notices', ''),
            data.get('is_top', False),
            data.get('is_carousel', False),
            data.get('sort_order', 0),
            creator_id,
            data.get('status', 'published'),
            1 if data.get('no_review_needed') else 0
        ))


        activity_id = cursor.lastrowid
        print(f"活动创建成功，ID: {activity_id}")


        conn.commit()


        try:
            Logger.log_operation(
                user_id=creator_id,
                user_name=request.current_user.get('real_name', '管理员'),
                operation_type='创建活动',
                operation_module='活动管理',
                operation_desc=f"创建活动: {data.get('title', '')}",
                request_params=data
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")


        cursor.close()
        conn.close()


        return jsonify(success_response({'activity_id': activity_id}, "活动创建成功"))


    except Exception as e:
        print(f"\n!!!!! 创建活动失败 !!!!!")
        print(f"异常信息: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"创建活动失败: {str(e)}")), 500



@app.route('/api/admin/activities/<int:activity_id>', methods=['PUT'])
@admin_or_teacher_required
def admin_update_activity(activity_id):
    """管理员更新活动"""
    try:
        data = request.get_json()


        print(f"\n===== 更新活动 ID: {activity_id} =====")
        print(f"接收到的完整数据: {data}")


        # ★ 修复：添加 no_review_needed 到可更新字段
        updatable_fields = [
            'title', 'description', 'category', 'cover_images',
            'registration_start', 'registration_end', 'activity_start', 'activity_end', 'cancel_deadline',
            'location', 'latitude', 'longitude', 'registration_requirements',
            'fee_details', 'base_fee', 'insurance_fee', 'transport_fee', 'meal_fee',
            'max_participants', 'notices', 'is_top', 'is_carousel', 'sort_order', 'status',
            'no_review_needed'
        ]


        update_fields = []
        params = []


        for field in updatable_fields:
            if field in data:
                if field == 'additional_fees' and data[field]:
                    update_fields.append(f"{field} = %s")
                    params.append(json.dumps(data[field]))
                elif field == 'cover_images' and data[field]:
                    update_fields.append(f"{field} = %s")
                    params.append(json.dumps(data[field]))
                elif field == 'cover_images' and not data[field]:
                    # 空数组也要正确处理
                    update_fields.append(f"{field} = %s")
                    params.append(json.dumps([]))
                else:
                    update_fields.append(f"{field} = %s")
                    params.append(data[field])


        if not update_fields:
            return jsonify(error_response("没有可更新的字段")), 400


        params.append(activity_id)


        conn = get_db_connection()
        cursor = conn.cursor()


        sql = f"UPDATE activities SET {', '.join(update_fields)} WHERE id = %s"


        print(f"执行SQL: {sql}")
        print(f"参数数量: {len(params)}")


        cursor.execute(sql, params)
        affected_rows = cursor.rowcount
        print(f"影响行数: {affected_rows}")


        conn.commit()


        try:
            Logger.log_audit(
                user_id=request.current_user['id'],
                user_name=request.current_user.get('real_name', '管理员'),
                action='修改活动',
                target_type='activities',
                target_id=activity_id,
                old_value=None,
                new_value=data
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")


        cursor.close()
        conn.close()


        return jsonify(success_response(None, "活动更新成功"))
    except Exception as e:
        print(f"\n!!!!! 更新活动异常 !!!!!")
        print(f"异常信息: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"更新活动失败: {str(e)}")), 500



@app.route('/api/admin/activities/<int:activity_id>/save-as-new', methods=['POST'])
@admin_or_teacher_required
def admin_save_activity_as_new(activity_id):
    """管理员将活动保存为新活动"""
    try:
        data = request.get_json()
        creator_id = request.current_user['id']


        new_title = data.get('title', '')


        print(f"\n===== 保存为新活动 =====")
        print(f"原活动ID: {activity_id}")
        print(f"新活动标题: {new_title}")
        print(f"接收到的 is_top: {data.get('is_top')}")
        print(f"接收到的 is_carousel: {data.get('is_carousel')}")
        print(f"接收到的 no_review_needed: {data.get('no_review_needed')}")


        conn = get_db_connection()
        cursor = conn.cursor()


        # 检查活动名是否重复
        cursor.execute("SELECT id FROM activities WHERE title = %s", (new_title,))
        existing = cursor.fetchone()


        if existing:
            cursor.close()
            conn.close()
            print(f"活动名重复，已存在ID: {existing['id']}")
            return jsonify(error_response("活动名称已存在，请修改后再保存")), 400


        # 获取原活动信息
        cursor.execute("SELECT * FROM activities WHERE id = %s", (activity_id,))
        original = cursor.fetchone()


        if not original:
            cursor.close()
            conn.close()
            return jsonify(error_response("原活动不存在")), 404


        # 处理封面图片
        cover_images = data.get('cover_images', original.get('cover_images'))
        if cover_images and isinstance(cover_images, list):
            cover_images = json.dumps(cover_images)
        elif cover_images and isinstance(cover_images, str):
            pass
        else:
            cover_images = None


        # 处理附加费用
        additional_fees = data.get('additional_fees', original.get('additional_fees'))
        if additional_fees and isinstance(additional_fees, list):
            additional_fees = json.dumps(additional_fees)
        elif additional_fees and isinstance(additional_fees, str):
            pass
        else:
            additional_fees = None


        # 创建新活动
        cursor.execute("""
            INSERT INTO activities 
            (title, description, category, cover_images, registration_start, registration_end,
             activity_start, activity_end, cancel_deadline, location, latitude, longitude, registration_requirements,
             fee_details, base_fee, insurance_fee, transport_fee, meal_fee, additional_fees, max_participants, notices,
             is_top, is_carousel, sort_order, creator_id, status, no_review_needed)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data.get('title', original['title']),
            data.get('description', original['description']),
            data.get('category', original['category']),
            cover_images,
            data.get('registration_start', original['registration_start']),
            data.get('registration_end', original['registration_end']),
            data.get('activity_start', original['activity_start']),
            data.get('activity_end', original['activity_end']),
            data.get('cancel_deadline', original.get('cancel_deadline')),
            data.get('location', original['location']),
            data.get('latitude', original.get('latitude')),
            data.get('longitude', original.get('longitude')),
            data.get('registration_requirements', original.get('registration_requirements')),
            data.get('fee_details', original.get('fee_details')),
            data.get('base_fee', original['base_fee']),
            data.get('insurance_fee', original.get('insurance_fee', 0)),
            data.get('transport_fee', original.get('transport_fee', 0)),
            data.get('meal_fee', original.get('meal_fee', 0)),
            additional_fees,
            data.get('max_participants', original['max_participants']),
            data.get('notices', original.get('notices')),
            data.get('is_top', False),
            data.get('is_carousel', False),
            0,
            creator_id,
            'published',
            1 if data.get('no_review_needed') else 0
        ))


        new_activity_id = cursor.lastrowid
        print(f"新活动创建成功，ID: {new_activity_id}")


        conn.commit()


        # 验证是否保存成功
        cursor.execute("SELECT is_top, is_carousel, no_review_needed FROM activities WHERE id = %s", (new_activity_id,))
        verify = cursor.fetchone()
        print(f"===== 验证新活动 =====")
        print(f"is_top: {verify['is_top']}")
        print(f"is_carousel: {verify['is_carousel']}")
        print(f"no_review_needed: {verify['no_review_needed']}")


        try:
            Logger.log_operation(
                user_id=creator_id,
                user_name=request.current_user.get('real_name', '管理员'),
                operation_type='保存为新活动',
                operation_module='活动管理',
                operation_desc=f"从活动ID {activity_id} 保存为新活动: {data.get('title', '')}",
                request_params=data
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")


        cursor.close()
        conn.close()


        return jsonify(success_response({'activity_id': new_activity_id}, "新活动创建成功"))
    except Exception as e:
        print(f"\n!!!!! 保存为新活动失败 !!!!!")
        print(f"异常信息: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"保存为新活动失败: {str(e)}")), 500



@app.route('/api/admin/activities/<int:activity_id>/copy', methods=['POST'])
@admin_or_teacher_required
def admin_copy_activity(activity_id):
    """管理员复制活动"""
    try:
        data = request.get_json()
        creator_id = request.current_user['id']

        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取原活动信息
        cursor.execute("SELECT * FROM activities WHERE id = %s", (activity_id,))
        original = cursor.fetchone()

        if not original:
            return jsonify(error_response("原活动不存在")), 404

        # 复制活动，允许覆盖部分字段
        new_title = data.get('title', f"{original['title']} - 副本")
        new_registration_start = data.get('registration_start', original['registration_start'])
        new_registration_end = data.get('registration_end', original['registration_end'])
        new_activity_start = data.get('activity_start', original['activity_start'])
        new_activity_end = data.get('activity_end', original['activity_end'])

        cursor.execute("""
            INSERT INTO activities 
            (title, description, category, cover_image_url, registration_start, registration_end,
             activity_start, activity_end, location, latitude, longitude, registration_requirements,
             fee_details, base_fee, additional_fees, max_participants, notices,
             is_top, sort_order, creator_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            new_title, original['description'], original['category'],
            original['cover_image_url'], new_registration_start, new_registration_end,
            new_activity_start, new_activity_end, original['location'],
            original['latitude'], original['longitude'], original['registration_requirements'],
            original['fee_details'], original['base_fee'], original['additional_fees'],
            original['max_participants'], original['notices'],
            False, 0, creator_id
        ))

        new_activity_id = cursor.lastrowid

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response({'activity_id': new_activity_id}, "活动复制成功"))
    except Exception as e:
        return jsonify(error_response(f"复制活动失败: {str(e)}")), 500

@app.route('/api/admin/activities/<int:activity_id>', methods=['DELETE'])
@admin_or_teacher_required
def admin_delete_activity(activity_id):
    """删除活动及其所有相关数据（级联删除）"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        print(f"\n===== admin_delete_activity id={activity_id} =====")

        # 检查活动是否存在
        cursor.execute("SELECT id, title FROM activities WHERE id = %s", (activity_id,))
        activity = cursor.fetchone()
        if not activity:
            cursor.close()
            conn.close()
            return jsonify(error_response("活动不存在")), 404

        # 统计关联数据数量，返回给前端做二次确认提示
        cursor.execute(
            "SELECT COUNT(*) as count FROM user_activities WHERE activity_id = %s",
            (activity_id,)
        )
        reg_count = cursor.fetchone()['count']

        cursor.execute(
            "SELECT COUNT(*) as count FROM activity_photos WHERE activity_id = %s",
            (activity_id,)
        )
        photo_count = cursor.fetchone()['count']

        print(f"关联报名记录: {reg_count}, 照片: {photo_count}")

        # 级联删除所有关联数据
        # 1. 删除保险记录（关联 user_activities）
        cursor.execute("""
            DELETE FROM insurances WHERE user_activity_id IN (
                SELECT id FROM user_activities WHERE activity_id = %s
            )
        """, (activity_id,))
        print(f"删除保险记录: {cursor.rowcount}")

        # 2. 删除报名记录
        cursor.execute("DELETE FROM user_activities WHERE activity_id = %s", (activity_id,))
        print(f"删除报名记录: {cursor.rowcount}")

        # 3. 删除活动照片记录
        cursor.execute("DELETE FROM activity_photos WHERE activity_id = %s", (activity_id,))
        print(f"删除照片记录: {cursor.rowcount}")

        # 4. 删除活动本身
        cursor.execute("DELETE FROM activities WHERE id = %s", (activity_id,))
        print(f"删除活动: {cursor.rowcount}")

        conn.commit()

        try:
            Logger.log_audit(
                user_id=request.current_user['id'],
                user_name=request.current_user.get('real_name', '管理员'),
                action='强制删除活动',
                target_type='activities',
                target_id=activity_id,
                old_value={'title': activity['title'], 'reg_count': reg_count},
                new_value=None
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response({
            'deleted_registrations': reg_count,
            'deleted_photos': photo_count
        }, "活动及所有相关数据已删除"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"删除失败: {str(e)}")), 500




@app.route('/api/admin/activities/reorder', methods=['PUT'])
@admin_required
def admin_reorder_activities():
    """管理员重新排序活动"""
    try:
        data = request.get_json()
        activity_orders = data.get('activity_orders', [])  # [{id: 1, sort_order: 0}, ...]

        if not activity_orders:
            return jsonify(error_response("排序数据不能为空")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 批量更新排序
        for item in activity_orders:
            cursor.execute("""
                UPDATE activities SET sort_order = %s WHERE id = %s
            """, (item['sort_order'], item['id']))

        conn.commit()

        # ===== 新增：记录操作日志 =====
        try:
            Logger.log_operation(
                user_id=request.current_user['id'],
                user_name=request.current_user.get('real_name', '管理员'),
                operation_type='活动排序',
                operation_module='活动管理',
                operation_desc=f"重新排序 {len(activity_orders)} 个活动",
                request_params=data
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "活动排序更新成功"))
    except Exception as e:
        return jsonify(error_response(f"更新排序失败: {str(e)}")), 500

@app.route('/api/admin/registrations', methods=['GET'])
@admin_or_teacher_required
def admin_get_registrations():
    """管理员获取报名记录"""
    try:
        status = request.args.get('status', '')
        activity_id = request.args.get('activity_id', '')
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        offset = (page - 1) * limit

        conn = get_db_connection()
        cursor = conn.cursor()

        where_clause = "WHERE 1=1"
        params = []

        if status:
            where_clause += " AND ua.status = %s"
            params.append(status)

        if activity_id:
            where_clause += " AND ua.activity_id = %s"
            params.append(activity_id)

        cursor.execute(f"""
            SELECT ua.*, u.real_name, u.phone, a.title as activity_title
            FROM user_activities ua
            JOIN users u ON ua.user_id = u.id
            JOIN activities a ON ua.activity_id = a.id
            {where_clause}
            ORDER BY ua.registration_date DESC
            LIMIT %s OFFSET %s
        """, params + [limit, offset])

        registrations = cursor.fetchall()

        for reg in registrations:
            if reg['selected_fees']:
                reg['selected_fees'] = json.loads(reg['selected_fees'])

        cursor.close()
        conn.close()

        return jsonify(success_response(registrations))
    except Exception as e:
        return jsonify(error_response(f"获取报名记录失败: {str(e)}")), 500


@app.route('/api/admin/registrations/<int:registration_id>/review', methods=['PUT'])
@admin_or_teacher_required
def admin_review_registration(registration_id):
    """管理员审核报名"""
    try:
        data = request.get_json()
        status = data.get('status')
        admin_notes = data.get('admin_notes', '')

        if status not in ['approved', 'rejected']:
            return jsonify(error_response("无效的审核状态")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取报名详情
        cursor.execute("""
            SELECT ua.*, u.id as user_id, u.real_name, a.title as activity_title
            FROM user_activities ua
            JOIN users u ON ua.user_id = u.id
            JOIN activities a ON ua.activity_id = a.id
            WHERE ua.id = %s
        """, (registration_id,))
        reg_detail = cursor.fetchone()

        cursor.execute("""
            UPDATE user_activities 
            SET status = %s, admin_notes = %s 
            WHERE id = %s
        """, (status, admin_notes, registration_id))

        conn.commit()

        # 记录审计日志和发送通知
        if reg_detail:
            try:
                Logger.log_audit(
                    user_id=request.current_user['id'],
                    user_name=request.current_user.get('real_name', '管理员'),
                    action=f"审核报名-{status}",
                    target_type='user_activities',
                    target_id=registration_id,
                    old_value={'status': 'pending'},
                    new_value={'status': status, 'admin_notes': admin_notes}
                )

                if status == 'approved':
                    Notifier.send_registration_approved(
                        user_id=reg_detail['user_id'],
                        activity_title=reg_detail['activity_title']
                    )
            except Exception as e:
                print(f"日志/通知失败: {str(e)}")

        cursor.close()
        conn.close()

        message = "报名审核通过" if status == 'approved' else "报名审核未通过"
        return jsonify(success_response(None, message))
    except Exception as e:
        try:
            Logger.log_exception(e, user_id=request.current_user.get('id'))
        except:
            pass
        return jsonify(error_response(f"审核报名失败: {str(e)}")), 500


@app.route('/api/admin/users', methods=['GET'])
@admin_required
def admin_get_users():
    """管理员获取用户列表"""
    try:
        search = request.args.get('search', '')
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        offset = (page - 1) * limit

        conn = get_db_connection()
        cursor = conn.cursor()

        where_clause = "WHERE 1=1"
        params = []

        if search:
            where_clause += " AND (real_name LIKE %s OR phone LIKE %s)"
            params.extend([f'%{search}%', f'%{search}%'])

        cursor.execute(f"""
            SELECT id, real_name, phone, gender, role, user_type, created_at,
                   has_annual_insurance, annual_insurance_start, annual_insurance_end
            FROM users 
            {where_clause}
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
        """, params + [limit, offset])

        users = cursor.fetchall()

        # 格式化每个用户的日期
        for user in users:
            if user.get('created_at'):
                user['created_at_cn'] = format_datetime_to_chinese(user['created_at'])
            if user.get('annual_insurance_start'):
                user['annual_insurance_start_cn'] = format_datetime_to_chinese(user['annual_insurance_start'])
            if user.get('annual_insurance_end'):
                user['annual_insurance_end_cn'] = format_datetime_to_chinese(user['annual_insurance_end'])

            # 检查年险有效期
            if user.get('has_annual_insurance') and user.get('annual_insurance_end'):
                end_date = user['annual_insurance_end']
                if hasattr(end_date, 'date'):
                    end_date = end_date.date()
                elif isinstance(end_date, str):
                    try:
                        end_date = datetime.strptime(end_date.split(' ')[0], '%Y-%m-%d').date()
                    except:
                        end_date = None

                if end_date:
                    user['insurance_valid'] = end_date >= datetime.now().date()
                else:
                    user['insurance_valid'] = False
            else:
                user['insurance_valid'] = False

        cursor.close()
        conn.close()

        return jsonify(success_response(users))
    except Exception as e:
        return jsonify(error_response(f"获取用户列表失败: {str(e)}")), 500


@app.route('/api/users/activities', methods=['GET'])
@login_required
def get_user_activities():
    """获取用户活动记录"""
    try:
        user_id = request.current_user['id']

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT ua.*, a.title, a.cover_images, a.activity_start, a.activity_end, a.cancel_deadline
            FROM user_activities ua 
            JOIN activities a ON ua.activity_id = a.id 
            WHERE ua.user_id = %s 
            ORDER BY ua.registration_date DESC
        """, (user_id,))

        activities = cursor.fetchall()

        # 按状态分组
        result = {
            'pending': [],
            'approved': [],
            'completed': [],
            'cancelled': [],
            'rejected': []
        }

        for activity in activities:
            # 处理封面图片
            if activity.get('cover_images'):
                try:
                    if isinstance(activity['cover_images'], str):
                        activity['cover_images'] = json.loads(activity['cover_images'])
                except:
                    activity['cover_images'] = []
            else:
                activity['cover_images'] = []

            # 如果没有封面图片，设置默认图片
            if not activity['cover_images']:
                activity['cover_images'] = ['https://picsum.photos/750/400?random=1']

            # 为了兼容前端，添加 cover_image_url 字段（取第一张图）
            activity['cover_image_url'] = activity['cover_images'][0] if activity['cover_images'] else None

            # 格式化日期为中文
            if activity.get('activity_start'):
                activity['activity_start_cn'] = format_datetime_to_chinese(activity['activity_start'])
            if activity.get('activity_end'):
                activity['activity_end_cn'] = format_datetime_to_chinese(activity['activity_end'])
            if activity.get('registration_date'):
                activity['registration_date_cn'] = format_datetime_to_chinese(activity['registration_date'])
            if activity.get('cancel_deadline'):
                activity['cancel_deadline_cn'] = format_datetime_to_chinese(activity['cancel_deadline'])

            if activity['selected_fees']:
                activity['selected_fees'] = json.loads(activity['selected_fees'])
            result[activity['status']].append(activity)

        cursor.close()
        conn.close()

        return jsonify(success_response(result))
    except Exception as e:
        return jsonify(error_response(f"获取活动记录失败: {str(e)}")), 500

@app.route('/api/admin/products', methods=['GET'])
@admin_required
def admin_get_products():
    """管理员获取商品列表"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM products ORDER BY created_at DESC")
        products = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify(success_response(products))
    except Exception as e:
        return jsonify(error_response(f"获取商品列表失败: {str(e)}")), 500

@app.route('/api/admin/products', methods=['POST'])
@admin_required
def admin_create_product():
    """管理员创建商品"""
    try:
        data = request.get_json()

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO products (name, description, price, image_url, stock, category)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            data['name'], data.get('description', ''), data['price'],
            data.get('image_url', ''), data.get('stock', 0), data.get('category', '')
        ))

        product_id = cursor.lastrowid

        conn.commit()

        # ===== 新增：记录操作日志 =====
        try:
            Logger.log_operation(
                user_id=request.current_user['id'],
                user_name=request.current_user.get('real_name', '管理员'),
                operation_type='创建商品',
                operation_module='商品管理',
                operation_desc=f"创建商品: {data['name']}",
                request_params=data
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response({'product_id': product_id}, "商品创建成功"))
    except Exception as e:
        return jsonify(error_response(f"创建商品失败: {str(e)}")), 500

@app.route('/api/admin/photos/<int:photo_id>', methods=['DELETE'])
@admin_required
def admin_delete_photo(photo_id):
    """管理员删除活动照片"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取照片信息
        cursor.execute("SELECT * FROM activity_photos WHERE id = %s", (photo_id,))
        photo = cursor.fetchone()

        if not photo:
            cursor.close()
            conn.close()
            return jsonify(error_response("照片不存在")), 404

        # 删除数据库记录
        cursor.execute("DELETE FROM activity_photos WHERE id = %s", (photo_id,))

        # 尝试删除文件
        try:
            file_path = photo['photo_url'].replace(Config.SERVER_HOST, '.')
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"文件删除成功: {file_path}")
        except Exception as e:
            print(f"文件删除失败: {e}")
            pass  # 文件删除失败不影响数据库操作

        conn.commit()

        # ===== 新增：记录操作日志 =====
        try:
            Logger.log_operation(
                user_id=request.current_user['id'],
                user_name=request.current_user.get('real_name', '管理员'),
                operation_type='删除活动照片',
                operation_module='照片管理',
                operation_desc=f"删除照片ID: {photo_id}",
                request_params={'photo_id': photo_id}
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "照片删除成功"))
    except Exception as e:
        print(f"删除照片失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"删除照片失败: {str(e)}")), 500

# ==================== 文件上传 ====================
@app.route('/api/upload/image', methods=['POST'])
def upload_image():
    """通用图片上传"""
    try:
        timestamp = request.args.get('t', 'no_timestamp')
        random_id = request.args.get('r', 'no_random')

        print(f"\n===== 图片上传请求 =====")
        print(f"时间戳: {timestamp}, 随机数: {random_id}")
        print(f"请求文件: {request.files}")

        if 'image' not in request.files:
            return jsonify(error_response("没有上传文件")), 400

        file = request.files['image']

        print(f"文件名: {file.filename}")
        print(f"文件类型: {file.content_type}")

        if file.filename == '':
            return jsonify(error_response("没有选择文件")), 400

        image_url = save_uploaded_file(file, 'images')

        print(f"保存后的URL: {image_url}")
        print(f"===== 上传完成 =====\n")

        if not image_url:
            return jsonify(error_response("文件上传失败")), 400

        return jsonify(success_response({'image_url': image_url}, "图片上传成功"))
    except Exception as e:
        print(f"图片上传异常: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"图片上传失败: {str(e)}")), 500


@app.route('/api/upload/avatar', methods=['POST'])
@login_required
def upload_avatar():
    """上传用户头像"""
    try:
        if 'avatar' not in request.files:
            return jsonify(error_response("没有上传文件")), 400

        file = request.files['avatar']
        avatar_url = save_uploaded_file(file, 'avatars')

        if not avatar_url:
            return jsonify(error_response("头像上传失败")), 400

        # 更新用户头像
        user_id = request.current_user['id']
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("UPDATE users SET avatar_url = %s WHERE id = %s", (avatar_url, user_id))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response({'avatar_url': avatar_url}, "头像上传成功"))
    except Exception as e:
        return jsonify(error_response(f"上传头像失败: {str(e)}")), 500


# ==================== 定时清理任务 ====================
@app.route('/api/admin/cleanup/photos', methods=['POST'])
@admin_required
def cleanup_old_photos():
    """清理旧照片"""
    try:
        # 计算2个月前的日期
        cutoff_date = datetime.now() - timedelta(days=60)

        conn = get_db_connection()
        cursor = conn.cursor()

        # 查找需要删除的照片
        cursor.execute("""
                SELECT ap.* FROM activity_photos ap
                JOIN activities a ON ap.activity_id = a.id
                WHERE a.activity_end < %s
            """, (cutoff_date,))

        old_photos = cursor.fetchall()
        deleted_count = 0

        for photo in old_photos:
            try:
                # 删除文件
                file_path = photo['photo_url'].replace(Config.SERVER_HOST, '.')
                if os.path.exists(file_path):
                    os.remove(file_path)

                # 删除数据库记录
                cursor.execute("DELETE FROM activity_photos WHERE id = %s", (photo['id'],))
                deleted_count += 1
            except:
                continue  # 继续处理其他照片

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(
            {'deleted_count': deleted_count},
            f"清理完成，删除了 {deleted_count} 张照片"
        ))
    except Exception as e:
        return jsonify(error_response(f"清理照片失败: {str(e)}")), 500


# ==================== 统计数据 ====================
@app.route('/api/admin/statistics', methods=['GET'])
@admin_required
def get_admin_statistics():
    """获取管理员统计数据"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 统计用户数量
        cursor.execute("SELECT COUNT(*) as count FROM users WHERE role = 'user'")
        result = cursor.fetchone()
        total_users = result['count'] if result else 0

        # 统计活动数量
        cursor.execute("SELECT COUNT(*) as count FROM activities WHERE status = 'published'")
        result = cursor.fetchone()
        total_activities = result['count'] if result else 0

        # 统计待审核报名
        cursor.execute("SELECT COUNT(*) as count FROM user_activities WHERE status = 'pending'")
        result = cursor.fetchone()
        pending_registrations = result['count'] if result else 0

        # 统计待退款申请
        cursor.execute("SELECT COUNT(*) as count FROM user_activities WHERE refund_status = 'pending'")
        result = cursor.fetchone()
        pending_refunds = result['count'] if result else 0

        cursor.close()
        conn.close()

        statistics = {
            'total_users': total_users,
            'total_activities': total_activities,
            'pending_registrations': pending_registrations,
            'pending_refunds': pending_refunds
        }

        return jsonify(success_response(statistics))
    except Exception as e:
        print(f"统计数据异常: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"获取统计数据失败: {str(e)}")), 500


# 在 app.py 中添加以下接口

@app.route('/api/auth/wechat-login', methods=['POST'])
def wechat_login_api():
    """微信登录接口"""
    try:
        data = request.get_json()
        code = data.get('code')
        user_info = data.get('userInfo', {})

        print(f"\n===== 微信登录请求 =====")
        print(f"Code: {code}")
        print(f"UserInfo: {user_info}")

        if not code:
            return jsonify(error_response("缺少微信登录code")), 400

        # 获取微信openid
        wechat_data = wechat_login(code)
        if not wechat_data:
            return jsonify(error_response("获取微信信息失败")), 400

        openid = wechat_data['openid']
        print(f"获取到OpenID: {openid}")

        # 创建或更新用户
        user_id = create_or_update_wechat_user(openid, user_info)
        if not user_id:
            return jsonify(error_response("用户创建失败")), 500

        # 获取完整用户信息
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute("SELECT * FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        print(f"微信登录成功，用户ID: {user_id}")

        return jsonify(success_response({
            'user': user,
            'openid': openid,
            'message': '微信登录成功'
        }))

    except Exception as e:
        print(f"微信登录异常: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"微信登录失败: {str(e)}")), 500


@app.route('/api/auth/password-login', methods=['POST'])
def password_login():
    """账号密码登录接口"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')

        print(f"登录请求 - 用户名: {username}, 密码: {password}")  # 调试日志

        if not username or not password:
            return jsonify(error_response("用户名和密码不能为空")), 400

        # 从数据库验证用户名密码
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM users WHERE username = %s AND password = %s",
                       (username, password))
        user = cursor.fetchone()  # 因为用了DictCursor，直接就是字典！

        print(f"数据库查询结果: {user}")  # 调试日志

        if not user:
            cursor.close()
            conn.close()
            print("用户名或密码错误")  # 调试日志
            return jsonify(error_response("用户名或密码错误")), 401

        # 转换日期和Decimal类型
        from decimal import Decimal

        # 处理日期字段
        date_fields = ['created_at', 'annual_insurance_start', 'annual_insurance_end']
        for field in date_fields:
            if field in user and user[field] is not None:
                if hasattr(user[field], 'strftime'):
                    user[field] = user[field].strftime('%Y-%m-%d %H:%M:%S')

        cursor.close()
        conn.close()

        print(f"登录成功: {user['username']}")  # 调试日志

        return jsonify(success_response({
            'user': user,
            'message': '登录成功'
        }))

    except Exception as e:
        print(f"!!!!! 登录接口异常 !!!!!")  # 调试日志
        print(f"异常类型: {type(e)}")
        print(f"异常信息: {str(e)}")

        import traceback
        traceback.print_exc()

        return jsonify(error_response(f"登录失败: {str(e)}")), 500


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """退出登录"""
    return jsonify(success_response(None, "退出成功"))


# ==================== 新增接口 ====================

@app.route('/api/admin/users/<int:user_id>', methods=['GET'])
@admin_or_teacher_required
def admin_get_user_detail(user_id):
    """管理员获取单个用户详情"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()


        cursor.execute("SELECT * FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()


        if not user:
            cursor.close()
            conn.close()
            return jsonify(error_response("用户不存在")), 404


        if user.get('has_annual_insurance') and user.get('annual_insurance_end'):
            end_date = user['annual_insurance_end']
            if hasattr(end_date, 'date'):
                end_date = end_date.date()
            elif isinstance(end_date, str):
                try:
                    end_date = datetime.strptime(end_date.split(' ')[0], '%Y-%m-%d').date()
                except:
                    end_date = None
            if end_date:
                user['insurance_valid'] = end_date >= datetime.now().date()
            else:
                user['insurance_valid'] = False
        else:
            user['insurance_valid'] = False


        cursor.execute("""
            SELECT a.activity_start, a.activity_end, a.title
            FROM user_activities ua
            JOIN activities a ON ua.activity_id = a.id
            WHERE ua.user_id = %s AND ua.status IN ('approved', 'completed')
            AND a.activity_end >= NOW()
            ORDER BY a.activity_start ASC LIMIT 1
        """, (user_id,))
        ongoing_activity = cursor.fetchone()


        if ongoing_activity:
            user['has_ongoing_activity'] = True
            user['ongoing_activity_title'] = ongoing_activity['title']
            user['ongoing_activity_start'] = format_datetime_to_chinese(ongoing_activity['activity_start'])
            user['ongoing_activity_end'] = format_datetime_to_chinese(ongoing_activity['activity_end'])
            if not user['insurance_valid']:
                user['has_temporary_insurance'] = True
        else:
            user['has_ongoing_activity'] = False
            user['has_temporary_insurance'] = False


        if user.get('created_at'):
            user['created_at_cn'] = format_datetime_to_chinese(user['created_at'])
        if user.get('annual_insurance_start'):
            user['annual_insurance_start_cn'] = format_datetime_to_chinese(user['annual_insurance_start'])
        if user.get('annual_insurance_end'):
            user['annual_insurance_end_cn'] = format_datetime_to_chinese(user['annual_insurance_end'])


        # 活动历史
        cursor.execute("""
            SELECT ua.*, a.title as activity_title, a.activity_start, a.activity_end
            FROM user_activities ua
            JOIN activities a ON ua.activity_id = a.id
            WHERE ua.user_id = %s ORDER BY ua.registration_date DESC LIMIT 20
        """, (user_id,))
        user_activities = cursor.fetchall()
        from decimal import Decimal
        from datetime import timedelta as td_type, date as date_type
        for ua in user_activities:
            for key, val in list(ua.items()):
                if isinstance(val, Decimal):
                    ua[key] = float(val)
                elif hasattr(val, 'strftime'):
                    ua[key] = val.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(val, td_type):
                    total_secs = int(val.total_seconds())
                    ua[key] = f"{total_secs // 3600:02d}:{(total_secs % 3600) // 60:02d}"
        user['activity_history'] = user_activities


        # 课程历史
        cursor.execute("""
            SELECT tcb.*, tc.title as course_title, tc.course_type,
                   tcs.schedule_date, tcs.start_time, tcs.end_time,
                   t.real_name as teacher_name
            FROM teacher_course_bookings tcb
            JOIN teacher_courses tc ON tcb.course_id = tc.id
            LEFT JOIN teacher_course_schedules tcs ON tcb.schedule_id = tcs.id
            LEFT JOIN users t ON tcb.teacher_id = t.id
            WHERE tcb.user_id = %s ORDER BY tcb.booking_date DESC LIMIT 20
        """, (user_id,))
        user_courses = cursor.fetchall()
        for uc in user_courses:
            for key, val in list(uc.items()):
                if isinstance(val, Decimal):
                    uc[key] = float(val)
                elif hasattr(val, 'strftime'):
                    uc[key] = val.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(val, td_type):
                    total_secs = int(val.total_seconds())
                    uc[key] = f"{total_secs // 3600:02d}:{(total_secs % 3600) // 60:02d}"
        user['course_history'] = user_courses


        # 订单历史
        cursor.execute("""
            SELECT * FROM orders WHERE user_id = %s ORDER BY created_at DESC LIMIT 20
        """, (user_id,))
        user_orders = cursor.fetchall()
        for order in user_orders:
            for key, val in list(order.items()):
                if isinstance(val, Decimal):
                    order[key] = float(val)
                elif hasattr(val, 'strftime'):
                    order[key] = val.strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("""
                SELECT oi.*, p.name, p.image_url FROM order_items oi
                JOIN products p ON oi.product_id = p.id WHERE oi.order_id = %s
            """, (order['id'],))
            order['items'] = cursor.fetchall()
            for item in order['items']:
                if isinstance(item.get('unit_price'), Decimal):
                    item['unit_price'] = float(item['unit_price'])
            if order['items']:
                order['product_name'] = order['items'][0].get('name', '')
            else:
                order['product_name'] = '订单#' + str(order['id'])
        user['order_history'] = user_orders


        # 会员卡 + 核销历史 + 期限卡状态计算
        cursor.execute("""
            SELECT umc.*, mc.card_name, mc.card_image, mc.card_type as mc_card_type,
                   mc.valid_days as mc_valid_days
            FROM user_membership_cards umc
            LEFT JOIN membership_cards mc ON umc.card_id = mc.id
            WHERE umc.user_id = %s ORDER BY umc.purchase_date DESC
        """, (user_id,))
        user_cards = cursor.fetchall()


        today = date_type.today()
        for uc in user_cards:
            for key, val in list(uc.items()):
                if isinstance(val, Decimal):
                    uc[key] = float(val)
                elif isinstance(val, date_type) and not isinstance(val, datetime):
                    uc[key] = val.strftime('%Y-%m-%d')
                elif hasattr(val, 'strftime'):
                    uc[key] = val.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(val, td_type):
                    total_secs = int(val.total_seconds())
                    uc[key] = f"{total_secs // 3600:02d}:{(total_secs % 3600) // 60:02d}"


            # 计算期限卡状态和剩余天数
            if uc.get('card_type') == 'period':
                total_days = uc.get('mc_valid_days') or uc.get('valid_days') or 0
                uc['total_days'] = total_days


                if not uc.get('activated') or uc['activated'] == 0:
                    uc['period_status'] = '未激活'
                    uc['remaining_days'] = total_days
                elif uc.get('status') == 'expired':
                    uc['period_status'] = '已过期'
                    uc['remaining_days'] = 0
                else:
                    end_date_str = uc.get('end_date', '')
                    if end_date_str:
                        try:
                            ed = datetime.strptime(str(end_date_str)[:10], '%Y-%m-%d').date()
                        except:
                            ed = None
                    else:
                        ed = None


                    if ed and ed >= today:
                        uc['period_status'] = '已激活生效中'
                        uc['remaining_days'] = (ed - today).days
                    else:
                        uc['period_status'] = '已过期'
                        uc['remaining_days'] = 0
            else:
                uc['period_status'] = None
                uc['remaining_days'] = None
                uc['total_days'] = None


            # 核销记录
            cursor.execute("""
                SELECT * FROM membership_consume_logs
                WHERE user_membership_card_id = %s ORDER BY created_at DESC LIMIT 50
            """, (uc['id'],))
            logs = cursor.fetchall()
            for lg in logs:
                for key, val in list(lg.items()):
                    if hasattr(val, 'strftime'):
                        lg[key] = val.strftime('%Y-%m-%d %H:%M:%S')
            uc['consume_logs'] = logs


        user['membership_cards_detail'] = user_cards


        cursor.close()
        conn.close()


        return jsonify(success_response(user))
    except Exception as e:
        return jsonify(error_response(f"获取用户详情失败: {str(e)}")), 500




@app.route('/api/admin/activities/<int:activity_id>/export-insurance', methods=['GET'])
@admin_or_teacher_required
def export_insurance(activity_id):
    """导出活动所有参与人员的保险信息为Excel"""
    print(f"\n===== 开始导出活动 {activity_id} 的保险信息 =====")

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取活动信息
        cursor.execute("SELECT title FROM activities WHERE id = %s", (activity_id,))
        activity = cursor.fetchone()

        if not activity:
            cursor.close()
            conn.close()
            return jsonify(error_response("活动不存在")), 404

        activity_title = activity['title']
        print(f"活动标题: {activity_title}")

        # 获取已通过审核的所有报名用户的完整信息
        cursor.execute("""
            SELECT u.real_name, u.id_card, u.gender, u.age, u.phone, u.nation,
                   u.blood_type, u.emergency_contact_name, u.emergency_contact_phone,
                   u.allergy_history, u.contraindications,
                   u.has_annual_insurance, u.annual_insurance_start, u.annual_insurance_end,
                   ua.total_amount, ua.skip_insurance
            FROM user_activities ua
            JOIN users u ON ua.user_id = u.id
            WHERE ua.activity_id = %s AND ua.status = 'approved'
            ORDER BY ua.registration_date
        """, (activity_id,))

        participants = cursor.fetchall()
        print(f"查询到 {len(participants)} 个参与人员")

        if len(participants) > 0:
            print("第一个参与者数据:", participants[0])

        cursor.close()
        conn.close()

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "参与人员保险信息"

        # 设置表头
        headers = [
            '姓名', '身份证号', '性别', '年龄', '手机号', '民族',
            '血型', '紧急联系人', '紧急联系电话', '过敏史', '禁忌症',
            '年险状态', '年险开始', '年险结束', '报名费用', '跳过保险费'
        ]
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color="1296DB", end_color="1296DB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 填充数据
        for participant in participants:
            # 处理日期字段
            annual_start = participant.get('annual_insurance_start')
            if annual_start:
                if hasattr(annual_start, 'strftime'):
                    annual_start = annual_start.strftime('%Y-%m-%d %H:%M')
                else:
                    annual_start = str(annual_start)
            else:
                annual_start = ''

            annual_end = participant.get('annual_insurance_end')
            if annual_end:
                if hasattr(annual_end, 'strftime'):
                    annual_end = annual_end.strftime('%Y-%m-%d %H:%M')
                else:
                    annual_end = str(annual_end)
            else:
                annual_end = ''

            row = [
                participant.get('real_name', ''),
                participant.get('id_card', ''),
                participant.get('gender', ''),
                str(participant.get('age', '')),
                participant.get('phone', ''),
                participant.get('nation', ''),
                participant.get('blood_type', ''),
                participant.get('emergency_contact_name', ''),
                participant.get('emergency_contact_phone', ''),
                participant.get('allergy_history', ''),
                participant.get('contraindications', ''),
                '是' if participant.get('has_annual_insurance') else '否',
                annual_start,
                annual_end,
                str(participant.get('total_amount', '')),
                '是' if participant.get('skip_insurance') else '否'
            ]
            ws.append(row)
            print(f"添加行数据: {row[0]}")

        # 调整列宽
        column_widths = [12, 20, 8, 8, 15, 10, 10, 12, 15, 20, 20, 12, 18, 18, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # 保存文件
        filename = f"{activity_title}-全部人员保险信息.xlsx"
        file_path = os.path.join(Config.UPLOAD_FOLDER, 'exports', filename)

        os.makedirs(os.path.join(Config.UPLOAD_FOLDER, 'exports'), exist_ok=True)

        wb.save(file_path)

        print(f"文件保存成功: {file_path}")
        print(f"文件大小: {os.path.getsize(file_path)} bytes")

        file_url = f"{Config.SERVER_HOST}/uploads/exports/{filename}"

        return jsonify(success_response({
            'file_url': file_url,
            'filename': filename
        }, "导出成功"))

    except Exception as e:
        print(f"\n!!!!! 导出失败 !!!!!")
        print(f"异常信息: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"导出失败: {str(e)}")), 500


@app.route('/api/admin/activities/<int:activity_id>/export-no-insurance', methods=['GET'])
@admin_or_teacher_required
def export_no_insurance(activity_id):
    """导出活动中未购买保险的人员名单"""
    print(f"\n===== 开始导出活动 {activity_id} 的未购保险人员 =====")

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取活动信息
        cursor.execute("SELECT title FROM activities WHERE id = %s", (activity_id,))
        activity = cursor.fetchone()

        if not activity:
            cursor.close()
            conn.close()
            return jsonify(error_response("活动不存在")), 404

        activity_title = activity['title']
        print(f"活动标题: {activity_title}")

        # 获取已通过但没有年险或跳过保险费的用户
        cursor.execute("""
            SELECT u.real_name, u.id_card, u.gender, u.age, u.phone, u.nation,
                   u.blood_type, u.emergency_contact_name, u.emergency_contact_phone,
                   u.allergy_history, u.contraindications,
                   ua.skip_insurance, ua.total_amount
            FROM user_activities ua
            JOIN users u ON ua.user_id = u.id
            WHERE ua.activity_id = %s 
            AND ua.status = 'approved'
            AND (
                u.has_annual_insurance = FALSE 
                OR u.has_annual_insurance IS NULL 
                OR ua.skip_insurance = TRUE
                OR u.annual_insurance_end IS NULL
                OR u.annual_insurance_end < NOW()
            )
            ORDER BY ua.registration_date
        """, (activity_id,))

        participants = cursor.fetchall()
        print(f"查询到 {len(participants)} 个未购保险人员")

        if len(participants) > 0:
            print("第一个未购保险人员:", participants[0])

        cursor.close()
        conn.close()

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "未购保险人员"

        # 设置表头
        headers = [
            '姓名', '身份证号', '性别', '年龄', '手机号', '民族',
            '血型', '紧急联系人', '紧急联系电话', '过敏史', '禁忌症',
            '跳过保险费', '报名费用'
        ]
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 填充数据
        for participant in participants:
            row = [
                participant.get('real_name', ''),
                participant.get('id_card', ''),
                participant.get('gender', ''),
                str(participant.get('age', '')),
                participant.get('phone', ''),
                participant.get('nation', ''),
                participant.get('blood_type', ''),
                participant.get('emergency_contact_name', ''),
                participant.get('emergency_contact_phone', ''),
                participant.get('allergy_history', ''),
                participant.get('contraindications', ''),
                '是' if participant.get('skip_insurance') else '否',
                str(participant.get('total_amount', ''))
            ]
            ws.append(row)
            print(f"添加行数据: {row[0]}")

        # 调整列宽
        column_widths = [12, 20, 8, 8, 15, 10, 10, 12, 15, 20, 20, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # 保存文件
        filename = f"{activity_title}-未购保险人员.xlsx"
        file_path = os.path.join(Config.UPLOAD_FOLDER, 'exports', filename)

        os.makedirs(os.path.join(Config.UPLOAD_FOLDER, 'exports'), exist_ok=True)

        wb.save(file_path)

        print(f"文件保存成功: {file_path}")
        print(f"文件大小: {os.path.getsize(file_path)} bytes")

        file_url = f"{Config.SERVER_HOST}/uploads/exports/{filename}"

        return jsonify(success_response({
            'file_url': file_url,
            'filename': filename
        }, "导出成功"))

    except Exception as e:
        print(f"\n!!!!! 导出失败 !!!!!")
        print(f"异常信息: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"导出失败: {str(e)}")), 500

# ==================== 收货地址管理 ====================

@app.route('/api/addresses', methods=['GET'])
@login_required
def get_user_addresses():
    """获取用户收货地址列表"""
    try:
        user_id = request.current_user['id']

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM user_addresses 
            WHERE user_id = %s 
            ORDER BY is_default DESC, created_at DESC
        """, (user_id,))

        addresses = cursor.fetchall()
        cursor.close()
        conn.close()

        return jsonify(success_response(addresses))
    except Exception as e:
        return jsonify(error_response(f"获取地址列表失败: {str(e)}")), 500


@app.route('/api/addresses', methods=['POST'])
@login_required
def create_address():
    """创建收货地址"""
    try:
        data = request.get_json()
        user_id = request.current_user['id']

        conn = get_db_connection()
        cursor = conn.cursor()

        # 如果设置为默认地址，先取消其他默认地址
        if data.get('is_default', False):
            cursor.execute("""
                UPDATE user_addresses SET is_default = FALSE WHERE user_id = %s
            """, (user_id,))

        cursor.execute("""
            INSERT INTO user_addresses 
            (user_id, receiver_name, receiver_phone, province, city, district, detail_address, is_default)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            user_id,
            data['receiver_name'],
            data['receiver_phone'],
            data['province'],
            data['city'],
            data['district'],
            data['detail_address'],
            data.get('is_default', False)
        ))

        address_id = cursor.lastrowid

        conn.commit()

        # ===== 新增：记录操作日志 =====
        try:
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='添加地址',
                operation_module='地址管理',
                operation_desc=f"添加收货地址: {data['receiver_name']} {data['province']}{data['city']}",
                request_params=data
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response({'address_id': address_id}, "地址添加成功"))
    except Exception as e:
        return jsonify(error_response(f"添加地址失败: {str(e)}")), 500


@app.route('/api/addresses/<int:address_id>', methods=['PUT'])
@login_required
def update_address(address_id):
    """更新收货地址"""
    try:
        data = request.get_json()
        user_id = request.current_user['id']

        conn = get_db_connection()
        cursor = conn.cursor()

        # 如果设置为默认地址，先取消其他默认地址
        if data.get('is_default', False):
            cursor.execute("""
                UPDATE user_addresses SET is_default = FALSE WHERE user_id = %s
            """, (user_id,))

        updatable_fields = ['receiver_name', 'receiver_phone', 'province', 'city', 'district', 'detail_address',
                            'is_default']
        update_fields = []
        params = []

        for field in updatable_fields:
            if field in data:
                update_fields.append(f"{field} = %s")
                params.append(data[field])

        params.extend([address_id, user_id])

        sql = f"UPDATE user_addresses SET {', '.join(update_fields)} WHERE id = %s AND user_id = %s"
        cursor.execute(sql, params)

        conn.commit()

        # ===== 新增：记录操作日志 =====
        try:
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='修改地址',
                operation_module='地址管理',
                operation_desc=f"修改地址ID: {address_id}",
                request_params=data
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "地址更新成功"))
    except Exception as e:
        return jsonify(error_response(f"更新地址失败: {str(e)}")), 500


@app.route('/api/addresses/<int:address_id>', methods=['DELETE'])
@login_required
def delete_address(address_id):
    """删除收货地址"""
    try:
        user_id = request.current_user['id']

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM user_addresses WHERE id = %s AND user_id = %s", (address_id, user_id))

        conn.commit()

        # ===== 新增：记录操作日志 =====
        try:
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='删除地址',
                operation_module='地址管理',
                operation_desc=f"删除地址ID: {address_id}",
                request_params={'address_id': address_id}
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "地址删除成功"))
    except Exception as e:
        return jsonify(error_response(f"删除地址失败: {str(e)}")), 500


# ==================== 保险凭证审核 ====================

@app.route('/api/insurance-submissions', methods=['POST'])
@login_required
def submit_insurance():
    """提交保险凭证"""
    try:
        data = request.get_json()
        user_id = request.current_user['id']

        # 验证日期时间格式
        start_datetime = data.get('start_date')  # 格式：2024-03-01 08:00:00
        end_datetime = data.get('end_date')

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO insurance_submissions 
            (user_id, certificate_image_url, start_date, end_date, status)
            VALUES (%s, %s, %s, %s, 'pending')
        """, (
            user_id,
            data['certificate_image_url'],
            start_datetime,
            end_datetime
        ))

        submission_id = cursor.lastrowid

        conn.commit()

        # ===== 新增：记录日志和通知管理员 =====
        try:
            # 记录操作日志
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='提交保险凭证',
                operation_module='保险管理',
                operation_desc=f"提交保险凭证，保障期间:{start_datetime} ~ {end_datetime}",
                request_params=data
            )

            # 通知管理员
            Notifier.notify_admin_new_insurance(
                user_name=request.current_user.get('real_name', '用户')
            )
        except Exception as e:
            print(f"日志/通知记录失败（不影响主流程）: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response({'submission_id': submission_id}, "保险凭证已提交，等待审核"))
    except Exception as e:
        return jsonify(error_response(f"提交失败: {str(e)}")), 500


@app.route('/api/insurance-submissions', methods=['GET'])
@login_required
def get_insurance_submissions():
    """获取用户的保险凭证提交记录"""
    try:
        user_id = request.current_user['id']

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM insurance_submissions 
            WHERE user_id = %s 
            ORDER BY created_at DESC
        """, (user_id,))

        submissions = cursor.fetchall()
        cursor.close()
        conn.close()

        return jsonify(success_response(submissions))
    except Exception as e:
        return jsonify(error_response(f"获取记录失败: {str(e)}")), 500


@app.route('/api/admin/insurance-submissions', methods=['GET'])
@admin_required
def admin_get_insurance_submissions():
    """管理员获取待审核的保险凭证"""
    try:
        status = request.args.get('status', 'pending')

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT s.*, u.real_name, u.phone 
            FROM insurance_submissions s
            JOIN users u ON s.user_id = u.id
            WHERE s.status = %s
            ORDER BY s.created_at DESC
        """, (status,))

        submissions = cursor.fetchall()

        # 格式化日期为中文
        for submission in submissions:
            if submission.get('created_at'):
                submission['created_at_cn'] = format_datetime_to_chinese(submission['created_at'])
            if submission.get('start_date'):
                submission['start_date_cn'] = format_datetime_to_chinese(submission['start_date'])
            if submission.get('end_date'):
                submission['end_date_cn'] = format_datetime_to_chinese(submission['end_date'])

        cursor.close()
        conn.close()

        return jsonify(success_response(submissions))
    except Exception as e:
        return jsonify(error_response(f"获取审核列表失败: {str(e)}")), 500


@app.route('/api/admin/insurance-submissions/<int:submission_id>/review', methods=['PUT'])
@admin_required
def review_insurance_submission(submission_id):
    """审核保险凭证"""
    try:
        data = request.get_json()
        status = data.get('status')
        admin_notes = data.get('admin_notes', '')

        if status not in ['approved', 'rejected']:
            return jsonify(error_response("无效的审核状态")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM insurance_submissions WHERE id = %s", (submission_id,))
        submission = cursor.fetchone()

        if not submission:
            cursor.close()
            conn.close()
            return jsonify(error_response("记录不存在")), 404

        cursor.execute("""
            UPDATE insurance_submissions 
            SET status = %s, admin_notes = %s, reviewed_at = NOW()
            WHERE id = %s
        """, (status, admin_notes, submission_id))

        # 如果审核通过，更新用户的年险信息（datetime格式）
        if status == 'approved':
            cursor.execute("""
                UPDATE users 
                SET has_annual_insurance = TRUE,
                    annual_insurance_start = %s,
                    annual_insurance_end = %s
                WHERE id = %s
            """, (submission['start_date'], submission['end_date'], submission['user_id']))

        conn.commit()

        # ===== 新增：记录审计日志和通知用户 =====
        try:
            # 记录审计日志
            Logger.log_audit(
                user_id=request.current_user['id'],
                user_name=request.current_user.get('real_name', '管理员'),
                action='审核保险凭证',
                target_type='insurance_submissions',
                target_id=submission_id,
                old_value={'status': 'pending'},
                new_value={'status': status, 'admin_notes': admin_notes}
            )

            # 如果审核通过，通知用户
            if status == 'approved':
                Notifier.send_insurance_approved(user_id=submission['user_id'])
        except Exception as e:
            print(f"日志/通知记录失败（不影响主流程）: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "审核完成"))
    except Exception as e:
        return jsonify(error_response(f"审核失败: {str(e)}")), 500


# ==================== 微信支付 ====================

@app.route('/api/payment/create-order', methods=['POST'])
@login_required
def create_payment_order():
    """创建支付订单（活动报名、课程报名或商品购买）"""
    conn = None
    cursor = None
    try:
        data = request.get_json()
        order_type = data.get('type')  # 'activity', 'course', 'product'
        order_id = data.get('order_id')
        total_amount = float(data.get('total_amount', 0))
        user_id = request.current_user['id']
        openid = request.current_user.get('openid', '')

        if order_type not in ('activity', 'course', 'product', 'membership'):

            return jsonify(error_response("无效的订单类型")), 400
        if not order_id:
            return jsonify(error_response("缺少订单ID")), 400

        # 微信支付金额单位是分
        total_fee = int(round(total_amount * 100))
        if total_fee <= 0:
            return jsonify(error_response("支付金额无效")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 根据类型验证订单并生成 out_trade_no
        timestamp = str(int(time.time()))

        if order_type == 'activity':
            cursor.execute(
                "SELECT * FROM user_activities WHERE id = %s AND user_id = %s",
                (order_id, user_id)
            )
            record = cursor.fetchone()
            if not record:
                return jsonify(error_response("报名记录不存在")), 404
            if record['payment_status'] != 'unpaid':
                return jsonify(error_response("该订单已支付或已退款")), 400
            if record['status'] not in ('approved',):
                return jsonify(error_response("报名尚未审核通过，无法支付")), 400
            out_trade_no = f"ACT{order_id}_{timestamp}"
            description = '活动报名'
            cursor.execute(
                "UPDATE user_activities SET out_trade_no = %s WHERE id = %s",
                (out_trade_no, order_id)
            )

        elif order_type == 'course':
            cursor.execute(
                "SELECT * FROM teacher_course_bookings WHERE id = %s AND user_id = %s",
                (order_id, user_id)
            )
            record = cursor.fetchone()
            if not record:
                return jsonify(error_response("报名记录不存在")), 404
            if record['payment_status'] != 'unpaid':
                return jsonify(error_response("该订单已支付或已退款")), 400
            if record['status'] not in ('approved',):
                return jsonify(error_response("报名尚未审核通过，无法支付")), 400
            out_trade_no = f"CRS{order_id}_{timestamp}"
            description = '课程报名'
            cursor.execute(
                "UPDATE teacher_course_bookings SET out_trade_no = %s WHERE id = %s",
                (out_trade_no, order_id)
            )

        elif order_type == 'membership':
            cursor.execute(
                "SELECT * FROM user_membership_cards WHERE id = %s AND user_id = %s",
                (order_id, user_id)
            )
            record = cursor.fetchone()
            if not record:
                return jsonify(error_response("会员卡记录不存在")), 404
            if record['payment_status'] != 'unpaid':
                return jsonify(error_response("该订单已支付或已退款")), 400
            out_trade_no = f"MBR{order_id}_{timestamp}"
            description = f'会员卡购买'
            cursor.execute(
                "UPDATE user_membership_cards SET out_trade_no = %s WHERE id = %s",
                (out_trade_no, order_id)
            )

        else:  # product
            cursor.execute(
                "SELECT * FROM orders WHERE id = %s AND user_id = %s",
                (order_id, user_id)
            )
            record = cursor.fetchone()
            if not record:
                return jsonify(error_response("订单不存在")), 404
            if record['payment_status'] != 'unpaid':
                return jsonify(error_response("该订单已支付或已退款")), 400
            out_trade_no = record['order_no']
            description = '商品购买'
            cursor.execute(
                "UPDATE orders SET out_trade_no = %s WHERE id = %s",
                (out_trade_no, order_id)
            )

        conn.commit()

        # 调用微信支付统一下单
        import wechat_pay
        prepay_id = wechat_pay.create_jsapi_order(
            openid=openid,
            out_trade_no=out_trade_no,
            total_fen=total_fee,
            description=description
        )

        # 生成前端支付参数
        pay_params = wechat_pay.generate_jsapi_sign(prepay_id)

        cursor.close()
        conn.close()
        conn = None

        return jsonify(success_response(pay_params, "预支付订单创建成功"))

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        return jsonify(error_response(f"创建支付订单失败: {str(e)}")), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.route('/api/payment/callback', methods=['POST'])
def payment_callback():
    """微信支付回调"""
    conn = None
    cursor = None
    try:
        body = request.data.decode('utf-8')


        # DEBUG 模式：支持手动模拟回调
        if Config.DEBUG:
            data = json.loads(body) if body else {}
            out_trade_no = data.get('out_trade_no', '')
            transaction_id = data.get('transaction_id', f'mock_txn_{int(time.time())}')
        else:
            # 生产模式：验签 + 解密
            import wechat_pay
            if not wechat_pay.verify_callback_signature(request.headers, body):
                return jsonify({"code": "FAIL", "message": "签名验证失败"}), 400


            callback_data = json.loads(body)
            resource = callback_data.get('resource', {})
            decrypted = wechat_pay.decrypt_callback_data(resource)


            out_trade_no = decrypted.get('out_trade_no', '')
            transaction_id = decrypted.get('transaction_id', '')
            trade_state = decrypted.get('trade_state', '')


            if trade_state != 'SUCCESS':
                return jsonify({"code": "SUCCESS", "message": ""})


        if not out_trade_no:
            return jsonify({"code": "FAIL", "message": "缺少 out_trade_no"}), 400


        conn = get_db_connection()
        cursor = conn.cursor()
        conn.begin()


        # 根据 out_trade_no 前缀判断订单类型
        if out_trade_no.startswith('ACT'):
            # 活动报名支付
            cursor.execute(
                "SELECT * FROM user_activities WHERE out_trade_no = %s FOR UPDATE",
                (out_trade_no,)
            )
            record = cursor.fetchone()
            if not record:
                conn.rollback()
                return jsonify({"code": "FAIL", "message": "订单不存在"}), 404


            # 幂等检查
            if record['payment_status'] == 'paid':
                conn.rollback()
                return jsonify({"code": "SUCCESS", "message": ""})


            # 更新支付状态
            cursor.execute("""
                UPDATE user_activities
                SET payment_status = 'paid', transaction_id = %s, paid_at = NOW()
                WHERE id = %s
            """, (transaction_id, record['id']))


            # 发放积分
            total_amount = float(record['total_amount'])
            add_user_points(record['user_id'], total_amount, 'activity_payment',
                            booking_id=record['id'],
                            description=f'活动报名支付¥{total_amount}')


        elif out_trade_no.startswith('CRS'):
            # 课程报名支付
            cursor.execute(
                "SELECT * FROM teacher_course_bookings WHERE out_trade_no = %s FOR UPDATE",
                (out_trade_no,)
            )
            record = cursor.fetchone()
            if not record:
                conn.rollback()
                return jsonify({"code": "FAIL", "message": "订单不存在"}), 404


            if record['payment_status'] == 'paid':
                conn.rollback()
                return jsonify({"code": "SUCCESS", "message": ""})


            cursor.execute("""
                UPDATE teacher_course_bookings
                SET payment_status = 'paid', paid_at = NOW()
                WHERE id = %s
            """, (record['id'],))


            # 发放积分
            total_amount = float(record['payment_amount'])
            add_user_points(record['user_id'], total_amount, 'course_payment',
                            booking_id=record['id'],
                            description=f'课程报名支付¥{total_amount}')


        elif out_trade_no.startswith('MBR'):
            # 会员卡购买支付
            cursor.execute(
                "SELECT * FROM user_membership_cards WHERE out_trade_no = %s FOR UPDATE",
                (out_trade_no,)
            )
            record = cursor.fetchone()
            if not record:
                conn.rollback()
                return jsonify({"code": "FAIL", "message": "订单不存在"}), 404


            if record['payment_status'] == 'paid':
                conn.rollback()
                return jsonify({"code": "SUCCESS", "message": ""})


            # 次卡：支付后直接激活
            # 期限卡：支付后等待用户手动激活
            activated = 1 if record['card_type'] == 'times' else 0


            cursor.execute("""
                UPDATE user_membership_cards
                SET payment_status = 'paid', paid_at = NOW(),
                    activated = %s
                WHERE id = %s
            """, (activated, record['id']))


            # 发放积分
            total_amount = float(record['purchase_amount'])
            add_user_points(record['user_id'], total_amount, 'membership_payment',
                            description=f'购买会员卡支付¥{total_amount}')


        else:
            # 商品订单支付（out_trade_no = order_no）
            cursor.execute(
                "SELECT * FROM orders WHERE out_trade_no = %s FOR UPDATE",
                (out_trade_no,)
            )
            record = cursor.fetchone()
            if not record:
                conn.rollback()
                return jsonify({"code": "FAIL", "message": "订单不存在"}), 404


            if record['payment_status'] == 'paid':
                conn.rollback()
                return jsonify({"code": "SUCCESS", "message": ""})


            cursor.execute("""
                UPDATE orders
                SET payment_status = 'paid', status = 'paid', transaction_id = %s, paid_at = NOW()
                WHERE id = %s
            """, (transaction_id, record['id']))


            # 发放积分
            total_amount = float(record['total_amount'])
            add_user_points(record['user_id'], total_amount, 'product_payment',
                            order_id=record['id'],
                            description=f'商品购买支付¥{total_amount}')


            # ===== 新增：支付成功后创建配送单并通知教练 =====
            if record.get('delivery_teacher_id'):
                try:
                    # 查询订单中的商品名称
                    cursor.execute("""
                        SELECT oi.*, p.name FROM order_items oi
                        JOIN products p ON oi.product_id = p.id
                        WHERE oi.order_id = %s
                    """, (record['id'],))
                    order_items = cursor.fetchall()
                    product_names = ', '.join([it['name'] for it in order_items])


                    # 创建配送单
                    cursor.execute("""
                        INSERT INTO delivery_orders
                        (order_id, user_id, teacher_id, user_phone, delivery_address, product_name, status)
                        VALUES (%s, %s, %s, %s, %s, %s, 'pending')
                    """, (record['id'], record['user_id'], record['delivery_teacher_id'],
                          record.get('receiver_phone', ''), record.get('shipping_address', ''), product_names))


                    # 查询买家姓名
                    cursor.execute("SELECT real_name FROM users WHERE id = %s", (record['user_id'],))
                    buyer = cursor.fetchone()
                    buyer_name = buyer['real_name'] if buyer else '用户'


                    # 发送站内通知给教练
                    Notifier.send_delivery_notification(
                        teacher_id=record['delivery_teacher_id'],
                        user_name=buyer_name,
                        product_name=product_names,
                        order_no=record.get('order_no', out_trade_no)
                    )
                except Exception as de:
                    print(f"创建配送单/通知失败（不影响支付）: {str(de)}")
                    import traceback
                    traceback.print_exc()


        conn.commit()
        return jsonify({"code": "SUCCESS", "message": ""})


    except Exception as e:
        if conn:
            conn.rollback()
        print(f"支付回调处理失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"code": "FAIL", "message": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



# ==================== 管理员商品管理增强 ====================

@app.route('/api/admin/products/<int:product_id>', methods=['PUT'])
@admin_required
def admin_update_product(product_id):
    """管理员更新商品"""
    try:
        data = request.get_json()

        updatable_fields = ['name', 'description', 'price', 'image_url', 'stock', 'category', 'status']
        update_fields = []
        params = []

        for field in updatable_fields:
            if field in data:
                update_fields.append(f"{field} = %s")
                params.append(data[field])

        if not update_fields:
            return jsonify(error_response("没有可更新的字段")), 400

        params.append(product_id)

        conn = get_db_connection()
        cursor = conn.cursor()

        sql = f"UPDATE products SET {', '.join(update_fields)} WHERE id = %s"
        cursor.execute(sql, params)

        conn.commit()

        # ===== 新增：记录审计日志 =====
        try:
            Logger.log_audit(
                user_id=request.current_user['id'],
                user_name=request.current_user.get('real_name', '管理员'),
                action='修改商品',
                target_type='products',
                target_id=product_id,
                old_value=None,
                new_value=data
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "商品更新成功"))
    except Exception as e:
        return jsonify(error_response(f"更新商品失败: {str(e)}")), 500


@app.route('/api/admin/products/<int:product_id>', methods=['DELETE'])
@admin_required
def admin_delete_product(product_id):
    """管理员删除商品"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 软删除：更改状态为inactive
        cursor.execute("UPDATE products SET status = 'inactive' WHERE id = %s", (product_id,))

        conn.commit()

        # ===== 新增：记录操作日志 =====
        try:
            Logger.log_operation(
                user_id=request.current_user['id'],
                user_name=request.current_user.get('real_name', '管理员'),
                operation_type='下架商品',
                operation_module='商品管理',
                operation_desc=f"下架商品ID: {product_id}",
                request_params={'product_id': product_id}
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "商品已下架"))
    except Exception as e:
        return jsonify(error_response(f"删除商品失败: {str(e)}")), 500


def format_date_to_chinese(date_obj):
    """格式化日期为中文：2024年03月01日"""
    if not date_obj:
        return ''

    try:
        if isinstance(date_obj, str):
            # 如果是字符串，先转换为日期对象
            if ' ' in date_obj:
                date_obj = datetime.strptime(date_obj.split(' ')[0], '%Y-%m-%d')
            else:
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d')

        if hasattr(date_obj, 'strftime'):
            return date_obj.strftime('%Y年%m月%d日')

        return str(date_obj)
    except:
        return str(date_obj) if date_obj else ''


def format_datetime_to_chinese(datetime_obj):
    """格式化日期时间为中文：2024年03月01日 08:00"""
    if not datetime_obj:
        return ''

    try:
        if isinstance(datetime_obj, str):
            if len(datetime_obj) > 10:
                datetime_obj = datetime.strptime(datetime_obj[:19], '%Y-%m-%d %H:%M:%S')
            else:
                datetime_obj = datetime.strptime(datetime_obj, '%Y-%m-%d')

        if hasattr(datetime_obj, 'strftime'):
            return datetime_obj.strftime('%Y年%m月%d日 %H:%M')

        return str(datetime_obj)
    except:
        return str(datetime_obj) if datetime_obj else ''


# ==================== 退款相关接口 ====================

@app.route('/api/users/activities/<int:registration_id>/refund', methods=['POST'])
@login_required
def request_refund(registration_id):
    """用户申请退款"""
    try:
        data = request.get_json()
        refund_reason = data.get('refund_reason', '')
        user_id = request.current_user['id']

        if not refund_reason:
            return jsonify(error_response("请输入退款原因")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查报名记录
        cursor.execute("""
            SELECT * FROM user_activities 
            WHERE id = %s AND user_id = %s AND status = 'approved'
        """, (registration_id, user_id))

        registration = cursor.fetchone()
        if not registration:
            cursor.close()
            conn.close()
            return jsonify(error_response("报名记录不存在或不可退款")), 404

        if registration['payment_status'] != 'paid':
            cursor.close()
            conn.close()
            return jsonify(error_response("该报名未支付，无需退款")), 400

        # 更新退款状态
        cursor.execute("""
            UPDATE user_activities 
            SET refund_reason = %s, refund_status = 'pending', refund_request_time = NOW()
            WHERE id = %s
        """, (refund_reason, registration_id))

        conn.commit()
        # ===== 新增：记录日志和通知管理员 =====
        try:
            # 记录操作日志
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='申请退款',
                operation_module='退款管理',
                operation_desc=f"申请退款，原因:{refund_reason}",
                request_params={'registration_id': registration_id, 'reason': refund_reason}
            )

            # 获取活动信息
            cursor_temp = conn.cursor()
            cursor_temp.execute("""
                SELECT a.title, ua.total_amount
                FROM user_activities ua
                JOIN activities a ON ua.activity_id = a.id
                WHERE ua.id = %s
            """, (registration_id,))
            activity_info = cursor_temp.fetchone()
            cursor_temp.close()

            if activity_info:
                # 通知管理员
                Notifier.notify_admin_new_refund(
                    activity_title=activity_info['title'],
                    user_name=request.current_user.get('real_name', '用户'),
                    amount=activity_info['total_amount']
                )
        except Exception as e:
            print(f"日志/通知记录失败（不影响主流程）: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "退款申请已提交，等待审核"))
    except Exception as e:
        return jsonify(error_response(f"申请退款失败: {str(e)}")), 500


@app.route('/api/course-booking/<int:booking_id>/refund', methods=['POST'])
@login_required
def request_course_refund(booking_id):
    """用户申请课程退款"""
    try:
        data = request.get_json()
        refund_reason = data.get('refund_reason', '')
        user_id = request.current_user['id']

        if not refund_reason:
            return jsonify(error_response("请输入退款原因")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM teacher_course_bookings
            WHERE id = %s AND user_id = %s AND status = 'approved'
        """, (booking_id, user_id))

        booking = cursor.fetchone()
        if not booking:
            cursor.close()
            conn.close()
            return jsonify(error_response("课程记录不存在或不可退款")), 404

        if booking['payment_status'] != 'paid':
            cursor.close()
            conn.close()
            return jsonify(error_response("该课程未支付，无需退款")), 400

        cursor.execute("""
            UPDATE teacher_course_bookings
            SET refund_reason = %s, refund_status = 'pending', refund_request_time = NOW()
            WHERE id = %s
        """, (refund_reason, booking_id))

        conn.commit()

        try:
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='申请课程退款',
                operation_module='退款管理',
                operation_desc=f"申请课程退款，原因:{refund_reason}",
                request_params={'booking_id': booking_id, 'reason': refund_reason}
            )
        except Exception as e:
            print(f"日志记录失败（不影响主流程）: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "退款申请已提交，等待审核"))
    except Exception as e:
        return jsonify(error_response(f"申请退款失败: {str(e)}")), 500


@app.route('/api/order/<int:order_id>/refund', methods=['POST'])
@login_required
def request_order_refund(order_id):
    """商品订单退款申请（用户侧）"""
    try:
        data = request.get_json()
        refund_reason = data.get('refund_reason', '')
        user_id = request.current_user['id']

        if not refund_reason:
            return jsonify(error_response("请输入退款原因")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM orders
            WHERE id = %s AND user_id = %s
        """, (order_id, user_id))

        order = cursor.fetchone()
        if not order:
            cursor.close()
            conn.close()
            return jsonify(error_response("订单不存在")), 404

        if order['payment_status'] != 'paid':
            cursor.close()
            conn.close()
            return jsonify(error_response("该订单未支付，无需退款")), 400

        if order.get('refund_status') and order['refund_status'] != 'none':
            cursor.close()
            conn.close()
            return jsonify(error_response("该订单已申请过退款")), 400

        cursor.execute("""
            UPDATE orders
            SET refund_reason = %s, refund_status = 'pending', refund_request_time = NOW()
            WHERE id = %s
        """, (refund_reason, order_id))

        conn.commit()

        try:
            Logger.log_operation(
                user_id=user_id,
                user_name=request.current_user.get('real_name', '用户'),
                operation_type='申请退款',
                operation_module='退款管理',
                operation_desc=f"商品订单退款申请，订单号:{order.get('order_no', '')}, 原因:{refund_reason}",
                request_params={'order_id': order_id, 'reason': refund_reason}
            )

            Notifier.notify_admin_new_refund(
                activity_title=f"商品订单 {order.get('order_no', '')}",
                user_name=request.current_user.get('real_name', '用户'),
                amount=order.get('total_amount', 0)
            )
        except Exception as e:
            print(f"日志/通知记录失败（不影响主流程）: {str(e)}")

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "退款申请已提交，等待审核"))
    except Exception as e:
        return jsonify(error_response(f"申请退款失败: {str(e)}")), 500


@app.route('/api/admin/refunds', methods=['GET'])
@admin_required
def admin_get_refunds():
    """管理员获取退款申请列表，支持 type 参数区分活动/课程/商品"""
    try:
        status = request.args.get('status', 'pending')
        refund_type = request.args.get('type', 'activity')

        conn = get_db_connection()
        cursor = conn.cursor()

        if refund_type == 'course':
            cursor.execute("""
                SELECT tcb.*, u.real_name, u.phone, tc.title as course_title
                FROM teacher_course_bookings tcb
                JOIN users u ON tcb.user_id = u.id
                JOIN teacher_courses tc ON tcb.course_id = tc.id
                WHERE tcb.refund_status = %s
                ORDER BY tcb.refund_request_time DESC
            """, (status,))
        elif refund_type == 'product':
            cursor.execute("""
                SELECT o.*, u.real_name, u.phone
                FROM orders o
                JOIN users u ON o.user_id = u.id
                WHERE o.refund_status = %s
                ORDER BY o.refund_request_time DESC
            """, (status,))
        else:
            cursor.execute("""
                SELECT ua.*, u.real_name, u.phone, a.title as activity_title
                FROM user_activities ua
                JOIN users u ON ua.user_id = u.id
                JOIN activities a ON ua.activity_id = a.id
                WHERE ua.refund_status = %s
                ORDER BY ua.refund_request_time DESC
            """, (status,))

        refunds = cursor.fetchall()

        for refund in refunds:
            if refund.get('refund_request_time'):
                refund['refund_request_time_cn'] = format_datetime_to_chinese(refund['refund_request_time'])

        cursor.close()
        conn.close()

        return jsonify(success_response(refunds))
    except Exception as e:
        return jsonify(error_response(f"获取退款列表失败: {str(e)}")), 500


@app.route('/api/admin/refunds/<int:registration_id>/review', methods=['PUT'])
@admin_required
def admin_review_refund(registration_id):
    """管理员审核退款"""
    try:
        data = request.get_json()
        status = data.get('status')
        admin_notes = data.get('admin_notes', '')
        refund_type = data.get('type', 'activity')


        if status not in ['approved', 'rejected']:
            return jsonify(error_response("无效的审核状态")), 400


        conn = get_db_connection()
        cursor = conn.cursor()


        if refund_type == 'course':
            table = 'teacher_course_bookings'
            amount_field = 'payment_amount'
        elif refund_type == 'product':
            table = 'orders'
            amount_field = 'total_amount'
        else:
            table = 'user_activities'
            amount_field = 'total_amount'


        # 更新退款审核状态
        cursor.execute(f"""
            UPDATE {table}
            SET refund_status = %s, refund_admin_notes = %s
            WHERE id = %s
        """, (status, admin_notes, registration_id))


        if status == 'approved':
            cursor.execute(f"SELECT * FROM {table} WHERE id = %s", (registration_id,))
            record = cursor.fetchone()


            if not record:
                conn.rollback()
                return jsonify(error_response("记录不存在")), 404


            # 微信退款
            out_trade_no = record.get('out_trade_no', '')
            total_amount = float(record[amount_field])
            total_fen = int(round(total_amount * 100))


            if out_trade_no:
                try:
                    import wechat_pay
                    out_refund_no = f"REF{registration_id}_{int(time.time())}"
                    wechat_pay.create_refund(
                        out_trade_no=out_trade_no,
                        out_refund_no=out_refund_no,
                        total_fen=total_fen,
                        refund_fen=total_fen,
                        reason=admin_notes or '管理员审核退款'
                    )
                except Exception as e:
                    conn.rollback()
                    return jsonify(error_response(f"微信退款失败: {str(e)}")), 500


            # 更新支付状态为已退款
            cursor.execute(f"""
                UPDATE {table}
                SET payment_status = 'refunded', refunded_at = NOW()
                WHERE id = %s
            """, (registration_id,))


            # ★ 新增：退款通过后取消注册并减少人数
            if refund_type == 'activity':
                cursor.execute("""
                    UPDATE user_activities SET status = 'cancelled' WHERE id = %s
                """, (registration_id,))
                cursor.execute("""
                    UPDATE activities SET current_participants = GREATEST(0, current_participants - 1)
                    WHERE id = (SELECT activity_id FROM user_activities WHERE id = %s)
                """, (registration_id,))


            elif refund_type == 'course':
                cursor.execute("""
                    UPDATE teacher_course_bookings SET status = 'cancelled' WHERE id = %s
                """, (registration_id,))
                # 释放私教时间段
                if record.get('schedule_id'):
                    cursor.execute("""
                        UPDATE teacher_course_schedules
                        SET is_booked = 0, booking_id = NULL
                        WHERE id = %s
                    """, (record['schedule_id'],))
                # 减少团课人数
                cursor.execute("SELECT course_type FROM teacher_courses WHERE id = %s", (record['course_id'],))
                course_info = cursor.fetchone()
                if course_info and course_info['course_type'] == 'group':
                    cursor.execute("""
                        UPDATE teacher_courses
                        SET current_participants = GREATEST(0, current_participants - 1)
                        WHERE id = %s
                    """, (record['course_id'],))


            elif refund_type == 'product':
                cursor.execute("UPDATE orders SET status = 'refunded' WHERE id = %s", (registration_id,))


            # 扣除积分
            deduct_user_points(record['user_id'], total_amount, 'refund',
                               order_id=registration_id if refund_type == 'product' else None,
                               booking_id=registration_id if refund_type != 'product' else None,
                               description=f'退款扣除积分¥{total_amount}')


        conn.commit()


        # 记录审计日志和通知用户
        try:
            cursor2 = conn.cursor()
            cursor2.execute(f"SELECT * FROM {table} WHERE id = %s", (registration_id,))
            refund_detail = cursor2.fetchone()


            if refund_detail:
                Logger.log_audit(
                    user_id=request.current_user['id'],
                    user_name=request.current_user.get('real_name', '管理员'),
                    action='审核退款',
                    target_type=table,
                    target_id=registration_id,
                    old_value={'refund_status': 'pending'},
                    new_value={'refund_status': status, 'admin_notes': admin_notes, 'type': refund_type}
                )


                title = ''
                if refund_type == 'activity':
                    cursor2.execute("""
                        SELECT a.title FROM user_activities ua
                        JOIN activities a ON ua.activity_id = a.id WHERE ua.id = %s
                    """, (registration_id,))
                    info = cursor2.fetchone()
                    title = info['title'] if info else ''
                elif refund_type == 'course':
                    cursor2.execute("""
                        SELECT tc.title FROM teacher_course_bookings tcb
                        JOIN teacher_courses tc ON tcb.course_id = tc.id WHERE tcb.id = %s
                    """, (registration_id,))
                    info = cursor2.fetchone()
                    title = info['title'] if info else ''
                else:
                    title = f"商品订单 {refund_detail.get('order_no', '')}"


                if status == 'approved':
                    Notifier.send_refund_approved(
                        user_id=refund_detail['user_id'],
                        activity_title=title,
                        amount=refund_detail.get('total_amount', refund_detail.get('payment_amount', 0))
                    )
                elif status == 'rejected':
                    Notifier.send_refund_rejected(
                        user_id=refund_detail['user_id'],
                        item_title=title,
                        reject_reason=admin_notes or '管理员拒绝退款'
                    )


            cursor2.close()
        except Exception as e:
            print(f"日志/通知记录失败（不影响主流程）: {str(e)}")


        cursor.close()
        conn.close()


        message = "退款审核通过" if status == 'approved' else "退款审核未通过"
        return jsonify(success_response(None, message))
    except Exception as e:
        return jsonify(error_response(f"审核退款失败: {str(e)}")), 500




@app.route('/api/orders/<int:order_id>', methods=['DELETE'])
@login_required
def delete_order(order_id):
    """取消订单（用户取消待支付订单）"""
    try:
        user_id = request.current_user['id']

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM orders WHERE id = %s AND user_id = %s", (order_id, user_id))
        order = cursor.fetchone()

        if not order:
            cursor.close()
            conn.close()
            return jsonify(error_response("订单不存在")), 404

        if order['payment_status'] == 'paid':
            cursor.close()
            conn.close()
            return jsonify(error_response("已支付的订单不能取消，请申请退款")), 400

        # 取消订单（恢复库存）
        cursor.execute("""
            SELECT oi.product_id, oi.quantity FROM order_items oi WHERE oi.order_id = %s
        """, (order_id,))
        items = cursor.fetchall()

        for item in items:
            cursor.execute("""
                UPDATE products SET stock = stock + %s WHERE id = %s
            """, (item['quantity'], item['product_id']))

        cursor.execute(
            "UPDATE orders SET status = 'cancelled' WHERE id = %s",
            (order_id,)
        )

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(None, "订单已取消"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"取消失败: {str(e)}")), 500



# ==================== 通知相关接口 ====================

@app.route('/api/notifications', methods=['GET'])
@login_required
def get_user_notifications():
    """获取用户通知列表"""
    try:
        user_id = request.current_user['id']
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        offset = (page - 1) * limit

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM notification_logs
            WHERE user_id = %s
            ORDER BY is_read ASC, created_at DESC
            LIMIT %s OFFSET %s
        """, (user_id, limit, offset))

        notifications = cursor.fetchall()

        # 格式化日期
        for notif in notifications:
            if notif.get('created_at'):
                notif['created_at_cn'] = format_datetime_to_chinese(notif['created_at'])
            if notif.get('sent_at'):
                notif['sent_at_cn'] = format_datetime_to_chinese(notif['sent_at'])

        cursor.close()
        conn.close()

        return jsonify(success_response(notifications))
    except Exception as e:
        return jsonify(error_response(f"获取通知列表失败: {str(e)}")), 500


@app.route('/api/notifications/<int:notification_id>/read', methods=['PUT'])
@login_required
def mark_notification_as_read(notification_id):
    """标记通知为已读"""
    try:
        user_id = request.current_user['id']

        conn = get_db_connection()
        cursor = conn.cursor()

        # 确认通知属于当前用户
        cursor.execute("""
            SELECT * FROM notification_logs 
            WHERE id = %s AND user_id = %s
        """, (notification_id, user_id))

        notification = cursor.fetchone()

        if not notification:
            cursor.close()
            conn.close()
            return jsonify(error_response("通知不存在")), 404

        # 标记为已读
        cursor.execute("""
            UPDATE notification_logs 
            SET is_read = 1
            WHERE id = %s
        """, (notification_id,))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(None, "已标记为已读"))
    except Exception as e:
        return jsonify(error_response(f"标记已读失败: {str(e)}")), 500


@app.route('/api/admin/logs/operations', methods=['GET'])
@admin_required
def get_operation_logs():
    """管理员查看操作日志"""
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM operation_logs
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
        """, (limit, offset))

        logs = cursor.fetchall()

        for log in logs:
            if log.get('created_at'):
                log['created_at_cn'] = format_datetime_to_chinese(log['created_at'])

        cursor.close()
        conn.close()

        return jsonify(success_response(logs))
    except Exception as e:
        return jsonify(error_response(f"获取操作日志失败: {str(e)}")), 500


@app.route('/api/admin/logs/errors', methods=['GET'])
@admin_required
def get_error_logs():
    """管理员查看错误日志"""
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM error_logs
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
        """, (limit, offset))

        logs = cursor.fetchall()

        for log in logs:
            if log.get('created_at'):
                log['created_at_cn'] = format_datetime_to_chinese(log['created_at'])

        cursor.close()
        conn.close()

        return jsonify(success_response(logs))
    except Exception as e:
        return jsonify(error_response(f"获取错误日志失败: {str(e)}")), 500


@app.route('/api/admin/backup/manual', methods=['POST'])
@admin_required
def manual_backup():
    """管理员手动触发数据库备份"""
    try:
        from backup import backup_database

        # 记录审计日志
        Logger.log_audit(
            user_id=request.current_user['id'],
            user_name=request.current_user.get('real_name', '管理员'),
            action='手动备份数据库',
            target_type='system',
            target_id=None
        )

        success = backup_database()

        if success:
            return jsonify(success_response(None, "数据库备份成功"))
        else:
            return jsonify(error_response("数据库备份失败")), 500

    except Exception as e:
        Logger.log_exception(e, user_id=request.current_user['id'])
        return jsonify(error_response(f"备份失败: {str(e)}")), 500


@app.route('/api/admin/users/<int:user_id>/role', methods=['PUT'])
@admin_required
def admin_update_user_role(user_id):
    """管理员修改用户权限（已弃用，统一走 set_user_type）"""
    # 为了兼容性保留此接口，内部转发到 set_user_type 逻辑
    try:
        data = request.get_json()
        new_role = data.get('role')

        if new_role not in ['user', 'admin']:
            return jsonify(error_response("无效的角色")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            cursor.close()
            conn.close()
            return jsonify(error_response("用户不存在")), 404

        # 同步 user_type
        if new_role == 'admin':
            new_user_type = 'admin'
        else:
            new_user_type = 'user'

        cursor.execute("""
            UPDATE users SET role = %s, user_type = %s WHERE id = %s
        """, (new_role, new_user_type, user_id))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(None, f"用户权限已更新为{new_role}"))
    except Exception as e:
        return jsonify(error_response(f"修改权限失败: {str(e)}")), 500




# ==================== 积分系统工具函数 ====================
def add_user_points(user_id, amount, change_type, order_id=None, booking_id=None, description=None):
    """增加用户积分（100元=1积分）"""
    try:
        points = int(amount / 100)
        if points == 0:
            return

        conn = get_db_connection()
        cursor = conn.cursor()

        # 更新用户总积分
        cursor.execute("""
            UPDATE users SET total_points = total_points + %s WHERE id = %s
        """, (points, user_id))

        # 获取更新后的积分
        cursor.execute("SELECT total_points FROM users WHERE id = %s", (user_id,))
        current_points = cursor.fetchone()['total_points']

        # 记录积分变化
        cursor.execute("""
            INSERT INTO user_points
            (user_id, points_change, current_points, change_type, related_order_id, related_booking_id, description)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (user_id, points, current_points, change_type, order_id, booking_id,
              description or f'消费¥{amount}获得{points}积分'))

        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"积分增加失败: {str(e)}")


def deduct_user_points(user_id, amount, change_type, order_id=None, booking_id=None, description=None):
    """扣除用户积分（退款时）"""
    try:
        points = int(amount / 100)
        if points == 0:
            return

        conn = get_db_connection()
        cursor = conn.cursor()

        # 更新用户总积分（不能为负数）
        cursor.execute("""
            UPDATE users SET total_points = GREATEST(0, total_points - %s) WHERE id = %s
        """, (points, user_id))

        # 获取更新后的积分
        cursor.execute("SELECT total_points FROM users WHERE id = %s", (user_id,))
        current_points = cursor.fetchone()['total_points']

        # 记录积分变化
        cursor.execute("""
            INSERT INTO user_points
            (user_id, points_change, current_points, change_type, related_order_id, related_booking_id, description)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (user_id, -points, current_points, change_type, order_id, booking_id,
              description or f'退款¥{amount}扣除{points}积分'))

        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"积分扣除失败: {str(e)}")


# ==================== 筛选分类管理 ====================
@app.route('/api/filter-categories', methods=['GET'])
def get_filter_categories():
    """获取筛选分类列表"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM filter_categories
            WHERE status = 'active'
            ORDER BY sort_order ASC, id ASC
        """)
        categories = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify(success_response(categories, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


@app.route('/api/filter-categories', methods=['POST'])
@admin_required
def create_filter_category():
    """创建筛选分类（管理员）"""
    try:
        data = request.json
        category_name = data.get('category_name', '').strip()
        category_type = data.get('category_type', 'course')
        sort_order = data.get('sort_order', 0)

        if not category_name:
            return jsonify(error_response("分类名称不能为空")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查是否已存在
        cursor.execute("SELECT id FROM filter_categories WHERE category_name = %s", (category_name,))
        if cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify(error_response("该分类已存在")), 400

        # 创建分类
        cursor.execute("""
            INSERT INTO filter_categories (category_name, category_type, sort_order)
            VALUES (%s, %s, %s)
        """, (category_name, category_type, sort_order))

        category_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response({'id': category_id}, "创建成功"))
    except Exception as e:
        return jsonify(error_response(f"创建失败: {str(e)}")), 500


@app.route('/api/filter-categories/<int:category_id>', methods=['PUT'])
@admin_required
def update_filter_category(category_id):
    """更新筛选分类"""
    try:
        data = request.json

        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查分类是否存在
        cursor.execute("SELECT id FROM filter_categories WHERE id = %s", (category_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify(error_response("分类不存在")), 404

        # 构建更新语句
        updates = []
        params = []

        if 'category_name' in data:
            updates.append("category_name = %s")
            params.append(data['category_name'])
        if 'sort_order' in data:
            updates.append("sort_order = %s")
            params.append(data['sort_order'])
        if 'status' in data:
            updates.append("status = %s")
            params.append(data['status'])

        if updates:
            params.append(category_id)
            cursor.execute(f"""
                UPDATE filter_categories SET {', '.join(updates)} WHERE id = %s
            """, params)
            conn.commit()

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "更新成功"))
    except Exception as e:
        return jsonify(error_response(f"更新失败: {str(e)}")), 500


@app.route('/api/filter-categories/<int:category_id>', methods=['DELETE'])
@admin_required
def delete_filter_category(category_id):
    """删除筛选分类"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查是否有课程使用该分类
        cursor.execute("SELECT COUNT(*) as count FROM teacher_courses WHERE category_id = %s", (category_id,))
        result = cursor.fetchone()
        if result['count'] > 0:
            cursor.close()
            conn.close()
            return jsonify(error_response("该分类正在被使用，无法删除")), 400

        # 删除分类
        cursor.execute("DELETE FROM filter_categories WHERE id = %s", (category_id,))
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(None, "删除成功"))
    except Exception as e:
        return jsonify(error_response(f"删除失败: {str(e)}")), 500


# ==================== 教练列表和详情 ====================
@app.route('/api/teachers', methods=['GET'])
def get_teachers():
    """获取教练列表"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        search = request.args.get('search', '')

        query = """
            SELECT
                u.id, u.real_name as name, u.avatar_url, u.phone,
                u.teacher_cover_image, u.teacher_intro, u.teacher_detail,
                COUNT(DISTINCT tc.id) as course_count,
                COUNT(DISTINCT tcb.id) as total_lessons
            FROM users u
            LEFT JOIN teacher_courses tc ON u.id = tc.teacher_id AND tc.status = 'active'
            LEFT JOIN teacher_course_bookings tcb ON u.id = tcb.teacher_id AND tcb.status = 'approved'
            WHERE u.user_type = 'teacher'
        """

        params = []
        if search:
            query += " AND u.real_name LIKE %s"
            params.append(f'%{search}%')

        query += " GROUP BY u.id ORDER BY total_lessons DESC, u.id ASC"

        cursor.execute(query, params)
        teachers = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify(success_response(teachers, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


@app.route('/api/teachers/<int:teacher_id>', methods=['GET'])
def get_teacher_detail(teacher_id):
    """获取教练详情"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取教练基本信息
        cursor.execute("""
            SELECT
                u.id, u.real_name as name, u.avatar_url, u.phone,
                u.teacher_cover_image, u.teacher_intro, u.teacher_detail,
                u.total_points
            FROM users u
            WHERE u.id = %s AND u.user_type = 'teacher'
        """, (teacher_id,))

        teacher = cursor.fetchone()
        if not teacher:
            cursor.close()
            conn.close()
            return jsonify(error_response("教练不存在")), 404

        # 获取教练的私教课程
        cursor.execute("""
            SELECT
                tc.id, tc.title, tc.cover_image, tc.description,
                tc.duration, tc.price, fc.category_name
            FROM teacher_courses tc
            LEFT JOIN filter_categories fc ON tc.category_id = fc.id
            WHERE tc.teacher_id = %s AND tc.course_type = 'private' AND tc.status = 'active'
            ORDER BY tc.created_at DESC
        """, (teacher_id,))
        private_courses = cursor.fetchall()

        # 获取教练的团课（最近的）
        cursor.execute("""
            SELECT
                tc.id, tc.title, tc.cover_image, tc.description,
                tc.duration, tc.price, tc.course_start, tc.course_end,
                tc.max_participants, tc.current_participants, tc.location,
                fc.category_name
            FROM teacher_courses tc
            LEFT JOIN filter_categories fc ON tc.category_id = fc.id
            WHERE tc.teacher_id = %s AND tc.course_type = 'group'
                AND tc.status = 'active' AND tc.course_start >= NOW()
            ORDER BY tc.course_start ASC
            LIMIT 10
        """, (teacher_id,))
        group_courses = cursor.fetchall()

        # 统计累计课时
        cursor.execute("""
            SELECT COUNT(*) as total_lessons
            FROM teacher_course_bookings
            WHERE teacher_id = %s AND status = 'approved'
        """, (teacher_id,))
        stats = cursor.fetchone()

        cursor.close()
        conn.close()

        result = {
            'teacher': teacher,
            'private_courses': private_courses,
            'group_courses': group_courses,
            'total_lessons': stats['total_lessons']
        }

        return jsonify(success_response(result, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


@app.route('/api/teachers/<int:teacher_id>/profile', methods=['PUT'])
@login_required
def update_teacher_profile(teacher_id):
    """更新教练信息（教练本人或管理员）"""
    try:
        current_user = request.current_user

        # 权限检查：只有教练本人或管理员可以修改
        if current_user['id'] != teacher_id and current_user.get('user_type') != 'admin':
            return jsonify(error_response("无权限")), 403

        data = request.json

        conn = get_db_connection()
        cursor = conn.cursor()

        # 构建更新语句
        updates = []
        params = []

        if 'teacher_cover_image' in data:
            updates.append("teacher_cover_image = %s")
            params.append(data['teacher_cover_image'])
        if 'teacher_intro' in data:
            updates.append("teacher_intro = %s")
            params.append(data['teacher_intro'])
        if 'teacher_detail' in data:
            updates.append("teacher_detail = %s")
            params.append(data['teacher_detail'])
        if 'avatar_url' in data:
            updates.append("avatar_url = %s")
            params.append(data['avatar_url'])

        if updates:
            params.append(teacher_id)
            cursor.execute(f"""
                UPDATE users SET {', '.join(updates)} WHERE id = %s AND user_type = 'teacher'
            """, params)
            conn.commit()

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "更新成功"))
    except Exception as e:
        return jsonify(error_response(f"更新失败: {str(e)}")), 500


# ==================== 教练课程管理 ====================
@app.route('/api/teacher-courses', methods=['GET'])
def get_teacher_courses():
    """获取教练课程列表"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        course_type = request.args.get('course_type', '')
        teacher_id = request.args.get('teacher_id', '')
        category_id = request.args.get('category_id', '')
        search = request.args.get('search', '')
        date = request.args.get('date', '')
        show_all = request.args.get('show_all', '')

        print(f"\n===== get_teacher_courses =====")
        print(f"course_type={course_type}, teacher_id={teacher_id}, show_all={show_all}")
        print(f"search={search}, date={date}, category_id={category_id}")

        query = """
            SELECT
                tc.*,
                u.real_name as teacher_name,
                u.avatar_url as teacher_avatar,
                fc.category_name
            FROM teacher_courses tc
            LEFT JOIN users u ON tc.teacher_id = u.id
            LEFT JOIN filter_categories fc ON tc.category_id = fc.id
            WHERE 1=1
        """
        params = []

        # 调试：先不加任何过滤，看看总共有多少课程
        cursor.execute("SELECT id, title, status, course_type FROM teacher_courses")
        all_courses = cursor.fetchall()
        print(f"数据库中所有课程: {all_courses}")

        # 管理员和教练查看全部课程（含inactive），普通用户只看active
        current_user = None
        try:
            current_user = get_current_user()
        except Exception as e:
            print(f"获取当前用户失败(不影响): {e}")

        print(f"当前用户: {current_user}")

        if not show_all and (not current_user or (current_user.get('role') != 'admin' and current_user.get('user_type') != 'teacher')):
            query += " AND tc.status = 'active'"
            print("已添加 status='active' 过滤")
        else:
            print("未添加 status 过滤（show_all 或 管理员/教练）")

        if course_type:
            query += " AND tc.course_type = %s"
            params.append(course_type)

        if teacher_id:
            query += " AND (tc.teacher_id = %s OR JSON_CONTAINS(tc.teacher_ids, CAST(%s AS JSON)))"
            params.append(teacher_id)
            params.append(int(teacher_id))

        if category_id:
            query += " AND tc.category_id = %s"
            params.append(category_id)

        if search:
            query += " AND tc.title LIKE %s"
            params.append(f'%{search}%')

        if date and course_type == 'group':
            # 所选日期在课程起止日期范围内即显示
            query += " AND DATE(tc.course_start) <= %s AND DATE(tc.course_end) >= %s"
            params.append(date)
            params.append(date)

        if course_type == 'group':
            query += " ORDER BY tc.course_start ASC"
        else:
            query += " ORDER BY tc.created_at DESC"

        print(f"最终SQL: {query}")
        print(f"参数: {params}")

        cursor.execute(query, params)
        courses = cursor.fetchall()
        print(f"查询到 {len(courses)} 条课程")

        # 格式化日期和Decimal字段
        from decimal import Decimal
        date_fields = ['course_start', 'course_end', 'registration_start', 'registration_end',
                       'cancel_deadline', 'created_at', 'updated_at']
        decimal_fields = ['price']
        for c in courses:
            for field in date_fields:
                if field in c and c[field] is not None:
                    if hasattr(c[field], 'strftime'):
                        c[field] = c[field].strftime('%Y-%m-%d %H:%M')
            for field in decimal_fields:
                if field in c and isinstance(c.get(field), Decimal):
                    c[field] = float(c[field])

        import json as json_lib
        all_teacher_ids = set()
        for c in courses:
            if c.get('teacher_ids'):
                ids = c['teacher_ids'] if isinstance(c['teacher_ids'], list) else json_lib.loads(c['teacher_ids'])
                all_teacher_ids.update(ids)
                c['teacher_ids'] = ids
            else:
                c['teacher_ids'] = []

        if all_teacher_ids:
            placeholders = ','.join(['%s'] * len(all_teacher_ids))
            cursor.execute(f"SELECT id, real_name FROM users WHERE id IN ({placeholders})", list(all_teacher_ids))
            teacher_map = {t['id']: t['real_name'] for t in cursor.fetchall()}
            for c in courses:
                if c.get('teacher_ids'):
                    c['teacher_names'] = [teacher_map.get(tid, '') for tid in c['teacher_ids']]
        else:
            for c in courses:
                c['teacher_names'] = []

        cursor.close()
        conn.close()

        return jsonify(success_response(courses, "获取成功"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"获取失败: {str(e)}")), 500



@app.route('/api/teacher-courses/<int:course_id>', methods=['GET'])
def get_teacher_course_detail(course_id):
    """获取课程详情"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                tc.*,
                u.real_name as teacher_name,
                u.avatar_url as teacher_avatar,
                u.phone as teacher_phone,
                fc.category_name
            FROM teacher_courses tc
            LEFT JOIN users u ON tc.teacher_id = u.id
            LEFT JOIN filter_categories fc ON tc.category_id = fc.id
            WHERE tc.id = %s
        """, (course_id,))

        course = cursor.fetchone()
        if not course:
            cursor.close()
            conn.close()
            return jsonify(error_response("课程不存在")), 404

        print(f"\n===== get_teacher_course_detail id={course_id} =====")

        # ★ 关键修复：格式化日期字段，和 get_teacher_courses 保持一致
        from decimal import Decimal
        date_fields = ['course_start', 'course_end', 'registration_start', 'registration_end',
                       'cancel_deadline', 'created_at', 'updated_at']
        for field in date_fields:
            if field in course and course[field] is not None:
                if hasattr(course[field], 'strftime'):
                    course[field] = course[field].strftime('%Y-%m-%d %H:%M')
                    print(f"格式化 {field}: {course[field]}")

        decimal_fields = ['price']
        for field in decimal_fields:
            if field in course and isinstance(course.get(field), Decimal):
                course[field] = float(course[field])

        # 解析 teacher_ids
        import json as json_lib
        if course.get('teacher_ids'):
            ids = course['teacher_ids'] if isinstance(course['teacher_ids'], list) else json_lib.loads(course['teacher_ids'])
            course['teacher_ids'] = ids
            if ids:
                placeholders = ','.join(['%s'] * len(ids))
                cursor.execute(f"SELECT id, real_name FROM users WHERE id IN ({placeholders})", ids)
                teacher_map = {t['id']: t['real_name'] for t in cursor.fetchall()}
                course['teacher_names'] = [teacher_map.get(tid, '') for tid in ids]

        cursor.close()
        conn.close()

        print(f"返回课程数据: course_start={course.get('course_start')}, registration_start={course.get('registration_start')}")
        return jsonify(success_response(course, "获取成功"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"获取失败: {str(e)}")), 500



@app.route('/api/teacher-courses', methods=['POST'])
@login_required
def create_teacher_course():
    """创建教练课程（教练或管理员）"""
    try:
        current_user = request.current_user
        data = request.json

        # 权限检查
        if current_user.get('user_type') not in ['teacher', 'admin']:
            return jsonify(error_response("无权限创建课程")), 403

        # 如果是教练，只能创建自己的课程
        teacher_ids = data.get('teacher_ids', [])
        teacher_id = data.get('teacher_id')
        if current_user.get('user_type') == 'teacher':
            teacher_id = current_user['id']
            teacher_ids = [current_user['id']]
        elif teacher_ids:
            teacher_id = teacher_ids[0]

        if not teacher_ids and teacher_id:
            teacher_ids = [teacher_id]

        course_type = data.get('course_type', 'private')  # private 或 group
        title = data.get('title', '').strip()
        description = data.get('description', '')
        category_id = data.get('category_id')
        duration = data.get('duration', 60)
        price = data.get('price', 0)
        cover_image = data.get('cover_image', '')

        if not title:
            return jsonify(error_response("课程名称不能为空")), 400
        if not teacher_id:
            return jsonify(error_response("教练ID不能为空")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查教练是否存在
        cursor.execute("SELECT user_type FROM users WHERE id = %s", (teacher_id,))
        teacher = cursor.fetchone()
        if not teacher or teacher['user_type'] not in ['teacher', 'admin']:
            cursor.close()
            conn.close()
            return jsonify(error_response("教练不存在")), 404

        # 创建课程
        import json as json_lib
        teacher_ids_json = json_lib.dumps(teacher_ids)

        if course_type == 'group':
            # 团课：需要时间和地点
            course_start = data.get('course_start')
            course_end = data.get('course_end')
            registration_start = data.get('registration_start')
            registration_end = data.get('registration_end')
            cancel_deadline = data.get('cancel_deadline')
            location = data.get('location', '')
            max_participants = data.get('max_participants', 20)

            cursor.execute("""
                INSERT INTO teacher_courses
                (course_type, teacher_id, teacher_ids, title, cover_image, description, category_id,
                 duration, price, course_start, course_end, registration_start, registration_end,
                 cancel_deadline, location, max_participants)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (course_type, teacher_id, teacher_ids_json, title, cover_image, description, category_id,
                  duration, price, course_start, course_end, registration_start, registration_end,
                  cancel_deadline, location, max_participants))
        else:
            # 私教：不需要固定时间
            cursor.execute("""
                INSERT INTO teacher_courses
                (course_type, teacher_id, teacher_ids, title, cover_image, description, category_id, duration, price)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (course_type, teacher_id, teacher_ids_json, title, cover_image, description, category_id, duration, price))

        course_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response({'id': course_id}, "创建成功"))
    except Exception as e:
        return jsonify(error_response(f"创建失败: {str(e)}")), 500


@app.route('/api/teacher-courses/<int:course_id>', methods=['PUT'])
@login_required
def update_teacher_course(course_id):
    """更新课程信息"""
    import json as json_lib
    import traceback as tb
    try:
        current_user = request.current_user
        data = request.json

        print(f"\n===== update_teacher_course course_id={course_id} =====")
        print(f"收到数据: {data}")

        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查课程是否存在
        cursor.execute("SELECT teacher_id FROM teacher_courses WHERE id = %s", (course_id,))
        course = cursor.fetchone()
        if not course:
            cursor.close()
            conn.close()
            return jsonify(error_response("课程不存在")), 404

        # 权限检查
        if (current_user['id'] != course['teacher_id']
                and current_user.get('user_type') != 'admin'
                and current_user.get('role') != 'admin'):
            cursor.close()
            conn.close()
            return jsonify(error_response("无权限")), 403

        updates = []
        params = []

        # 逐字段安全处理
        if 'title' in data and data['title']:
            updates.append("title = %s")
            params.append(str(data['title']))

        if 'description' in data:
            updates.append("description = %s")
            params.append(data['description'] or '')

        if 'category_id' in data:
            updates.append("category_id = %s")
            params.append(data['category_id'])  # 允许 None

        if 'duration' in data and data['duration'] is not None:
            updates.append("duration = %s")
            params.append(int(data['duration']))

        if 'price' in data and data['price'] is not None:
            updates.append("price = %s")
            params.append(float(data['price']))

        if 'cover_image' in data:
            updates.append("cover_image = %s")
            params.append(data['cover_image'] or '')

        if 'no_review_needed' in data:
            updates.append("no_review_needed = %s")
            params.append(1 if data['no_review_needed'] else 0)

        if 'course_start' in data:
            updates.append("course_start = %s")
            params.append(data['course_start'])  # 允许 None（私教课无固定时间）

        if 'course_end' in data:
            updates.append("course_end = %s")
            params.append(data['course_end'])

        if 'location' in data:
            updates.append("location = %s")
            params.append(data['location'] or '')

        if 'max_participants' in data and data['max_participants'] is not None:
            updates.append("max_participants = %s")
            params.append(int(data['max_participants']))

        if 'registration_start' in data:
            updates.append("registration_start = %s")
            params.append(data['registration_start'])

        if 'registration_end' in data:
            updates.append("registration_end = %s")
            params.append(data['registration_end'])

        if 'cancel_deadline' in data:
            updates.append("cancel_deadline = %s")
            params.append(data['cancel_deadline'])

        if 'course_type' in data and data['course_type']:
            updates.append("course_type = %s")
            params.append(data['course_type'])

        # teacher_ids 处理
        if 'teacher_ids' in data:
            teacher_ids = data['teacher_ids']
            # 防止双重序列化
            if isinstance(teacher_ids, str):
                try:
                    teacher_ids = json_lib.loads(teacher_ids)
                except Exception:
                    teacher_ids = []
            if not isinstance(teacher_ids, list):
                teacher_ids = []

            updates.append("teacher_ids = %s")
            params.append(json_lib.dumps(teacher_ids))

            # teacher_id 只在有值时更新，避免 NULL 违反外键约束
            if teacher_ids and len(teacher_ids) > 0:
                updates.append("teacher_id = %s")
                params.append(int(teacher_ids[0]))

        if not updates:
            cursor.close()
            conn.close()
            return jsonify(success_response(None, "无需更新"))

        params.append(course_id)
        sql = f"UPDATE teacher_courses SET {', '.join(updates)} WHERE id = %s"
        print(f"执行SQL: {sql}")
        print(f"参数: {params}")

        cursor.execute(sql, params)
        conn.commit()

        cursor.close()
        conn.close()

        print("更新成功")
        return jsonify(success_response(None, "更新成功"))

    except Exception as e:
        print(f"\n!!!!! update_teacher_course 异常 !!!!!")
        print(f"异常类型: {type(e).__name__}")
        print(f"异常信息: {str(e)}")
        tb.print_exc()
        return jsonify(error_response(f"更新失败: {str(e)}")), 500




@app.route('/api/teacher-courses/<int:course_id>', methods=['DELETE'])
@login_required
def delete_teacher_course(course_id):
    """删除课程及其所有相关数据（级联删除）"""
    try:
        current_user = request.current_user

        conn = get_db_connection()
        cursor = conn.cursor()

        print(f"\n===== delete_teacher_course id={course_id} =====")

        # 检查课程是否存在
        cursor.execute("SELECT teacher_id, title FROM teacher_courses WHERE id = %s", (course_id,))
        course = cursor.fetchone()
        if not course:
            cursor.close()
            conn.close()
            return jsonify(error_response("课程不存在")), 404

        # 权限检查
        if (current_user['id'] != course['teacher_id']
                and current_user.get('user_type') != 'admin'
                and current_user.get('role') != 'admin'):
            cursor.close()
            conn.close()
            return jsonify(error_response("无权限")), 403

        # 统计关联数据
        cursor.execute(
            "SELECT COUNT(*) as count FROM teacher_course_bookings WHERE course_id = %s",
            (course_id,)
        )
        booking_count = cursor.fetchone()['count']
        print(f"关联预约记录: {booking_count}")

        # 级联删除
        # 1. 删除时间表
        cursor.execute("DELETE FROM teacher_course_schedules WHERE course_id = %s", (course_id,))
        print(f"删除时间表: {cursor.rowcount}")

        # 2. 删除所有预约记录
        cursor.execute("DELETE FROM teacher_course_bookings WHERE course_id = %s", (course_id,))
        print(f"删除预约记录: {cursor.rowcount}")

        # 3. 删除课程本身
        cursor.execute("DELETE FROM teacher_courses WHERE id = %s", (course_id,))
        print(f"删除课程: {cursor.rowcount}")

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response({
            'deleted_bookings': booking_count
        }, "课程及所有相关数据已删除"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"删除失败: {str(e)}")), 500




# ==================== 私教课程时间管理 ====================
@app.route('/api/teacher-courses/<int:course_id>/available-times', methods=['GET'])
def get_available_times(course_id):
    """获取私教课程的可用时间段"""
    try:
        date = request.args.get('date')  # 格式：2026-02-26
        if not date:
            return jsonify(error_response("日期参数不能为空")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查课程是否存在且为私教
        cursor.execute("""
            SELECT course_type, duration FROM teacher_courses WHERE id = %s AND status = 'active'
        """, (course_id,))
        course = cursor.fetchone()
        if not course:
            cursor.close()
            conn.close()
            return jsonify(error_response("课程不存在")), 404
        if course['course_type'] != 'private':
            cursor.close()
            conn.close()
            return jsonify(error_response("仅私教课程支持时间段查询")), 400

        # 获取该日期的所有时间段
        cursor.execute("""
            SELECT
                tcs.id, tcs.start_time, tcs.end_time, tcs.is_booked,
                tcb.status as booking_status
            FROM teacher_course_schedules tcs
            LEFT JOIN teacher_course_bookings tcb ON tcs.booking_id = tcb.id
            WHERE tcs.course_id = %s AND tcs.schedule_date = %s
            ORDER BY tcs.start_time ASC
        """, (course_id, date))

        time_slots = cursor.fetchall()

        # 如果该日期没有时间段，自动生成（08:00 ~ 22:00 按课程时长分段）
        if not time_slots:
            duration = int(course['duration'] or 60)
            start_hour = 8
            end_hour = 22
            generated_slots = []
            current_minutes = start_hour * 60
            end_minutes = end_hour * 60

            while current_minutes + duration <= end_minutes:
                s_h = current_minutes // 60
                s_m = current_minutes % 60
                e_minutes = current_minutes + duration
                e_h = e_minutes // 60
                e_m = e_minutes % 60
                start_time = f"{s_h:02d}:{s_m:02d}"
                end_time = f"{e_h:02d}:{e_m:02d}"

                cursor.execute("""
                    INSERT INTO teacher_course_schedules (course_id, schedule_date, start_time, end_time)
                    VALUES (%s, %s, %s, %s)
                """, (course_id, date, start_time, end_time))
                slot_id = cursor.lastrowid
                generated_slots.append({
                    'id': slot_id,
                    'start_time': start_time,
                    'end_time': end_time,
                    'available': True
                })
                current_minutes += duration

            conn.commit()
            cursor.close()
            conn.close()
            return jsonify(success_response(generated_slots, "获取成功"))

        # 标记可用性
        result = []
        for slot in time_slots:
            # 只有已预约且审核通过的才显示为不可用
            is_available = not (slot['is_booked'] and slot['booking_status'] == 'approved')
            # 格式化时间为 HH:MM
            st = slot['start_time']
            et = slot['end_time']
            start_str = f"{int(st.total_seconds()) // 3600:02d}:{(int(st.total_seconds()) % 3600) // 60:02d}" if hasattr(st, 'total_seconds') else str(st)[:5]
            end_str = f"{int(et.total_seconds()) // 3600:02d}:{(int(et.total_seconds()) % 3600) // 60:02d}" if hasattr(et, 'total_seconds') else str(et)[:5]
            result.append({
                'id': slot['id'],
                'start_time': start_str,
                'end_time': end_str,
                'available': is_available
            })

        cursor.close()
        conn.close()

        return jsonify(success_response(result, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


@app.route('/api/teacher-courses/<int:course_id>/schedules', methods=['POST'])
@login_required
def create_course_schedule(course_id):
    """创建私教课程时间段（教练）"""
    try:
        current_user = request.current_user
        data = request.json

        schedule_date = data.get('schedule_date')  # 格式：2026-02-26
        start_time = data.get('start_time')  # 格式：14:00
        end_time = data.get('end_time')  # 格式：15:00

        if not all([schedule_date, start_time, end_time]):
            return jsonify(error_response("日期和时间不能为空")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查课程是否存在
        cursor.execute("""
            SELECT teacher_id, course_type FROM teacher_courses WHERE id = %s AND status = 'active'
        """, (course_id,))
        course = cursor.fetchone()
        if not course:
            cursor.close()
            conn.close()
            return jsonify(error_response("课程不存在")), 404

        # 权限检查
        if current_user['id'] != course['teacher_id'] and current_user.get('user_type') != 'admin':
            cursor.close()
            conn.close()
            return jsonify(error_response("无权限")), 403

        if course['course_type'] != 'private':
            cursor.close()
            conn.close()
            return jsonify(error_response("仅私教课程可以添加时间段")), 400

        # 检查时间段是否已存在
        cursor.execute("""
            SELECT id FROM teacher_course_schedules
            WHERE course_id = %s AND schedule_date = %s AND start_time = %s
        """, (course_id, schedule_date, start_time))
        if cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify(error_response("该时间段已存在")), 400

        # 创建时间段
        cursor.execute("""
            INSERT INTO teacher_course_schedules (course_id, schedule_date, start_time, end_time)
            VALUES (%s, %s, %s, %s)
        """, (course_id, schedule_date, start_time, end_time))

        schedule_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response({'id': schedule_id}, "创建成功"))
    except Exception as e:
        return jsonify(error_response(f"创建失败: {str(e)}")), 500


# ==================== 会员卡管理 ====================
@app.route('/api/user-membership-cards/check', methods=['GET'])
@login_required
def check_membership_card():
    """检查用户是否有可用的会员卡"""
    try:
        current_user = request.current_user
        teacher_id = request.args.get('teacher_id')

        if not teacher_id:
            return jsonify(error_response("教练ID不能为空")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 查询用户的所有活跃会员卡
        cursor.execute("""
            SELECT
                umc.*,
                mc.card_name,
                mc.card_image,
                mc.times_count as card_total_times,
                u.real_name as teacher_name
            FROM user_membership_cards umc
            LEFT JOIN membership_cards mc ON umc.card_id = mc.id
            LEFT JOIN users u ON umc.teacher_id = u.id
            WHERE umc.user_id = %s AND umc.payment_status = 'paid'
            ORDER BY FIELD(umc.status, 'active', 'expired', 'used_up'), umc.purchase_date DESC
        """, (current_user['id'],))


        cards = cursor.fetchall()

        # 检查是否有可用的卡
        available_card = None
        for card in cards:
            if card['card_type'] == 'times':
                # 次卡：检查剩余次数
                if card['remaining_times'] > 0:
                    available_card = card
                    break
            elif card['card_type'] == 'period':
                # 期限卡：检查是否在有效期内
                if card['activated'] and card['start_date'] <= datetime.now().date() <= card['end_date']:
                    available_card = card
                    break

        cursor.close()
        conn.close()

        if available_card:
            return jsonify(success_response({
                'has_card': True,
                'card': available_card
            }, "有可用会员卡"))
        else:
            return jsonify(success_response({
                'has_card': False
            }, "无可用会员卡"))

    except Exception as e:
        return jsonify(error_response(f"查询失败: {str(e)}")), 500


@app.route('/api/membership-cards', methods=['GET'])
def get_membership_cards():
    """获取会员卡列表"""
    try:
        import json as json_lib
        from decimal import Decimal

        teacher_id = request.args.get('teacher_id')

        conn = get_db_connection()
        cursor = conn.cursor()

        query = "SELECT * FROM membership_cards WHERE status = 'active'"
        params = []

        if teacher_id:
            query += " AND (JSON_CONTAINS(teacher_ids, %s, '$') OR teacher_ids IS NULL OR teacher_ids = '[]')"
            params.append(str(int(teacher_id)))

        query += " ORDER BY price ASC"
        cursor.execute(query, params)
        cards = cursor.fetchall()

        # 收集所有教练ID
        all_teacher_ids = set()
        for card in cards:
            t_ids = card.get('teacher_ids')
            if t_ids:
                if isinstance(t_ids, str):
                    try:
                        t_ids = json_lib.loads(t_ids)
                    except Exception:
                        t_ids = []
                card['teacher_ids'] = t_ids
                all_teacher_ids.update([int(i) for i in t_ids])
            else:
                card['teacher_ids'] = []
            if isinstance(card.get('price'), Decimal):
                card['price'] = float(card['price'])

        # 批量查询教练真实姓名
        teacher_map = {}
        if all_teacher_ids:
            placeholders = ','.join(['%s'] * len(all_teacher_ids))
            cursor.execute(
                f"SELECT id, real_name FROM users WHERE id IN ({placeholders})",
                list(all_teacher_ids)
            )
            teacher_map = {t['id']: t['real_name'] for t in cursor.fetchall()}

        for card in cards:
            card['teachers'] = [
                {'id': tid, 'name': teacher_map.get(int(tid), f'教练{tid}')}
                for tid in card['teacher_ids']
            ]
            card['teacher_names'] = [teacher_map.get(int(tid), f'教练{tid}') for tid in card['teacher_ids']]

        cursor.close()
        conn.close()

        return jsonify(success_response(cards, "获取成功"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"获取失败: {str(e)}")), 500



@app.route('/api/membership-cards/<int:card_id>', methods=['GET'])
def get_membership_card_detail(card_id):
    """获取会员卡详情（含关联教练信息）"""
    try:
        import json as json_lib
        from decimal import Decimal
        from datetime import date as date_type


        conn = get_db_connection()
        cursor = conn.cursor()


        cursor.execute("SELECT * FROM membership_cards WHERE id = %s", (card_id,))
        card = cursor.fetchone()


        if not card:
            cursor.close()
            conn.close()
            return jsonify(error_response("会员卡不存在")), 404


        # 格式化 Decimal
        if isinstance(card.get('price'), Decimal):
            card['price'] = float(card['price'])


        # ★ 修复：格式化日期字段为 YYYY-MM-DD 字符串
        for date_field in ['validity_start', 'validity_end']:
            val = card.get(date_field)
            if val:
                if isinstance(val, date_type) and not isinstance(val, datetime):
                    card[date_field] = val.strftime('%Y-%m-%d')
                elif hasattr(val, 'strftime'):
                    card[date_field] = val.strftime('%Y-%m-%d')
                else:
                    card[date_field] = str(val)[:10] if val else ''
            else:
                card[date_field] = ''


        # 解析 teacher_ids 并获取教练真实姓名
        teacher_ids = card.get('teacher_ids')
        if teacher_ids:
            if isinstance(teacher_ids, str):
                try:
                    teacher_ids = json_lib.loads(teacher_ids)
                except Exception:
                    teacher_ids = []
        else:
            teacher_ids = []


        card['teacher_ids'] = teacher_ids
        card['teachers'] = []


        if teacher_ids:
            placeholders = ','.join(['%s'] * len(teacher_ids))
            cursor.execute(
                f"SELECT id, real_name FROM users WHERE id IN ({placeholders})",
                teacher_ids
            )
            teachers = cursor.fetchall()
            card['teachers'] = [{'id': t['id'], 'name': t['real_name']} for t in teachers]


        cursor.close()
        conn.close()


        return jsonify(success_response(card, "获取成功"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"获取失败: {str(e)}")), 500




@app.route('/api/membership-cards', methods=['POST'])
@admin_required
def create_membership_card():
    """创建会员卡（管理员）"""
    try:
        data = request.json

        card_name = data.get('card_name', '').strip()
        card_type = data.get('card_type')  # times 或 period
        price = data.get('price', 0)
        teacher_ids = data.get('teacher_ids', [])  # 教练ID数组

        if not card_name or not card_type:
            return jsonify(error_response("会员卡名称和类型不能为空")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 准备字段
        card_image = data.get('card_image', '')
        description = data.get('description', '')
        times_count = data.get('times_count') if card_type == 'times' else None
        valid_days = data.get('valid_days') if card_type == 'period' else None
        purchase_notes = data.get('purchase_notes', '')
        applicable_stores = data.get('applicable_stores', '')
        applicable_courses = data.get('applicable_courses', '')
        validity_start = data.get('validity_start') if card_type == 'period' else None
        validity_end = data.get('validity_end') if card_type == 'period' else None

        # 创建会员卡
        cursor.execute("""
            INSERT INTO membership_cards
            (card_name, card_type, card_image, description, times_count, valid_days,
             price, teacher_ids, purchase_notes, applicable_stores, applicable_courses,
             validity_start, validity_end)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (card_name, card_type, card_image, description, times_count, valid_days,
              price, json.dumps(teacher_ids), purchase_notes, applicable_stores, applicable_courses,
              validity_start, validity_end))

        card_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response({'id': card_id}, "创建成功"))
    except Exception as e:
        return jsonify(error_response(f"创建失败: {str(e)}")), 500


@app.route('/api/membership-cards/<int:card_id>', methods=['PUT'])
@admin_required
def update_membership_card(card_id):
    """更新会员卡（管理员）"""
    try:
        data = request.json


        conn = get_db_connection()
        cursor = conn.cursor()


        cursor.execute("SELECT * FROM membership_cards WHERE id = %s", (card_id,))
        card = cursor.fetchone()
        if not card:
            cursor.close()
            conn.close()
            return jsonify(error_response("会员卡不存在")), 404


        card_name = data.get('card_name', card['card_name'])
        card_type = data.get('card_type', card['card_type'])
        price = data.get('price', card['price'])
        card_image = data.get('card_image', card['card_image'])
        description = data.get('description', card['description'])
        times_count = data.get('times_count', card['times_count'])
        valid_days = data.get('valid_days', card['valid_days'])
        teacher_ids = data.get('teacher_ids', card['teacher_ids'])
        purchase_notes = data.get('purchase_notes', card['purchase_notes'])
        applicable_stores = data.get('applicable_stores', card['applicable_stores'])
        applicable_courses = data.get('applicable_courses', card.get('applicable_courses', ''))
        validity_start = data.get('validity_start', card.get('validity_start'))
        validity_end = data.get('validity_end', card.get('validity_end'))


        # ★ 修复：空字符串转None，防止MySQL DATE列报错
        if not validity_start or validity_start == '':
            validity_start = None
        if not validity_end or validity_end == '':
            validity_end = None


        if isinstance(teacher_ids, list):
            teacher_ids = json.dumps(teacher_ids)


        cursor.execute("""
            UPDATE membership_cards
            SET card_name = %s, card_type = %s, card_image = %s, description = %s,
                times_count = %s, valid_days = %s, price = %s, teacher_ids = %s,
                purchase_notes = %s, applicable_stores = %s, applicable_courses = %s,
                validity_start = %s, validity_end = %s
            WHERE id = %s
        """, (card_name, card_type, card_image, description,
              times_count, valid_days, price, teacher_ids,
              purchase_notes, applicable_stores, applicable_courses,
              validity_start, validity_end, card_id))


        conn.commit()
        cursor.close()
        conn.close()


        return jsonify(success_response(None, "更新成功"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"更新失败: {str(e)}")), 500



@app.route('/api/membership-cards/<int:card_id>', methods=['DELETE'])
@admin_required
def delete_membership_card(card_id):
    """删除会员卡（管理员）"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查会员卡是否存在
        cursor.execute("SELECT * FROM membership_cards WHERE id = %s", (card_id,))
        card = cursor.fetchone()
        if not card:
            cursor.close()
            conn.close()
            return jsonify(error_response("会员卡不存在")), 404

        # 检查是否有用户购买了此会员卡
        cursor.execute("SELECT COUNT(*) as count FROM user_membership_cards WHERE card_id = %s AND status = 'active'", (card_id,))
        active_count = cursor.fetchone()['count']

        if active_count > 0:
            # 如果有活跃用户，只更新状态为inactive
            cursor.execute("UPDATE membership_cards SET status = 'inactive' WHERE id = %s", (card_id,))
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify(success_response(None, "已下架（有用户正在使用）"))
        else:
            # 没有活跃用户，直接删除
            cursor.execute("DELETE FROM membership_cards WHERE id = %s", (card_id,))
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify(success_response(None, "删除成功"))
    except Exception as e:
        return jsonify(error_response(f"删除失败: {str(e)}")), 500


@app.route('/api/membership-cards/<int:card_id>/purchase', methods=['POST'])
@login_required
def purchase_membership_card(card_id):
    """购买会员卡 — 创建待支付记录"""
    try:
        import json as json_lib
        current_user = request.current_user
        data = request.json

        teacher_id = data.get('teacher_id')

        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取会员卡信息
        cursor.execute("SELECT * FROM membership_cards WHERE id = %s AND status = 'active'", (card_id,))
        card = cursor.fetchone()
        if not card:
            cursor.close()
            conn.close()
            return jsonify(error_response("会员卡不存在")), 404

        # 解析关联教练
        card_teacher_ids = card.get('teacher_ids')
        if card_teacher_ids:
            if isinstance(card_teacher_ids, str):
                try:
                    card_teacher_ids = json_lib.loads(card_teacher_ids)
                except Exception:
                    card_teacher_ids = []
        else:
            card_teacher_ids = []

        # 有关联教练时才校验
        if card_teacher_ids and len(card_teacher_ids) > 0:
            if teacher_id is None:
                cursor.close()
                conn.close()
                return jsonify(error_response("请选择教练")), 400
            if int(teacher_id) not in [int(t) for t in card_teacher_ids]:
                cursor.close()
                conn.close()
                return jsonify(error_response("该教练不在此会员卡关联范围内")), 400
        else:
            teacher_id = None

        from decimal import Decimal
        price = float(card['price']) if isinstance(card['price'], Decimal) else (card['price'] or 0)

        # 创建待支付记录
        if card['card_type'] == 'times':
            cursor.execute("""
                INSERT INTO user_membership_cards
                (user_id, card_id, teacher_id, card_type,
                 total_times, remaining_times,
                 purchase_amount, activated, payment_status)
                VALUES (%s, %s, %s, 'times', %s, %s, %s, 0, 'unpaid')
            """, (
                current_user['id'], card_id, teacher_id,
                card['times_count'], card['times_count'], price
            ))
        else:
            cursor.execute("""
                INSERT INTO user_membership_cards
                (user_id, card_id, teacher_id, card_type,
                 valid_days, validity_start, validity_end,
                 purchase_amount, activated, payment_status)
                VALUES (%s, %s, %s, 'period', %s, %s, %s, %s, 0, 'unpaid')
            """, (
                current_user['id'], card_id, teacher_id,
                card['valid_days'],
                card['validity_start'],
                card['validity_end'],
                price
            ))

        user_card_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response({
            'id': user_card_id,
            'need_payment': True,
            'payment_amount': price,
            'card_name': card['card_name']
        }, "请完成支付"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"购买失败: {str(e)}")), 500




@app.route('/api/user-membership-cards', methods=['GET'])
@login_required
def get_user_membership_cards():
    """获取用户的会员卡列表"""
    try:
        current_user = request.current_user

        conn = get_db_connection()
        cursor = conn.cursor()

        # 自动过期检测：期限卡激活后 end_date 已过
        cursor.execute("""
            UPDATE user_membership_cards
            SET status = 'expired'
            WHERE user_id = %s AND card_type = 'period' AND activated = 1
              AND end_date IS NOT NULL AND end_date < CURDATE() AND status = 'active'
        """, (current_user['id'],))

        # 激活有效期过期检测：未激活但 validity_end 已过
        cursor.execute("""
            UPDATE user_membership_cards
            SET status = 'expired'
            WHERE user_id = %s AND card_type = 'period' AND activated = 0
              AND validity_end IS NOT NULL AND validity_end < CURDATE() AND status = 'active'
        """, (current_user['id'],))

        conn.commit()

        # 只显示已支付或管理员赠送的卡
        cursor.execute("""
            SELECT
                umc.*,
                mc.card_name,
                mc.card_image,
                mc.times_count as card_total_times,
                u.real_name as teacher_name
            FROM user_membership_cards umc
            LEFT JOIN membership_cards mc ON umc.card_id = mc.id
            LEFT JOIN users u ON umc.teacher_id = u.id
            WHERE umc.user_id = %s
              AND (umc.payment_status = 'paid' OR umc.gift_from_admin = 1)
            ORDER BY FIELD(umc.status, 'active', 'expired', 'used_up'), umc.purchase_date DESC
        """, (current_user['id'],))

        cards = cursor.fetchall()

        from decimal import Decimal
        for card in cards:
            for field in ['start_date', 'end_date', 'validity_start', 'validity_end']:
                if card.get(field) and hasattr(card[field], 'strftime'):
                    card[field] = card[field].strftime('%Y-%m-%d')
            if card.get('purchase_date') and hasattr(card['purchase_date'], 'strftime'):
                card['purchase_date'] = card['purchase_date'].strftime('%Y-%m-%d %H:%M:%S')
            if card.get('activated_at') and hasattr(card['activated_at'], 'strftime'):
                card['activated_at'] = card['activated_at'].strftime('%Y-%m-%d %H:%M:%S')
            if isinstance(card.get('purchase_amount'), Decimal):
                card['purchase_amount'] = float(card['purchase_amount'])

        cursor.close()
        conn.close()

        return jsonify(success_response(cards, "获取成功"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"获取失败: {str(e)}")), 500





@app.route('/api/user-membership-cards/<int:card_id>/activate', methods=['PUT'])
@login_required
def activate_user_membership_card(card_id):
    """用户激活期限卡"""
    try:
        current_user = request.current_user

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM user_membership_cards
            WHERE id = %s AND user_id = %s
        """, (card_id, current_user['id']))
        user_card = cursor.fetchone()

        if not user_card:
            cursor.close()
            conn.close()
            return jsonify(error_response("会员卡不存在")), 404

        if user_card['card_type'] == 'times':
            cursor.close()
            conn.close()
            return jsonify(error_response("次卡无需激活")), 400

        if user_card['activated'] == 1:
            cursor.close()
            conn.close()
            return jsonify(error_response("该会员卡已激活")), 400

        if user_card['status'] != 'active':
            cursor.close()
            conn.close()
            return jsonify(error_response("该会员卡已过期或不可用")), 400

        # 检查激活有效期
        from datetime import date, timedelta
        today = date.today()

        validity_end = user_card.get('validity_end')
        if validity_end:
            if hasattr(validity_end, 'date'):
                validity_end = validity_end.date()
            elif isinstance(validity_end, str):
                from datetime import datetime
                validity_end = datetime.strptime(validity_end, '%Y-%m-%d').date()
            if today > validity_end:
                # 激活有效期已过，更新为过期状态
                cursor.execute(
                    "UPDATE user_membership_cards SET status = 'expired' WHERE id = %s",
                    (card_id,)
                )
                conn.commit()
                cursor.close()
                conn.close()
                return jsonify(error_response("该会员卡激活有效期已过，无法激活")), 400

        validity_start = user_card.get('validity_start')
        if validity_start:
            if hasattr(validity_start, 'date'):
                validity_start = validity_start.date()
            elif isinstance(validity_start, str):
                from datetime import datetime
                validity_start = datetime.strptime(validity_start, '%Y-%m-%d').date()
            if today < validity_start:
                cursor.close()
                conn.close()
                return jsonify(error_response("该会员卡激活时间未到")), 400

        # 激活：今天起算 valid_days 天
        valid_days = user_card.get('valid_days') or 0
        start_date = today
        end_date = today + timedelta(days=valid_days) if valid_days else None

        cursor.execute("""
            UPDATE user_membership_cards
            SET activated = 1, activated_at = NOW(),
                start_date = %s, end_date = %s
            WHERE id = %s
        """, (start_date, end_date, card_id))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response({
            'start_date': str(start_date),
            'end_date': str(end_date) if end_date else None
        }, "激活成功"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"激活失败: {str(e)}")), 500



@app.route('/api/admin/membership-cards/<int:card_id>/gift', methods=['POST'])
@admin_required
def admin_gift_membership_card(card_id):
    """管理员赠送会员卡给多个用户"""
    try:
        import json as json_lib
        import traceback as tb

        data = request.json
        user_ids = data.get('user_ids', [])

        print(f"\n===== admin_gift_membership_card card_id={card_id} =====")
        print(f"user_ids: {user_ids}")

        if not user_ids:
            return jsonify(error_response("请选择要赠送的用户")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取会员卡信息
        cursor.execute("SELECT * FROM membership_cards WHERE id = %s", (card_id,))
        card = cursor.fetchone()
        if not card:
            cursor.close()
            conn.close()
            return jsonify(error_response("会员卡不存在")), 404

        print(f"会员卡信息: {card}")

        # 处理 teacher_ids
        card_teacher_ids = card.get('teacher_ids')
        if card_teacher_ids:
            if isinstance(card_teacher_ids, str):
                try:
                    card_teacher_ids = json_lib.loads(card_teacher_ids)
                except Exception:
                    card_teacher_ids = []
        else:
            card_teacher_ids = []

        print(f"card_teacher_ids: {card_teacher_ids}")

        # 期限卡的有效期
        validity_start = card.get('validity_start')
        validity_end = card.get('validity_end')
        valid_days = card.get('valid_days')
        times_count = card.get('times_count')
        card_type = card.get('card_type')

        from decimal import Decimal
        price = float(card.get('price', 0)) if isinstance(card.get('price'), Decimal) else (card.get('price') or 0)

        gifted_count = 0
        for user_id in user_ids:
            try:
                # 检查用户是否存在
                cursor.execute("SELECT id, user_type FROM users WHERE id = %s", (user_id,))
                user = cursor.fetchone()
                if not user:
                    print(f"用户 {user_id} 不存在，跳过")
                    continue

                if card_type == 'times':
                    # 次卡：直接生效，无需激活
                    cursor.execute("""
                        INSERT INTO user_membership_cards
                        (user_id, card_id, card_type, teacher_id,
                         remaining_times, total_times,
                         purchase_amount, status, activated, gift_from_admin)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, 'active', 1, 1)
                    """, (
                        user_id, card_id, 'times',
                        None,  # 月卡/年卡类次卡无指定教练
                        times_count or 0,
                        times_count or 0,
                        0
                    ))
                else:
                    # 期限卡：未激活状态，等待用户激活
                    cursor.execute("""
                        INSERT INTO user_membership_cards
                        (user_id, card_id, card_type, teacher_id,
                         valid_days, validity_start, validity_end,
                         purchase_amount, status, activated, gift_from_admin)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'active', 0, 1)
                    """, (
                        user_id, card_id, 'period',
                        None,  # 赠送时不绑定教练，由用户激活时确定
                        valid_days,
                        validity_start,
                        validity_end,
                        0
                    ))

                gifted_count += 1
                print(f"成功赠送给用户 {user_id}")

            except Exception as e:
                print(f"赠送给用户 {user_id} 失败: {str(e)}")
                tb.print_exc()
                continue

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(
            {'gifted_count': gifted_count},
            f"成功赠送给{gifted_count}位用户"
        ))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"赠送失败: {str(e)}")), 500



@app.route('/api/teacher-courses/<int:course_id>/participants', methods=['GET'])
@login_required
def get_course_participants(course_id):
    """获取课程已报名人员列表"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查课程是否存在
        cursor.execute("SELECT id, title, course_type, max_participants, current_participants FROM teacher_courses WHERE id = %s", (course_id,))
        course = cursor.fetchone()
        if not course:
            cursor.close()
            conn.close()
            return jsonify(error_response("课程不存在")), 404

        # 获取所有预约人员（不限状态，参考活动人员详情）
        cursor.execute("""
            SELECT
                tcb.id as booking_id,
                tcb.user_id,
                tcb.status,
                tcb.booking_date,
                tcb.payment_amount,
                tcb.payment_status,
                tcb.created_at,
                tcs.schedule_date,
                tcs.start_time,
                tcs.end_time,
                u.real_name,
                u.phone,
                u.avatar_url,
                u.gender,
                u.id_card
            FROM teacher_course_bookings tcb
            LEFT JOIN users u ON tcb.user_id = u.id
            LEFT JOIN teacher_course_schedules tcs ON tcb.schedule_id = tcs.id
            WHERE tcb.course_id = %s AND tcb.status NOT IN ('cancelled', 'rejected')
            ORDER BY tcb.created_at DESC
        """, (course_id,))

        participants = cursor.fetchall()

        # 格式化日期字段
        from decimal import Decimal
        from datetime import timedelta as td_type
        for p in participants:
            for key, val in list(p.items()):
                if isinstance(val, td_type):
                    total_secs = int(val.total_seconds())
                    p[key] = f"{total_secs // 3600:02d}:{(total_secs % 3600) // 60:02d}"
                elif hasattr(val, 'strftime'):
                    p[key] = val.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(val, Decimal):
                    p[key] = float(val)

        cursor.close()
        conn.close()

        return jsonify(success_response({
            'course': course,
            'participants': participants,
            'total': len(participants)
        }, "获取成功"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"获取失败: {str(e)}")), 500



# ==================== 课程预约管理 ====================
@app.route('/api/course-bookings', methods=['POST'])
@login_required
def create_course_booking():
    """创建课程预约（用户先付款）"""
    try:
        current_user = request.current_user
        data = request.json

        course_id = data.get('course_id')
        schedule_id = data.get('schedule_id')  # 私教必填
        payment_amount = float(data.get('payment_amount', 0))
        use_membership = data.get('use_membership', False)  # 是否使用会员卡
        membership_card_id = data.get('membership_card_id')  # 用户会员卡ID

        if not course_id:
            return jsonify(error_response("课程ID不能为空")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取课程信息
        cursor.execute("""
            SELECT * FROM teacher_courses WHERE id = %s AND status = 'active'
        """, (course_id,))
        course = cursor.fetchone()
        if not course:
            cursor.close()
            conn.close()
            return jsonify(error_response("课程不存在")), 404

        # 检查课程类型
        if course['course_type'] == 'private' and not schedule_id:
            cursor.close()
            conn.close()
            return jsonify(error_response("私教课程需要选择时间段")), 400

        # 如果是私教，检查时间段是否可用
        if course['course_type'] == 'private':
            cursor.execute("""
                SELECT is_booked FROM teacher_course_schedules
                WHERE id = %s AND course_id = %s
            """, (schedule_id, course_id))
            schedule = cursor.fetchone()
            if not schedule:
                cursor.close()
                conn.close()
                return jsonify(error_response("时间段不存在")), 404
            if schedule['is_booked']:
                # 检查是否已审核通过
                cursor.execute("""
                    SELECT status FROM teacher_course_bookings
                    WHERE schedule_id = %s AND status = 'approved'
                """, (schedule_id,))
                if cursor.fetchone():
                    cursor.close()
                    conn.close()
                    return jsonify(error_response("该时间段已被预约")), 400

        # 如果是团课，检查人数是否已满
        if course['course_type'] == 'group':
            if int(course['current_participants']) >= int(course['max_participants']):
                cursor.close()
                conn.close()
                return jsonify(error_response("团课人数已满")), 400

        # 如果使用会员卡
        if use_membership and membership_card_id:
            cursor.execute("""
                SELECT * FROM user_membership_cards
                WHERE id = %s AND user_id = %s AND teacher_id = %s AND status = 'active'
            """, (membership_card_id, current_user['id'], course['teacher_id']))
            user_card = cursor.fetchone()

            if not user_card:
                cursor.close()
                conn.close()
                return jsonify(error_response("会员卡不可用")), 400

            # 次卡：检查剩余次数
            if user_card['card_type'] == 'times':
                if user_card['remaining_times'] <= 0:
                    cursor.close()
                    conn.close()
                    return jsonify(error_response("会员卡次数已用完")), 400
                # 扣除次数
                cursor.execute("""
                    UPDATE user_membership_cards
                    SET remaining_times = remaining_times - 1
                    WHERE id = %s
                """, (membership_card_id,))
                # 检查是否用完
                cursor.execute("""
                    UPDATE user_membership_cards
                    SET status = 'used_up'
                    WHERE id = %s AND remaining_times = 0
                """, (membership_card_id,))

            # 期限卡：激活并检查有效期
            elif user_card['card_type'] == 'period':
                if not user_card['activated']:
                    # 首次使用，激活会员卡
                    cursor.execute("""
                        SELECT valid_days FROM membership_cards WHERE id = %s
                    """, (user_card['card_id'],))
                    card_info = cursor.fetchone()
                    start_date = datetime.now().date()
                    end_date = start_date + timedelta(days=card_info['valid_days'])

                    cursor.execute("""
                        UPDATE user_membership_cards
                        SET activated = 1, start_date = %s, end_date = %s
                        WHERE id = %s
                    """, (start_date, end_date, membership_card_id))
                else:
                    # 检查是否在有效期内
                    if not (user_card['start_date'] <= datetime.now().date() <= user_card['end_date']):
                        cursor.close()
                        conn.close()
                        return jsonify(error_response("会员卡已过期")), 400

            payment_amount = 0  # 使用会员卡，金额为0

        # 判断是否免审核
        no_review = course.get('no_review_needed', 0)
        if no_review:
            booking_status = 'approved'
        else:
            booking_status = 'pending'

        # 创建预约记录
        cursor.execute("""
            INSERT INTO teacher_course_bookings
            (course_id, user_id, teacher_id, schedule_id, booking_date, payment_amount,
             use_membership, membership_card_id, status, payment_status)
            VALUES (%s, %s, %s, %s, NOW(), %s, %s, %s, %s, 'unpaid')
        """, (course_id, current_user['id'], course['teacher_id'], schedule_id,
              payment_amount, use_membership, membership_card_id, booking_status))

        booking_id = cursor.lastrowid

        # 如果是私教，标记时间段为已预约
        if course['course_type'] == 'private':
            cursor.execute("""
                UPDATE teacher_course_schedules
                SET is_booked = 1, booking_id = %s
                WHERE id = %s
            """, (booking_id, schedule_id))

        # 如果是团课，增加当前参与人数
        if course['course_type'] == 'group':
            cursor.execute("""
                UPDATE teacher_courses
                SET current_participants = current_participants + 1
                WHERE id = %s
            """, (course_id,))

        # 积分在支付回调中发放，此处不再处理

        conn.commit()
        cursor.close()
        conn.close()

        # 发送通知（免审核不通知）
        if not no_review:
            try:
                Notifier.notify_course_booking(
                    course_title=course['title'],
                    user_name=current_user.get('real_name', '用户'),
                    teacher_id=course['teacher_id']
                )
            except Exception as e:
                print(f"课程预约通知发送失败（不影响主流程）: {str(e)}")

        if no_review:
            return jsonify(success_response({'id': booking_id, 'need_payment': True}, "预约成功，请完成支付"))
        else:
            return jsonify(success_response({'id': booking_id, 'need_payment': False}, "预约成功，等待审核"))
    except Exception as e:
        return jsonify(error_response(f"预约失败: {str(e)}")), 500


@app.route('/api/course-bookings/<int:booking_id>/approve', methods=['PUT'])
@login_required
def approve_course_booking(booking_id):
    """审核通过课程预约（教练或管理员）"""
    try:
        current_user = request.current_user

        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取预约信息
        cursor.execute("""
            SELECT * FROM teacher_course_bookings WHERE id = %s
        """, (booking_id,))
        booking = cursor.fetchone()

        if not booking:
            cursor.close()
            conn.close()
            return jsonify(error_response("预约记录不存在")), 404

        # 权限检查：只有教练本人或管理员可以审核
        if current_user['id'] != booking['teacher_id'] and current_user.get('user_type') != 'admin':
            cursor.close()
            conn.close()
            return jsonify(error_response("无权限")), 403

        if booking['status'] != 'pending':
            cursor.close()
            conn.close()
            return jsonify(error_response(f"预约状态为{booking['status']}，无法审核")), 400

        # 更新预约状态为审核通过
        cursor.execute("""
            UPDATE teacher_course_bookings SET status = 'approved' WHERE id = %s
        """, (booking_id,))

        conn.commit()

        try:
            Logger.log_audit(
                user_id=current_user['id'],
                user_name=current_user.get('real_name', ''),
                action='审核通过课程预约',
                target_type='teacher_course_bookings',
                target_id=booking_id
            )
        except Exception:
            pass

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "审核通过"))
    except Exception as e:
        return jsonify(error_response(f"审核失败: {str(e)}")), 500


@app.route('/api/course-bookings/<int:booking_id>/reject', methods=['PUT'])
@login_required
def reject_course_booking(booking_id):
    """拒绝课程预约（教练或管理员）- 自动退款"""
    try:
        current_user = request.current_user
        data = request.json
        reject_reason = data.get('reject_reason', '教练拒绝')

        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取预约信息
        cursor.execute("""
            SELECT * FROM teacher_course_bookings WHERE id = %s
        """, (booking_id,))
        booking = cursor.fetchone()

        if not booking:
            cursor.close()
            conn.close()
            return jsonify(error_response("预约记录不存在")), 404

        # 权限检查
        if current_user['id'] != booking['teacher_id'] and current_user.get('user_type') != 'admin':
            cursor.close()
            conn.close()
            return jsonify(error_response("无权限")), 403

        if booking['status'] != 'pending':
            cursor.close()
            conn.close()
            return jsonify(error_response(f"预约状态为{booking['status']}，无法拒绝")), 400

        # 更新预约状态为拒绝
        cursor.execute("""
            UPDATE teacher_course_bookings
            SET status = 'rejected', reject_reason = %s
            WHERE id = %s
        """, (reject_reason, booking_id))

        # 如果是私教，释放时间段
        if booking['schedule_id']:
            cursor.execute("""
                UPDATE teacher_course_schedules
                SET is_booked = 0, booking_id = NULL
                WHERE id = %s
            """, (booking['schedule_id'],))

        # 如果是团课，减少当前参与人数
        cursor.execute("""
            SELECT course_type FROM teacher_courses WHERE id = %s
        """, (booking['course_id'],))
        course = cursor.fetchone()
        if course and course['course_type'] == 'group':
            cursor.execute("""
                UPDATE teacher_courses
                SET current_participants = GREATEST(0, current_participants - 1)
                WHERE id = %s
            """, (booking['course_id'],))

        # 如果使用了会员卡，退还次数或释放期限卡
        if booking['use_membership'] and booking['membership_card_id']:
            cursor.execute("""
                SELECT card_type FROM user_membership_cards WHERE id = %s
            """, (booking['membership_card_id'],))
            user_card = cursor.fetchone()

            if user_card:
                if user_card['card_type'] == 'times':
                    # 退还次数
                    cursor.execute("""
                        UPDATE user_membership_cards
                        SET remaining_times = remaining_times + 1, status = 'active'
                        WHERE id = %s
                    """, (booking['membership_card_id'],))
                # 期限卡不需要特殊处理（只是没有消耗）

        # 如果实际支付了钱，自动退款并扣除积分
        if booking['payment_amount'] > 0:
            # TODO: 调用微信退款API
            # 这里应该调用微信支付的退款接口

            # 扣除积分
            deduct_user_points(booking['user_id'], float(booking['payment_amount']), 'refund',
                               booking_id=booking_id,
                               description=f'课程预约被拒绝，退款¥{booking["payment_amount"]}')

        conn.commit()

        try:
            Logger.log_audit(
                user_id=current_user['id'],
                user_name=current_user.get('real_name', ''),
                action='拒绝课程预约',
                target_type='teacher_course_bookings',
                target_id=booking_id
            )
        except Exception:
            pass

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "已拒绝并自动退款"))
    except Exception as e:
        return jsonify(error_response(f"操作失败: {str(e)}")), 500


@app.route('/api/course-bookings', methods=['GET'])
@login_required
def get_course_bookings():
    """获取课程预约列表"""
    try:
        current_user = request.current_user

        # 查询参数
        status = request.args.get('status', '')  # pending, approved, rejected, cancelled
        role = request.args.get('role', 'user')  # user 或 teacher
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        offset = (page - 1) * limit

        conn = get_db_connection()
        cursor = conn.cursor()

        # 构建查询
        query = """
            SELECT
                tcb.*,
                tc.title as course_title,
                tc.course_type,
                tc.cover_image as course_image,
                tcs.schedule_date,
                tcs.start_time,
                tcs.end_time,
                u.real_name as user_name,
                u.phone as user_phone,
                t.real_name as teacher_name
            FROM teacher_course_bookings tcb
            LEFT JOIN teacher_courses tc ON tcb.course_id = tc.id
            LEFT JOIN teacher_course_schedules tcs ON tcb.schedule_id = tcs.id
            LEFT JOIN users u ON tcb.user_id = u.id
            LEFT JOIN users t ON tcb.teacher_id = t.id
            WHERE 1=1
        """
        params = []

        # 根据角色过滤
        if role == 'admin' and current_user.get('role') == 'admin':
            pass  # 管理员（含教练）可查看所有预约
        elif role == 'teacher':
            # 教练查看自己的预约
            query += " AND tcb.teacher_id = %s"
            params.append(current_user['id'])
        else:
            # 用户查看自己的预约
            query += " AND tcb.user_id = %s"
            params.append(current_user['id'])

        # 根据状态过滤
        if status:
            query += " AND tcb.status = %s"
            params.append(status)

        # 分页
        query += " ORDER BY tcb.created_at DESC LIMIT %s OFFSET %s"
        params.extend([limit, offset])

        cursor.execute(query, params)
        bookings = cursor.fetchall()

        # 格式化不可序列化的字段
        from decimal import Decimal
        from datetime import timedelta, datetime as dt_type
        for b in bookings:
            for key, val in list(b.items()):
                if isinstance(val, timedelta):
                    total_secs = int(val.total_seconds())
                    b[key] = f"{total_secs // 3600:02d}:{(total_secs % 3600) // 60:02d}"
                elif isinstance(val, Decimal):
                    b[key] = float(val)
                elif isinstance(val, dt_type):
                    b[key] = val.strftime('%Y-%m-%d %H:%M:%S')

        # 获取总数
        count_where = " WHERE 1=1"
        count_params = []
        if role == 'admin' and current_user.get('role') == 'admin':
            pass
        elif role == 'teacher':
            count_where += " AND tcb.teacher_id = %s"
            count_params.append(current_user['id'])
        else:
            count_where += " AND tcb.user_id = %s"
            count_params.append(current_user['id'])
        if status:
            count_where += " AND tcb.status = %s"
            count_params.append(status)

        cursor.execute("SELECT COUNT(*) as total FROM teacher_course_bookings tcb" + count_where, count_params)
        total = cursor.fetchone()['total']

        cursor.close()
        conn.close()

        return jsonify(success_response({
            'list': bookings,
            'total': total,
            'page': page,
            'limit': limit
        }, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500

@app.route('/api/course-bookings/<int:booking_id>/cancel', methods=['PUT'])
@login_required
def cancel_course_booking(booking_id):
    """用户取消课程预约（仅限未支付）"""
    try:
        current_user = request.current_user


        conn = get_db_connection()
        cursor = conn.cursor()


        cursor.execute("""
            SELECT * FROM teacher_course_bookings WHERE id = %s AND user_id = %s
        """, (booking_id, current_user['id']))
        booking = cursor.fetchone()


        if not booking:
            cursor.close()
            conn.close()
            return jsonify(error_response("预约记录不存在")), 404


        if booking['status'] in ['cancelled', 'completed']:
            cursor.close()
            conn.close()
            return jsonify(error_response("该预约已取消或已完成")), 400


        # ★ 已付款不能直接取消，需走退款流程
        if booking['payment_status'] == 'paid':
            cursor.close()
            conn.close()
            return jsonify(error_response("已支付的预约请通过退款流程取消")), 400


        # 更新预约状态为取消
        cursor.execute("""
            UPDATE teacher_course_bookings SET status = 'cancelled' WHERE id = %s
        """, (booking_id,))


        # 如果是私教，释放时间段
        if booking['schedule_id']:
            cursor.execute("""
                UPDATE teacher_course_schedules
                SET is_booked = 0, booking_id = NULL
                WHERE id = %s
            """, (booking['schedule_id'],))


        # 如果是团课，减少当前参与人数
        cursor.execute("SELECT course_type FROM teacher_courses WHERE id = %s", (booking['course_id'],))
        course = cursor.fetchone()
        if course and course['course_type'] == 'group':
            cursor.execute("""
                UPDATE teacher_courses
                SET current_participants = GREATEST(0, current_participants - 1)
                WHERE id = %s
            """, (booking['course_id'],))


        # 如果使用了会员卡，退还次数
        if booking['use_membership'] and booking['membership_card_id']:
            cursor.execute("SELECT card_type FROM user_membership_cards WHERE id = %s", (booking['membership_card_id'],))
            user_card = cursor.fetchone()
            if user_card and user_card['card_type'] == 'times':
                cursor.execute("""
                    UPDATE user_membership_cards
                    SET remaining_times = remaining_times + 1, status = 'active'
                    WHERE id = %s
                """, (booking['membership_card_id'],))


        conn.commit()


        try:
            Logger.log_operation(
                user_id=current_user['id'],
                user_name=current_user.get('real_name', '用户'),
                operation_type='取消课程预约',
                operation_module='课程预约',
                operation_desc=f"取消课程预约ID: {booking_id}",
                request_params={'booking_id': booking_id}
            )
        except Exception as e:
            print(f"日志记录失败: {str(e)}")


        cursor.close()
        conn.close()


        return jsonify(success_response(None, "已取消预约"))
    except Exception as e:
        return jsonify(error_response(f"取消失败: {str(e)}")), 500

# ==================== 场馆首页内容管理 ====================
@app.route('/api/venue-content/link-options', methods=['GET'])
@login_required
def get_venue_content_link_options():
    """获取发布类型下拉选项（活动/课程/商品列表）"""
    try:
        link_type = request.args.get('type', '')
        search = request.args.get('search', '').strip()

        conn = get_db_connection()
        cursor = conn.cursor()

        if link_type == 'activity':
            query = "SELECT id, title, cover_images, created_at FROM activities WHERE 1=1"
            params = []
            if search:
                query += " AND title LIKE %s"
                params.append(f"%{search}%")
            query += " ORDER BY created_at DESC LIMIT 100"
        elif link_type == 'course':
            query = "SELECT id, title, cover_image, created_at FROM teacher_courses WHERE 1=1"
            params = []
            if search:
                query += " AND title LIKE %s"
                params.append(f"%{search}%")
            query += " ORDER BY created_at DESC LIMIT 100"
        elif link_type == 'product':
            query = "SELECT id, name as title, image_url, created_at FROM products WHERE 1=1"
            params = []
            if search:
                query += " AND name LIKE %s"
                params.append(f"%{search}%")
            query += " ORDER BY created_at DESC LIMIT 100"
        else:
            cursor.close()
            conn.close()
            return jsonify(success_response([], "请指定类型"))

        cursor.execute(query, params)
        items = cursor.fetchall()

        # 统一图片字段为 image_url
        for item in items:
            if link_type == 'activity':
                cover_images = item.pop('cover_images', None)
                if cover_images:
                    if isinstance(cover_images, str):
                        try:
                            cover_images = json.loads(cover_images)
                        except Exception:
                            cover_images = []
                    if isinstance(cover_images, list) and len(cover_images) > 0:
                        item['image_url'] = cover_images[0]
                    else:
                        item['image_url'] = ''
                else:
                    item['image_url'] = ''
            elif link_type == 'course':
                item['image_url'] = item.pop('cover_image', '') or ''

        cursor.close()
        conn.close()

        return jsonify(success_response(items, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


@app.route('/api/venue-content', methods=['GET'])
def get_venue_content():
    """获取场馆首页内容"""
    try:
        content_type = request.args.get('content_type', '')

        conn = get_db_connection()
        cursor = conn.cursor()

        query = "SELECT * FROM venue_content WHERE is_active = 1"
        params = []

        if content_type:
            query += " AND content_type = %s"
            params.append(content_type)

        query += " ORDER BY sort_order ASC, id DESC"

        cursor.execute(query, params)
        content = cursor.fetchall()

        # 对有 link_type 和 link_id 的条目，如果 title 或 image_url 为空，从关联项填充默认值
        for item in content:
            link_type = item.get('link_type', '')
            link_id = item.get('link_id')
            if not link_type or not link_id:
                continue

            need_title = not item.get('title')
            need_image = not item.get('image_url')
            if not need_title and not need_image:
                continue

            try:
                if link_type == 'activity':
                    cursor.execute("SELECT title, cover_images FROM activities WHERE id = %s", (link_id,))
                    linked = cursor.fetchone()
                    if linked:
                        if need_title and linked.get('title'):
                            item['title'] = linked['title']
                        if need_image and linked.get('cover_images'):
                            images = linked['cover_images']
                            if isinstance(images, str):
                                images = json.loads(images)
                            if isinstance(images, list) and len(images) > 0:
                                item['image_url'] = images[0]
                elif link_type == 'course':
                    cursor.execute("SELECT title, cover_image FROM teacher_courses WHERE id = %s", (link_id,))
                    linked = cursor.fetchone()
                    if linked:
                        if need_title and linked.get('title'):
                            item['title'] = linked['title']
                        if need_image and linked.get('cover_image'):
                            item['image_url'] = linked['cover_image']
                elif link_type == 'product':
                    cursor.execute("SELECT name, image_url FROM products WHERE id = %s", (link_id,))
                    linked = cursor.fetchone()
                    if linked:
                        if need_title and linked.get('name'):
                            item['title'] = linked['name']
                        if need_image and linked.get('image_url'):
                            item['image_url'] = linked['image_url']
            except Exception:
                pass

        cursor.close()
        conn.close()

        return jsonify(success_response(content, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


@app.route('/api/venue-content', methods=['POST'])
@admin_required
def create_venue_content():
    """创建场馆内容（管理员）"""
    try:
        data = request.json

        content_type = data.get('content_type')  # carousel, banner, text, activity
        title = data.get('title', '')
        image_url = data.get('image_url', '')
        content = data.get('content', '')
        link_type = data.get('link_type', '')
        link_id = data.get('link_id')
        sort_order = data.get('sort_order', 0)

        if not content_type:
            return jsonify(error_response("内容类型不能为空")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO venue_content
            (content_type, title, image_url, content, link_type, link_id, sort_order)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (content_type, title, image_url, content, link_type, link_id, sort_order))

        content_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response({'id': content_id}, "创建成功"))
    except Exception as e:
        return jsonify(error_response(f"创建失败: {str(e)}")), 500


@app.route('/api/venue-content/<int:content_id>', methods=['PUT'])
@admin_required
def update_venue_content(content_id):
    """更新场馆内容"""
    try:
        data = request.json

        conn = get_db_connection()
        cursor = conn.cursor()

        # 构建更新语句
        updates = []
        params = []

        for field in ['content_type', 'title', 'image_url', 'content', 'link_type', 'link_id', 'sort_order', 'is_active']:
            if field in data:
                updates.append(f"{field} = %s")
                params.append(data[field])

        if updates:
            params.append(content_id)
            cursor.execute(f"""
                UPDATE venue_content SET {', '.join(updates)} WHERE id = %s
            """, params)
            conn.commit()

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "更新成功"))
    except Exception as e:
        return jsonify(error_response(f"更新失败: {str(e)}")), 500


@app.route('/api/venue-content/<int:content_id>', methods=['DELETE'])
@admin_required
def delete_venue_content(content_id):
    """删除场馆内容"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM venue_content WHERE id = %s", (content_id,))
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(None, "删除成功"))
    except Exception as e:
        return jsonify(error_response(f"删除失败: {str(e)}")), 500


# ==================== 餐饮配送订单 ====================
@app.route('/api/delivery-orders', methods=['POST'])
@login_required
def create_delivery_order():
    """创建餐饮配送订单"""
    try:
        current_user = request.current_user
        data = request.json

        order_id = data.get('order_id')
        teacher_id = data.get('teacher_id')
        delivery_address = data.get('delivery_address', '')
        product_name = data.get('product_name', '')
        remarks = data.get('remarks', '')

        if not order_id:
            return jsonify(error_response("订单ID不能为空")), 400

        # 如果没有指定教练，尝试从最近的预约中获取
        if not teacher_id:
            conn = get_db_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT teacher_id FROM teacher_course_bookings
                WHERE user_id = %s AND status = 'approved'
                ORDER BY created_at DESC LIMIT 1
            """, (current_user['id'],))
            result = cursor.fetchone()
            if result:
                teacher_id = result['teacher_id']
            cursor.close()
            conn.close()

        if not teacher_id:
            return jsonify(error_response("请选择配送教练")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 创建配送订单
        cursor.execute("""
            INSERT INTO delivery_orders
            (order_id, user_id, teacher_id, user_phone, delivery_address, product_name, remarks)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (order_id, current_user['id'], teacher_id, current_user.get('phone', ''),
              delivery_address, product_name, remarks))

        delivery_id = cursor.lastrowid
        conn.commit()

        # TODO: 发送推送通知给教练
        # send_notification_to_teacher(teacher_id, delivery_id)

        cursor.close()
        conn.close()

        return jsonify(success_response({'id': delivery_id}, "配送订单已创建，已通知教练"))
    except Exception as e:
        return jsonify(error_response(f"创建失败: {str(e)}")), 500


@app.route('/api/delivery-orders', methods=['GET'])
@login_required
def get_delivery_orders():
    """获取配送订单列表"""
    try:
        current_user = request.current_user
        role = request.args.get('role', 'user')  # user 或 teacher
        status = request.args.get('status', '')

        conn = get_db_connection()
        cursor = conn.cursor()

        query = """
            SELECT
                do.*,
                u.real_name as user_name,
                t.real_name as teacher_name
            FROM delivery_orders do
            LEFT JOIN users u ON do.user_id = u.id
            LEFT JOIN users t ON do.teacher_id = t.id
            WHERE 1=1
        """
        params = []

        # 根据角色过滤
        if role == 'teacher':
            query += " AND do.teacher_id = %s"
            params.append(current_user['id'])
        else:
            query += " AND do.user_id = %s"
            params.append(current_user['id'])

        # 根据状态过滤
        if status:
            query += " AND do.status = %s"
            params.append(status)

        query += " ORDER BY do.created_at DESC"

        cursor.execute(query, params)
        orders = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify(success_response(orders, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


@app.route('/api/delivery-orders/<int:delivery_id>/confirm', methods=['PUT'])
@login_required
def confirm_delivery_order(delivery_id):
    """教练确认收到配送通知"""
    try:
        current_user = request.current_user

        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查订单
        cursor.execute("""
            SELECT teacher_id, status FROM delivery_orders WHERE id = %s
        """, (delivery_id,))
        order = cursor.fetchone()

        if not order:
            cursor.close()
            conn.close()
            return jsonify(error_response("配送订单不存在")), 404

        # 权限检查
        if current_user['id'] != order['teacher_id']:
            cursor.close()
            conn.close()
            return jsonify(error_response("无权限")), 403

        if order['status'] != 'pending':
            cursor.close()
            conn.close()
            return jsonify(error_response("订单状态不正确")), 400

        # 更新状态为已确认
        cursor.execute("""
            UPDATE delivery_orders
            SET status = 'confirmed', confirmed_at = NOW()
            WHERE id = %s
        """, (delivery_id,))

        conn.commit()

        # TODO: 发送推送通知给用户
        # send_notification_to_user(order['user_id'], '教练已确认收到配送通知')

        cursor.close()
        conn.close()

        return jsonify(success_response(None, "已确认"))
    except Exception as e:
        return jsonify(error_response(f"确认失败: {str(e)}")), 500


# ==================== 用户积分查询 ====================
@app.route('/api/user-points/current', methods=['GET'])
@login_required
def get_current_points():
    """获取当前用户积分"""
    try:
        current_user = request.current_user

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT total_points FROM users WHERE id = %s", (current_user['id'],))
        result = cursor.fetchone()

        cursor.close()
        conn.close()

        return jsonify(success_response({
            'total_points': result['total_points'] if result else 0
        }, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


@app.route('/api/user-points', methods=['GET'])
@login_required
def get_user_points_history():
    """获取用户积分记录"""
    try:
        current_user = request.current_user
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        offset = (page - 1) * limit

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM user_points
            WHERE user_id = %s
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
        """, (current_user['id'], limit, offset))
        records = cursor.fetchall()

        cursor.execute("SELECT COUNT(*) as total FROM user_points WHERE user_id = %s", (current_user['id'],))
        total = cursor.fetchone()['total']

        cursor.close()
        conn.close()

        return jsonify(success_response({
            'list': records,
            'total': total,
            'page': page,
            'limit': limit
        }, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


# ==================== 用户类型管理（设置教练）====================
@app.route('/api/users/<int:user_id>/set-type', methods=['PUT'])
@admin_required
def set_user_type(user_id):
    """设置用户类型（管理员），同时自动同步权限"""
    try:
        data = request.json
        user_type = data.get('user_type')

        if user_type not in ['user', 'teacher', 'admin']:
            return jsonify(error_response("用户类型不正确")), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify(error_response("用户不存在")), 404

        # 用户类型和权限自动同步
        if user_type == 'admin':
            new_role = 'admin'
        elif user_type == 'teacher':
            new_role = 'admin'  # 教练获得管理员权限
        else:
            new_role = 'user'

        cursor.execute("""
            UPDATE users SET user_type = %s, role = %s WHERE id = %s
        """, (user_type, new_role, user_id))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(None, f"用户类型已更新为{user_type}"))
    except Exception as e:
        return jsonify(error_response(f"更新失败: {str(e)}")), 500




# ==================== 用户详情增强（历史活动、历史课程、订单记录）====================
@app.route('/api/users/<int:user_id>/history', methods=['GET'])
@admin_required
def get_user_history(user_id):
    """获取用户历史记录（管理员）"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 历史活动
        cursor.execute("""
            SELECT
                ua.id, ua.status, ua.total_amount, ua.registration_date,
                a.title, a.activity_start, a.activity_end
            FROM user_activities ua
            LEFT JOIN activities a ON ua.activity_id = a.id
            WHERE ua.user_id = %s
            ORDER BY ua.registration_date DESC
            LIMIT 50
        """, (user_id,))
        activities = cursor.fetchall()

        # 历史课程
        cursor.execute("""
            SELECT
                tcb.id, tcb.status, tcb.payment_amount, tcb.created_at,
                tc.title, tc.course_type,
                tcs.schedule_date, tcs.start_time
            FROM teacher_course_bookings tcb
            LEFT JOIN teacher_courses tc ON tcb.course_id = tc.id
            LEFT JOIN teacher_course_schedules tcs ON tcb.schedule_id = tcs.id
            WHERE tcb.user_id = %s
            ORDER BY tcb.created_at DESC
            LIMIT 50
        """, (user_id,))
        courses = cursor.fetchall()

        # 订单记录
        cursor.execute("""
            SELECT
                o.id, o.total_amount, o.status, o.created_at,
                p.name as product_name, oi.quantity, oi.unit_price
            FROM orders o
            LEFT JOIN order_items oi ON o.id = oi.order_id
            LEFT JOIN products p ON oi.product_id = p.id
            WHERE o.user_id = %s
            ORDER BY o.created_at DESC
            LIMIT 50
        """, (user_id,))
        orders = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify(success_response({
            'activities': activities,
            'courses': courses,
            'orders': orders
        }, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


# ==================== 站点配置 ====================

@app.route('/api/site-settings/consult-info', methods=['GET'])
def get_consult_info():
    """获取咨询信息"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute("SELECT setting_value FROM site_settings WHERE setting_key = 'consult_info'")
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        value = row['setting_value'] if row else ''
        return jsonify(success_response({'consult_info': value}, '获取成功'))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


@app.route('/api/site-settings/consult-info', methods=['PUT'])
@admin_required
def update_consult_info():
    """更新咨询信息（管理员）"""
    try:
        data = request.get_json()
        consult_info = data.get('consult_info', '')

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO site_settings (setting_key, setting_value)
            VALUES ('consult_info', %s)
            ON DUPLICATE KEY UPDATE setting_value = %s
        """, (consult_info, consult_info))
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success_response(None, '更新成功'))
    except Exception as e:
        return jsonify(error_response(f"更新失败: {str(e)}")), 500

@app.route('/api/payment/get-trade-no', methods=['GET'])
@login_required
def get_trade_no():
    """DEBUG模式：获取订单的 out_trade_no"""
    if not Config.DEBUG:
        return jsonify(error_response("仅调试模式可用")), 403

    order_type = request.args.get('type')
    order_id = request.args.get('order_id')
    user_id = request.current_user['id']

    print(f"\n===== get_trade_no =====")
    print(f"type={order_type}, order_id={order_id}, user_id={user_id}")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        if order_type == 'activity':
            cursor.execute(
                "SELECT out_trade_no FROM user_activities WHERE id = %s AND user_id = %s",
                (order_id, user_id)
            )
        elif order_type == 'course':
            cursor.execute(
                "SELECT out_trade_no FROM teacher_course_bookings WHERE id = %s AND user_id = %s",
                (order_id, user_id)
            )
        elif order_type == 'membership':
            cursor.execute(
                "SELECT out_trade_no FROM user_membership_cards WHERE id = %s AND user_id = %s",
                (order_id, user_id)
            )
        else:
            # product
            cursor.execute(
                "SELECT out_trade_no FROM orders WHERE id = %s AND user_id = %s",
                (order_id, user_id)
            )

        record = cursor.fetchone()
        print(f"查询结果: {record}")

        cursor.close()
        conn.close()

        if not record:
            return jsonify(error_response("订单不存在")), 404

        if not record.get('out_trade_no'):
            return jsonify(error_response("订单号为空，请重新发起支付")), 404

        return jsonify(success_response(
            {'out_trade_no': record['out_trade_no']},
            "获取成功"
        ))

    except Exception as e:
        import traceback
        traceback.print_exc()
        cursor.close()
        conn.close()
        return jsonify(error_response(f"查询失败: {str(e)}")), 500


@app.route('/api/payment/mock-success', methods=['POST'])
def mock_payment_success():
    """仅DEBUG模式：模拟支付成功回调，用于本地测试"""
    if not Config.DEBUG:
        return jsonify(error_response("仅调试模式可用")), 403


    data = request.get_json()
    out_trade_no = data.get('out_trade_no')


    print(f"\n===== mock_payment_success =====")
    print(f"out_trade_no: {out_trade_no}")


    if not out_trade_no:
        return jsonify(error_response("缺少 out_trade_no")), 400


    conn = get_db_connection()
    cursor = conn.cursor()
    conn.begin()


    try:
        mock_transaction_id = f"mock_txn_{int(time.time())}"


        if out_trade_no.startswith('ACT'):
            cursor.execute(
                "SELECT * FROM user_activities WHERE out_trade_no = %s FOR UPDATE",
                (out_trade_no,)
            )
            record = cursor.fetchone()
            if not record:
                conn.rollback()
                return jsonify(error_response("活动订单不存在")), 404
            if record['payment_status'] == 'paid':
                conn.rollback()
                return jsonify(success_response(None, "已支付"))
            cursor.execute("""
                UPDATE user_activities
                SET payment_status = 'paid', transaction_id = %s, paid_at = NOW()
                WHERE id = %s
            """, (mock_transaction_id, record['id']))
            add_user_points(
                record['user_id'],
                float(record['total_amount']),
                'activity_payment',
                description=f'活动报名支付¥{record["total_amount"]}'
            )


        elif out_trade_no.startswith('CRS'):
            cursor.execute(
                "SELECT * FROM teacher_course_bookings WHERE out_trade_no = %s FOR UPDATE",
                (out_trade_no,)
            )
            record = cursor.fetchone()
            if not record:
                conn.rollback()
                return jsonify(error_response("课程订单不存在")), 404
            if record['payment_status'] == 'paid':
                conn.rollback()
                return jsonify(success_response(None, "已支付"))
            cursor.execute("""
                UPDATE teacher_course_bookings
                SET payment_status = 'paid', paid_at = NOW()
                WHERE id = %s
            """, (record['id'],))
            add_user_points(
                record['user_id'],
                float(record['payment_amount']),
                'course_payment',
                description=f'课程报名支付¥{record["payment_amount"]}'
            )


        elif out_trade_no.startswith('MBR'):
            cursor.execute(
                "SELECT * FROM user_membership_cards WHERE out_trade_no = %s FOR UPDATE",
                (out_trade_no,)
            )
            record = cursor.fetchone()
            if not record:
                conn.rollback()
                return jsonify(error_response("会员卡订单不存在")), 404
            if record['payment_status'] == 'paid':
                conn.rollback()
                return jsonify(success_response(None, "已支付"))
            # 次卡支付后直接激活，期限卡等待手动激活
            activated = 1 if record['card_type'] == 'times' else 0
            cursor.execute("""
                UPDATE user_membership_cards
                SET payment_status = 'paid', paid_at = NOW(), activated = %s
                WHERE id = %s
            """, (activated, record['id']))
            add_user_points(
                record['user_id'],
                float(record['purchase_amount']),
                'membership_payment',
                description=f'购买会员卡支付¥{record["purchase_amount"]}'
            )


        else:
            # 商品订单（out_trade_no = order_no，以OD开头）
            cursor.execute(
                "SELECT * FROM orders WHERE out_trade_no = %s FOR UPDATE",
                (out_trade_no,)
            )
            record = cursor.fetchone()
            if not record:
                conn.rollback()
                return jsonify(error_response("商品订单不存在")), 404
            if record['payment_status'] == 'paid':
                conn.rollback()
                return jsonify(success_response(None, "已支付"))
            cursor.execute("""
                UPDATE orders
                SET payment_status = 'paid', status = 'paid',
                    transaction_id = %s, paid_at = NOW()
                WHERE id = %s
            """, (mock_transaction_id, record['id']))
            add_user_points(
                record['user_id'],
                float(record['total_amount']),
                'product_payment',
                description=f'商品购买支付¥{record["total_amount"]}'
            )


            # ===== 新增：支付成功后创建配送单并通知教练 =====
            if record.get('delivery_teacher_id'):
                try:
                    # 查询订单中的商品名称
                    cursor.execute("""
                        SELECT oi.*, p.name FROM order_items oi
                        JOIN products p ON oi.product_id = p.id
                        WHERE oi.order_id = %s
                    """, (record['id'],))
                    order_items = cursor.fetchall()
                    product_names = ', '.join([it['name'] for it in order_items])


                    # 创建配送单
                    cursor.execute("""
                        INSERT INTO delivery_orders
                        (order_id, user_id, teacher_id, user_phone, delivery_address, product_name, status)
                        VALUES (%s, %s, %s, %s, %s, %s, 'pending')
                    """, (record['id'], record['user_id'], record['delivery_teacher_id'],
                          record.get('receiver_phone', ''), record.get('shipping_address', ''), product_names))


                    # 查询买家姓名
                    cursor.execute("SELECT real_name FROM users WHERE id = %s", (record['user_id'],))
                    buyer = cursor.fetchone()
                    buyer_name = buyer['real_name'] if buyer else '用户'


                    # 发送站内通知给教练
                    Notifier.send_delivery_notification(
                        teacher_id=record['delivery_teacher_id'],
                        user_name=buyer_name,
                        product_name=product_names,
                        order_no=record.get('order_no', out_trade_no)
                    )
                except Exception as de:
                    print(f"创建配送单/通知失败（不影响支付）: {str(de)}")
                    import traceback
                    traceback.print_exc()


        conn.commit()
        cursor.close()
        conn.close()


        print(f"模拟支付成功: {out_trade_no}")
        return jsonify(success_response(None, f"模拟支付成功"))


    except Exception as e:
        conn.rollback()
        import traceback
        traceback.print_exc()
        cursor.close()
        conn.close()
        return jsonify(error_response(f"模拟失败: {str(e)}")), 500


# ==================== 二维码核销 ====================


@app.route('/api/user/qrcode-data', methods=['GET'])
@login_required
def get_user_qrcode_data():
    """生成用户二维码数据和图片"""
    import hmac, hashlib
    try:
        user = request.current_user
        ts = str(int(time.time()))
        msg = f"{user['id']}:{ts}"
        sig = hmac.new(Config.QRCODE_SECRET.encode(), msg.encode(), hashlib.sha256).hexdigest()[:24]
        qr_content = f"{user['id']}:{ts}:{sig}"
        expires_at = int(ts) + Config.QRCODE_EXPIRE_SECONDS


        # 生成二维码图片
        import qrcode as qr_lib
        img = qr_lib.make(qr_content)
        qr_dir = os.path.join(Config.UPLOAD_FOLDER, 'qrcodes')
        os.makedirs(qr_dir, exist_ok=True)
        filename = f"qr_{user['id']}_{ts}.png"
        filepath = os.path.join(qr_dir, filename)
        img.save(filepath)
        image_url = f"{Config.SERVER_HOST}/uploads/qrcodes/{filename}"


        return jsonify(success_response({
            'qr_content': qr_content,
            'qr_image_url': image_url,
            'expires_at': expires_at,
            'user_name': user.get('real_name', ''),
            'avatar_url': user.get('avatar_url', '')
        }, "获取成功"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"生成二维码失败: {str(e)}")), 500




@app.route('/api/membership-cards/verify-qrcode', methods=['POST'])
@login_required
def verify_qrcode():
    """验证用户二维码，返回用户信息、可核销次卡、生效中的全部会员卡"""
    import hmac, hashlib
    try:
        current_user = request.current_user
        if current_user.get('user_type') not in ('teacher', 'admin') and current_user.get('role') != 'admin':
            return jsonify(error_response("无权限，仅教练或管理员可操作")), 403


        data = request.json
        qr_content = data.get('qr_content', '')
        parts = qr_content.split(':')
        if len(parts) != 3:
            return jsonify(error_response("二维码格式无效")), 400


        target_user_id, ts, sig = parts[0], parts[1], parts[2]


        msg = f"{target_user_id}:{ts}"
        expected_sig = hmac.new(Config.QRCODE_SECRET.encode(), msg.encode(), hashlib.sha256).hexdigest()[:24]
        if sig != expected_sig:
            return jsonify(error_response("二维码签名无效")), 400


        if int(time.time()) - int(ts) > Config.QRCODE_EXPIRE_SECONDS:
            return jsonify(error_response("二维码已过期，请让用户刷新")), 400


        target_user_id = int(target_user_id)


        conn = get_db_connection()
        cursor = conn.cursor()


        cursor.execute("SELECT id, real_name, phone, avatar_url, created_at FROM users WHERE id = %s",
                       (target_user_id,))
        target_user = cursor.fetchone()
        if not target_user:
            cursor.close()
            conn.close()
            return jsonify(error_response("用户不存在")), 404


        created_at = target_user.get('created_at')
        if created_at and hasattr(created_at, 'date'):
            cumulative_days = (datetime.now().date() - created_at.date()).days
        else:
            cumulative_days = 0


        cursor.execute("""
            SELECT COUNT(*) as cnt FROM user_activities
            WHERE user_id = %s AND status IN ('approved','completed')
        """, (target_user_id,))
        act_count = cursor.fetchone()['cnt']


        cursor.execute("""
            SELECT COUNT(*) as cnt FROM teacher_course_bookings
            WHERE user_id = %s AND status IN ('approved','completed')
        """, (target_user_id,))
        course_count = cursor.fetchone()['cnt']


        cursor.execute("""
            SELECT COUNT(*) as cnt FROM membership_consume_logs WHERE user_id = %s
        """, (target_user_id,))
        consume_count = cursor.fetchone()['cnt']


        cumulative_training_count = act_count + course_count + consume_count


        # 可核销的次卡
        cursor.execute("""
            SELECT umc.*, mc.card_name, mc.card_image
            FROM user_membership_cards umc
            LEFT JOIN membership_cards mc ON umc.card_id = mc.id
            WHERE umc.user_id = %s AND umc.card_type = 'times' AND umc.status = 'active'
              AND (umc.payment_status = 'paid' OR umc.gift_from_admin = 1)
            ORDER BY umc.purchase_date DESC
        """, (target_user_id,))
        times_cards = cursor.fetchall()


        from decimal import Decimal
        for card in times_cards:
            for key, val in list(card.items()):
                if isinstance(val, Decimal):
                    card[key] = float(val)
                elif hasattr(val, 'strftime'):
                    card[key] = val.strftime('%Y-%m-%d %H:%M:%S')


            cursor.execute("""
                SELECT * FROM membership_consume_logs
                WHERE user_membership_card_id = %s ORDER BY created_at DESC LIMIT 5
            """, (card['id'],))
            logs = cursor.fetchall()
            for lg in logs:
                if hasattr(lg.get('created_at'), 'strftime'):
                    lg['created_at'] = lg['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            card['recent_logs'] = logs


        # 生效中的全部会员卡
        cursor.execute("""
            SELECT umc.*, mc.card_name, mc.card_image, mc.valid_days as mc_valid_days
            FROM user_membership_cards umc
            LEFT JOIN membership_cards mc ON umc.card_id = mc.id
            WHERE umc.user_id = %s AND umc.status = 'active'
              AND (umc.payment_status = 'paid' OR umc.gift_from_admin = 1)
              AND (
                (umc.card_type = 'times' AND umc.remaining_times > 0)
                OR (umc.card_type = 'period' AND umc.activated = 1 AND umc.end_date >= CURDATE())
              )
            ORDER BY umc.purchase_date DESC
        """, (target_user_id,))
        active_cards = cursor.fetchall()


        from datetime import date as date_type
        today = date_type.today()


        # ★ 关键修复：先计算 remaining_days，再序列化日期
        for card in active_cards:
            if card['card_type'] == 'period':
                # 在序列化之前，end_date 还是原始的 date/datetime 对象
                raw_end_date = card.get('end_date')
                if raw_end_date:
                    if hasattr(raw_end_date, 'date'):
                        # datetime 对象
                        end_date_obj = raw_end_date.date()
                    elif isinstance(raw_end_date, date_type):
                        # date 对象
                        end_date_obj = raw_end_date
                    else:
                        end_date_obj = None


                    if end_date_obj:
                        card['remaining_days'] = (end_date_obj - today).days
                    else:
                        card['remaining_days'] = 0
                else:
                    card['remaining_days'] = 0


                card['total_days'] = card.get('mc_valid_days') or card.get('valid_days') or 0
            else:
                card['remaining_days'] = None
                card['total_days'] = None


            # 序列化所有日期和Decimal字段
            for key, val in list(card.items()):
                if isinstance(val, Decimal):
                    card[key] = float(val)
                elif isinstance(val, date_type) and not isinstance(val, datetime):
                    card[key] = val.strftime('%Y-%m-%d')
                elif hasattr(val, 'strftime'):
                    card[key] = val.strftime('%Y-%m-%d %H:%M:%S')


        cursor.close()
        conn.close()


        return jsonify(success_response({
            'user_info': {
                'id': target_user['id'],
                'real_name': target_user['real_name'],
                'phone': target_user['phone'],
                'avatar_url': target_user['avatar_url'],
                'cumulative_days': cumulative_days,
                'cumulative_training_count': cumulative_training_count
            },
            'times_cards': times_cards,
            'active_cards': active_cards
        }, "验证成功"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"验证失败: {str(e)}")), 500






@app.route('/api/user-membership-cards/<int:card_id>/consume', methods=['POST'])
@login_required
def consume_membership_card(card_id):
    """核销会员卡一次（教练或管理员操作）"""
    try:
        current_user = request.current_user
        if current_user.get('user_type') not in ('teacher', 'admin') and current_user.get('role') != 'admin':
            return jsonify(error_response("无权限")), 403


        conn = get_db_connection()
        cursor = conn.cursor()
        conn.begin()


        cursor.execute("""
            SELECT * FROM user_membership_cards WHERE id = %s FOR UPDATE
        """, (card_id,))
        card = cursor.fetchone()


        if not card:
            conn.rollback()
            cursor.close()
            conn.close()
            return jsonify(error_response("会员卡不存在")), 404


        if card['card_type'] != 'times':
            conn.rollback()
            cursor.close()
            conn.close()
            return jsonify(error_response("仅次卡支持核销")), 400


        if card['status'] != 'active':
            conn.rollback()
            cursor.close()
            conn.close()
            return jsonify(error_response("会员卡已过期或已用完")), 400


        if (card['remaining_times'] or 0) <= 0:
            conn.rollback()
            cursor.close()
            conn.close()
            return jsonify(error_response("剩余次数为0，无法核销")), 400


        remaining_before = card['remaining_times']
        remaining_after = remaining_before - 1


        # 扣减次数
        cursor.execute("""
            UPDATE user_membership_cards SET remaining_times = %s WHERE id = %s
        """, (remaining_after, card_id))


        # 如果用完，更新状态
        if remaining_after <= 0:
            cursor.execute("""
                UPDATE user_membership_cards SET status = 'used_up' WHERE id = %s
            """, (card_id,))


        # 插入核销日志
        cursor.execute("""
            INSERT INTO membership_consume_logs
            (user_membership_card_id, user_id, operator_id, operator_name,
             consume_type, remaining_before, remaining_after)
            VALUES (%s, %s, %s, %s, 'scan', %s, %s)
        """, (card_id, card['user_id'], current_user['id'],
              current_user.get('real_name', ''), remaining_before, remaining_after))


        conn.commit()
        cursor.close()
        conn.close()


        return jsonify(success_response({
            'remaining_times': remaining_after,
            'status': 'used_up' if remaining_after <= 0 else 'active'
        }, f"核销成功，剩余{remaining_after}次"))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error_response(f"核销失败: {str(e)}")), 500




@app.route('/api/user-membership-cards/<int:card_id>/consume-logs', methods=['GET'])
@login_required
def get_consume_logs(card_id):
    """获取会员卡核销历史"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()


        cursor.execute("""
            SELECT * FROM membership_consume_logs
            WHERE user_membership_card_id = %s
            ORDER BY created_at DESC
        """, (card_id,))
        logs = cursor.fetchall()


        for lg in logs:
            if hasattr(lg.get('created_at'), 'strftime'):
                lg['created_at'] = lg['created_at'].strftime('%Y-%m-%d %H:%M:%S')


        cursor.close()
        conn.close()


        return jsonify(success_response(logs, "获取成功"))
    except Exception as e:
        return jsonify(error_response(f"获取失败: {str(e)}")), 500


if __name__ == '__main__':
    app.run(debug=Config.DEBUG, host='0.0.0.0', port=Config.SERVER_PORT)